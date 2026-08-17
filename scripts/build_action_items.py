#!/usr/bin/env python3
"""
Build docs/Action_Items_QA.xlsx — daftar perbaikan hasil audit repo qa-cazh (13 Agustus 2026).

Pakai:
    python scripts/build_action_items.py

Sheet 1 "Repo Ini"  = yang harus kita perbaiki di stagingv3-automation.
Sheet 2 "qa-cazh"   = temuan di repo tim lain (kalau mau disetor ke mereka).

Sumber narasi lengkap: docs/REFERENSI_ELEMEN.md
File .xlsx adalah artifact hasil generate — jangan diedit manual, ubah script ini.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = "docs/Action_Items_QA.xlsx"

NAVY = "1F4E78"
SUBTLE = "595959"
PRIO_FILL = {"Tinggi": "FFC7CE", "Sedang": "FFEB9C", "Rendah": "C6EFCE"}
EFFORT_FILL = "D9E1F2"
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [
    ("ID", 11), ("Area", 16), ("Lokasi", 40), ("Temuan", 52),
    ("Kenapa penting", 50), ("Perbaikan", 54),
    ("Prioritas", 11), ("Effort", 9), ("Status", 12),
]
CENTER = {1, 7, 8, 9}

# ---------------------------------------------------------------------------
# SHEET 1 — perbaikan di repo ini
# ---------------------------------------------------------------------------
REPO_INI = [
    ["FIX-001", "Anti-flaky",
     "cypress/support/**, cypress/e2e/** (sisa 222 kemunculan) + konstanta SETTLE/DEBOUNCE/SAVE_WAIT di 6 POM",
     "cy.wait(<angka>) semula 262x, kini 222x (40 hilang otomatis lewat FIX-015). Konstanta hardcoded: SETTLE=1200 (JadwalPelajaran, Jurusan, Kalender, Kamar, Tag), DEBOUNCE=800, SAVE_WAIT=1500.",
     "Melanggar aturan repo sendiri di CLAUDE.md. Ini sumber flaky nomor satu, sekaligus bikin durasi run membengkak. Dengan retries: 0, satu wait yang kependekan = spec merah tanpa bug.",
     "Rencana 4 tahap (jangan sekali sapu): (1) TERAPKAN DULU pola yang sudah terbukti di repo ini — spec tingkat & tahun ajaran sudah pakai cy.intercept + cy.wait('@alias'); angkat jadi standar. (2) Sapu wait terbesar dulu: SAVE_WAIT=1500 dan DEBOUNCE=800 → intercept POST simpan dan GET list, alias-kan, tunggu alias. (3) SETTLE=1200 (settle modal) → ganti .should('be.visible') pada elemen dialog, bukan jeda. (4) Sisa wait kecil (300-600ms) dikerjakan sambil lalu tiap kali menyentuh POM-nya. Aturan baru: setiap cy.wait(angka) yang tersisa WAJIB ambil nilainya dari fixture.timeouts.* + komentar alasan, biar sisa utangnya kelihatan.",
     "Tinggi", "L", "Open"],

    ["FIX-002", "Auth",
     "cypress/support/pageobjects/LoginPage.js — loginViaSession()",
     "cy.session dipanggil tanpa opsi validate().",
     "Kalau cookie/token sudah kedaluwarsa, Cypress tetap memakai sesi dari cache. Spec gagal dengan error selector yang menyesatkan, bukan error auth.",
     "SELESAI: validate() ditambahkan — cy.visit(dashboard) lalu cy.url().should('not.include', loginPath). Validasi lewat URL, bukan lewat teks halaman seperti qa-cazh (QC-007) yang bocor.",
     "Tinggi", "S", "Done"],

    ["FIX-003", "Auth",
     "cypress/support/pageobjects/LoginPage.js — loginViaSession()",
     "cy.wait(1000) // wait app stabilize di dalam blok cy.session.",
     "Kasus khusus FIX-001 tapi dampaknya ke SEMUA spec, karena setiap spec lewat loginViaSession.",
     "SELESAI: diganti cy.intercept('POST', '**/api/auth/login').as('loginAPI') + cy.wait('@loginAPI') dengan assertion statusCode 200. Endpoint-nya sudah terverifikasi (dipakai spec tingkat & tahun ajaran) dan kini juga ada di fixture login.json -> api.login.",
     "Tinggi", "S", "Done"],

    ["FIX-004", "Tooling",
     "package.json — 10 script: test:mapel, test:mapel:tambah/edit/hapus/list, test:tag, test:tag:tambah/list/edit/hapus",
     "Pemisah antar perintah pakai '&': \"npm run report:clean && npx cypress run --spec ... & npm run report:build\".",
     "KOREKSI dari catatan awal: di Windows (npm memakai cmd.exe, script-shell tidak di-set) '&' TIDAK menjalankan paralel — urutannya tetap benar. Masalah sebenarnya terbukti lewat uji langsung: '&' MENELAN EXIT CODE. Langkah yang exit 1 tetap menghasilkan exit code 0 untuk keseluruhan script, jadi npm run test:mapel selalu dianggap sukses walau Cypress merah. Fatal begitu dipasang ke CI. Catatan: kalau script-shell diubah ke sh/bash, '&' baru benar-benar jadi background — jadi pola lama rapuh di dua arah.",
     "SELESAI: dipecah jadi cy:<modul> (Cypress saja) + test:<modul> = run-s report:clean cy:<modul> report:build --continue-on-error. Sudah diuji: --continue-on-error tetap menjalankan report:build walau test gagal, TAPI exit code akhirnya tetap 1.",
     "Tinggi", "S", "Done"],

    ["FIX-005", "Debug",
     "cypress/support/e2e.js:1-3",
     "Cypress.on('uncaught:exception', () => false) menelan SEMUA error runtime app tanpa filter apa pun.",
     "Bug JavaScript asli di aplikasi — yang justru harusnya jadi temuan QA — hilang diam-diam dan tidak pernah masuk bugs.csv.",
     "TAHAP 1 SELESAI (perilaku run sengaja TIDAK diubah): ada daftar NOISE (baru 'ResizeObserver loop'); error di luar daftar kini dicatat ke Cypress log + console.warn tapi MASIH di-swallow. TAHAP 2 (belum): setelah kelihatan error apa saja yang nyata muncul di run, pindahkan yang derau ke NOISE lalu ubah return false jadi return true supaya error app asli menggagalkan test.",
     "Sedang", "S", "Partial"],

    ["FIX-006", "Bilingual",
     "seluruh POM & spec",
     "Semua assertion mengandalkan teks Bahasa Indonesia ('Simpan', 'Batal', 'Hapus', 'Tambah'), padahal app punya switcher bahasa ID/EN di halaman login.",
     "Sekali akun atau preferensi browser berpindah ke English, hampir seluruh suite merah tanpa ada bug sungguhan. Ini risiko diam yang baru ketahuan saat run besar.",
     "Pilih satu: (a) kunci bahasa ke ID di awal cy.session, atau (b) pakai regex alternation /Simpan|Save/i. Pasangan kata ID+EN sudah tersedia di cypress/fixtures/app.json blok 'bahasa'.",
     "Sedang", "M", "Open"],

    ["FIX-007", "Coverage",
     "cypress.config.js",
     "Belum ada task Node untuk menangani file hasil download: deleteDownloads, findDownloadedFile, readExcel.",
     "Modul yang punya tombol Export/Download belum bisa diuji sampai isi filenya — cuma bisa dicek tombolnya diklik.",
     "SELESAI: 5 task ditambahkan di cypress.config.js — clearDownloads, findDownload, statDownload, readSheet, readTextFile. Sudah diverifikasi ter-registrasi. findDownload mengabaikan file .crdownload (unduhan Chrome yang belum selesai) dan mengembalikan null supaya bisa di-retry pakai .should('not.be.null') — bukan jeda angka. readSheet butuh 'npm i -D xlsx'; selama belum dipasang, task mengembalikan pesan error yang jelas alih-alih melempar exception yang menyesatkan.",
     "Sedang", "M", "Done"],

    ["FIX-008", "Fixture",
     "cypress/fixtures/login.json -> urls.forgotPassword",
     "Nilai '/forgot-password' belum pernah diverifikasi ke app; qa-cazh juga tidak punya datanya untuk dibandingkan.",
     "Kalau path-nya salah, spec Lupa Password sebenarnya sedang menguji halaman 404 dan tetap 'hijau' untuk assertion tertentu.",
     "Buka manual di browser, konfirmasi path-nya, lalu kunci di fixture.",
     "Rendah", "S", "Open"],

    ["FIX-009", "Fixture",
     "cypress/fixtures/app.json — blok routes.unverified, selectors.unverified, dan teks validasi di docs/REFERENSI_ELEMEN.md bagian 4.3",
     "Sekitar 16 route, 15 selector, dan 25 pesan validasi masih berstatus 'unverified' (baru dari qa-cazh).",
     "Kalau dipakai apa adanya, kita mewarisi kesalahan tim lain — persis seperti konflik route Legalitas dan Waktu Perizinan yang sudah terbukti salah di repo mereka.",
     "Verifikasi bertahap saat modulnya digarap, lewat checkpoint element analysis biasa. Yang terbukti benar naikkan ke blok 'verified'; yang salah HAPUS, jangan dibiarkan jadi jebakan.",
     "Sedang", "M", "Open"],

    ["FIX-010", "Navigasi",
     "seluruh POM",
     "Semua modul masuk lewat cy.visit(ROUTE) langsung. Tidak ada satu pun test yang menguji navigasi lewat sidebar.",
     "Sidebar rusak, menu hilang, atau salah tautan tidak akan pernah ketahuan oleh suite ini.",
     "Tambah minimal 1 TC navigasi sidebar per grup modul. Selectornya kini sudah diketahui: [data-slot=\"accordion-menu-item\"] dan [data-slot=\"accordion-menu-title\"] (status masih unverified).",
     "Rendah", "M", "Open"],

    ["FIX-011", "Arsitektur POM",
     "— (berlaku untuk modul yang akan datang)",
     "Ada modul yang TIDAK punya halaman sendiri: item sidebar-nya langsung membuka dialog di atas halaman aktif. Terkonfirmasi untuk Legalitas Bukti Bayar dan Pengaturan Perizinan.",
     "POM bergaya konstanta ROUTE tidak cocok untuk pola ini. Kalau dipaksa, hasilnya seperti error qa-cazh: PermissionTimePage mereka cy.visit ke halaman violation-type lalu klik sidebar.",
     "SELESAI lewat FIX-018: base class SidebarModalPage dibuat dengan anchorRoute + sidebarPath (tanpa konstanta ROUTE), dipakai LegalitasBuktiBayarPage & WaktuPerizinanPage.",
     "Rendah", "S", "Done"],

    ["FIX-012", "Data uji",
     "cypress/fixtures/app.json -> accounts.weakPin",
     "Akun cazhv3@pgl.my.id memicu banner + popup modal 'Perkuat Keamanan PIN Anda' di dashboard.",
     "Kalau akun ini dipakai untuk session global, dialognya menutupi UI dan membuat spec modul lain merah karena elemen tertutup — bukan karena bug.",
     "Pakai hanya di spec khusus PIN. Jangan dijadikan akun default, dan beri catatan di fixture (sudah ditulis).",
     "Rendah", "S", "Open"],

    ["FIX-013", "Bug kode",
     "cypress/support/pageobjects/KalenderPage.js:133 & 477",
     "elements.successToast dideklarasikan sebagai ()=>, tapi dipanggil successToast({ timeout: 8000 }) di moveKalenderInstansi(). Argumennya dibuang diam-diam.",
     "Kelas bug yang sama dengan QC-001 di repo qa-cazh, tapi ini di repo KITA. Kebetulan belum terasa karena nilainya sama dengan default; begitu ada yang mengirim timeout berbeda, diamnya baru menggigit.",
     "SELESAI: getter jadi (opts = {}) => cy.get(sel, { timeout: 8000, ...opts }) — default tetap, argumen kini benar-benar dipakai.",
     "Sedang", "S", "Done"],

    ["FIX-014", "Anti-flaky",
     "cypress/e2e/.../pengaturan mapel/listmapel.cy.js:138, 156, 235",
     "Tiga blok memakai .each(($row) => { cy.wrap($row)... }) — anti-pattern yang dilarang eksplisit di CLAUDE.md.",
     "cy.wrap di dalam .each() menahan referensi elemen lama; kalau tabel re-render di tengah iterasi, Cypress gagal dengan 'detached from the DOM'. Persis jenis kegagalan yang bikin run merah tanpa ada bug app.",
     "SELESAI: diubah ke .should(($rows) => { $rows.each(...) }) dengan Cypress.$ sinkron, plus pesan assertion per baris ('baris N kolom instansi') supaya kalau gagal langsung ketahuan baris mana.",
     "Tinggi", "S", "Done"],

    ["FIX-015", "Auth",
     "15 spec: kelas (4 + cleanupkelas), tahunajaran (4), tingkat (5), cleanup-tahunajaran",
     "Ada TIGA implementasi login di repo ini: LoginPage.loginViaSession (24 spec, id session-<email>), blok cy.session inline id 'admin-cazh-session' (14 spec, dua varian berbeda), dan 'admin-cleanup' (1 spec). Tidak satu pun blok inline punya validate().",
     "Akun yang sama login berkali-kali dalam satu run penuh, dan 15 spec tidak ikut menikmati perbaikan FIX-002/FIX-003. Sesi mati di spec-spec itu tetap dipulihkan dari cache lalu gagal dengan error selector yang menyesatkan.",
     "SELESAI: semua blok inline diganti LoginPage.loginViaSession(...). Repo kini punya SATU session id saja dan 41 spec lewat jalur yang sama. Efek samping bagus: 40 cy.wait(<angka>) ikut terhapus. Komentar TODO/konvensi usang juga dibersihkan.",
     "Tinggi", "M", "Done"],

    ["FIX-016", "Arsitektur POM",
     "cypress/support/pageobjects/base/ (helpers.js, CrudListPage.js, README.md)",
     "Tiap POM menyalin ulang pola yang sama: buka select Radix, tangani body scroll-lock, assertion toast, assertion semua-baris, cek persistence. TagPage.js ~570 baris, sekitar 60% berulang di POM lain.",
     "Duplikasi berarti bug ikut tersalin, dan perbaikan harus dilakukan N kali. Ini juga yang bikin modul baru mahal: tiap modul mulai dari 500+ baris, bukan dari config.",
     "SELESAI (fondasi): base class CrudListPage dibuat dari nol mengikuti konvensi repo ini — bukan salinan qa-cazh. POM turunan cukup deklarasi config (route, teks, fields, columns, api). Base memakai intercept+alias sebagai cara tunggu utama, jadi modul baru otomatis patuh aturan FIX-001 sejak awal. POM LAMA SENGAJA TIDAK di-refactor supaya run yang sudah jalan tidak terganggu.",
     "Sedang", "M", "Done"],

    ["FIX-017", "Coverage",
     "5 POM kerangka + 5 fixture kerangka (Kategori Inventaris, Tipe Pelanggaran, Kategori Pengumuman, Jenis Guru, Jenis Staff)",
     "Kelima modul ini berpola CRUD yang sama dengan modul yang sudah kita kuasai, tapi belum ada kerangkanya sama sekali.",
     "Tanpa kerangka, tiap modul baru mulai dari nol dan gampang menyimpang dari konvensi.",
     "SELESAI (kerangka saja): POM + fixture dibuat untuk 14 modul, nilai hipotesis ditandai (?) di POM dan TODO di fixture. BELUM ADA SPEC — sengaja, karena spec butuh PRD + TC sheet + element analysis dulu. Route Jenis Guru, Jenis Staff, Jenis Tagihan, Pengingat Tagihan, Pengaturan Aplikasi MURNI DUGAAN (qa-cazh tidak pernah cy.visit ke sana), wajib dibuka manual sebelum dipakai.",
     "Sedang", "M", "Partial"],

    ["FIX-018", "Arsitektur POM",
     "cypress/support/pageobjects/base/ — BasePage.js, SidebarModalPage.js, TabbedPage.js",
     "CrudListPage saja tidak cukup: ada modul yang TIDAK punya halaman (dialog dibuka dari sidebar), ada yang berbentuk tab, ada yang read-only.",
     "Kalau semua dipaksa ke satu bentuk, hasilnya seperti kekeliruan qa-cazh: PermissionTimePage mereka cy.visit ke halaman modul LAIN (violation-type) lalu klik sidebar, dan fixture-nya menyimpan dua kandidat route karena tidak pernah dipastikan.",
     "SELESAI: primitif bersama diangkat ke BasePage (Radix scroll-lock, select, switch, toast, sidebar, timeouts); CrudListPage di-refactor mewarisinya. Ditambah SidebarModalPage (anchorRoute + sidebarPath, upload file, time field React Aria per-segmen) dan TabbedPage (switchTab memilih lewat teks DAN memverifikasi data-state=active, bukan asal klik tab pertama seperti qa-cazh).",
     "Sedang", "M", "Done"],

    ["FIX-019", "Scope",
     "Pengaturan Aplikasi (qa-cazh: PGT-11) dan PPDB Pengaturan Web (AGT-7)",
     "qa-cazh menggabung banyak sub-fitur jadi satu kode modul: PGT-11 memuat Halaman Utama + Partner + Banner + SPMB; AGT-7 memuat 7 tab sekaligus.",
     "Empat sampai tujuh layar berbeda di bawah satu kode bikin TC sheet tidak terbaca dan cakupan tidak bisa diukur per fitur. Kalau kita ikut, laporan cakupan kita ikut menyesatkan.",
     "Keputusan dicatat di header POM masing-masing: PECAH jadi modul TC terpisah saat digarap. PengaturanAplikasiPage sengaja dibiarkan sebagai kerangka NAVIGASI saja (belum mendeklarasikan field) supaya tidak terlihat siap padahal isinya menebak 4 form sekaligus.",
     "Rendah", "S", "Open"],
    ["FIX-020", "Tooling",
     "package.json",
     "Script npm per modul cuma ada untuk mapel & tag. 7 modul lain (jurusan, kalender, kamar, kelas, tahunajaran, tingkat, jadwal pelajaran) harus dijalankan dengan mengetik --spec panjang manual.",
     "Menjalankan satu modul jadi merepotkan dan rawan salah ketik path — apalagi path-nya mengandung spasi.",
     "SELESAI: 64 script ditambahkan (cy:<modul>[:aksi] + test:<modul>[:aksi]) untuk seluruh 9 modul, semuanya memakai pola run-s sesuai FIX-004. Diverifikasi tidak ada script yang masih memakai pemisah '&'.",
     "Rendah", "S", "Done"],

    ["FIX-021", "Cleanup",
     "cypress/e2e/stagingv3/cleanup/ — modul tag, mapel, kalender",
     "Tiga modul belum punya spec cleanup. Setelah ditelusuri ketiganya TIDAK setara: hanya Tag yang memakai konvensi rerun-safe QA<ts><seq>.",
     "Tanpa cleanup, data uji menumpuk tiap rerun sampai list berat dan assertion jumlah baris jadi tidak deterministik.",
     "SEBAGIAN: cleanuptag.cy.js dibuat (aman — fixture punya testData.prefix = QA). Mapel & Kalender SENGAJA belum dibuat; lihat FIX-022 & FIX-023, keduanya butuh keputusan dulu.",
     "Sedang", "S", "Partial"],

    ["FIX-022", "Konvensi data",
     "cypress/e2e/.../pengaturan mapel/*.cy.js",
     "Modul Mapel TIDAK memakai konvensi rerun-safe repo ini. CLAUDE.md mewajibkan QA<6-digit-ts><seq>, tapi spec Mapel memakai Math.random().toString(36).slice(2,6) lalu menempelkannya ke nama asli, mis. 'Geografi A3F9'.",
     "Dua akibat: (1) data uji Mapel TIDAK BISA dibersihkan lewat prefix — spec cleanup berbasis prefix QA berisiko menghapus data asli, jadi sengaja tidak dibuat; (2) Math.random tidak dijamin unik lintas run, jadi bentrok nama masih mungkin.",
     "Butuh keputusan: (a) ubah penamaan Mapel ke QA<ts><seq> lalu buat cleanup berbasis prefix — paling bersih, tapi menyentuh 109 TC; ATAU (b) biarkan penamaannya, buat cleanup dengan strategi window 'Dibuat Pada' seperti Jadwal Pelajaran. Jangan dikerjakan sebelum dipilih.",
     "Sedang", "M", "Open"],

    ["FIX-023", "Cleanup",
     "cypress/e2e/.../pengaturan kalender/ + cypress/fixtures/kalender.json",
     "Modul Kalender tidak bisa dibersihkan lewat prefix: Nama Pekan adalah Radix Select (pilih dari daftar), bukan input teks bebas, jadi baris hasil test tidak bisa diberi nama unik. Fixture kalender juga tidak punya blok testData sama sekali.",
     "Cleanup berbasis prefix mustahil di sini. Kalau dipaksakan, yang terhapus bisa data asli.",
     "Kandidat: strategi window 'Dibuat Pada' seperti cypress/fixtures/jadwal_pelajaran.json -> cleanup.windowMinutes. TAPI perlu dipastikan dulu apakah tabel Kalender PUNYA kolom Dibuat Pada — peta kolom saat ini (instansi, awalPekan, namaPekan, header, edit, delete) menunjukkan TIDAK ADA. Kalau memang tidak ada, cleanup otomatis tidak mungkin dan harus manual.",
     "Sedang", "M", "Open"],

]

# ---------------------------------------------------------------------------
# SHEET 2 — temuan di repo qa-cazh
# ---------------------------------------------------------------------------
QA_CAZH = [
    ["QC-001", "Bug kode",
     "ViolationTypePage.js:217, 252, 327; InventoryCategoryPage.js:151",
     "this.elements.formModal({ timeout: 15000 }) — getter-nya () => cy.get(...) dan tidak menerima argumen.",
     "Argumen timeout dibuang diam-diam; yang berlaku tetap timeout default. Memberi rasa aman palsu di 4 titik.",
     "Ubah getter jadi menerima opsi: (opts = {}) => cy.get(sel, opts).",
     "Tinggi", "S", "Open"],

    ["QC-002", "Bug kode",
     "ViolationTypePage.js:34",
     "deleteModal = [role=\"dialog\"]:contains(\"Hapus\") — modal form ikut cocok kalau memuat kata 'Hapus'. deleteConfirmBtn lalu ambil tombol pertama ber-teks /hapus/.",
     "Bisa mengklik tombol yang salah, termasuk tombol hapus di baris tabel, bukan tombol konfirmasi di modal.",
     "Scope ke [data-slot=\"dialog-footer\"] dan bedakan modal lewat judulnya, bukan lewat :contains di seluruh dialog.",
     "Tinggi", "S", "Open"],

    ["QC-003", "Bug kode",
     "AnnouncementCategoryPage.js:31",
     "deleteModal = [role=\"dialog\"] polos — identik dengan formModal.",
     "Tidak ada cara membedakan modal hapus dan modal form. Assertion 'modal hapus muncul' bisa lolos padahal yang terbuka modal tambah.",
     "Bedakan lewat [data-slot=\"dialog-title\"] atau dialog-description.",
     "Tinggi", "S", "Open"],

    ["QC-004", "Bug kode",
     "ViolationTypePage.js:27-28",
     "Min poin = input[type=\"number\"]).first(), Max poin = .last().",
     "Kalau field number-nya cuma satu (atau tiga), keduanya menunjuk elemen yang sama tanpa error apa pun — test lulus dengan data salah.",
     "Scope lewat label: formItem('Poin Minimum').find('input').",
     "Tinggi", "S", "Open"],

    ["QC-005", "Bug kode",
     "InventoryCategoryPage.js:22",
     "modalNamaInput pakai OR-list ('input[data-slot=\"input\"], input:not([type=\"hidden\"]), ...') lalu .last().",
     "Posisional. Begitu urutan field digeser atau ada input pencarian di dalam dialog, yang keambil bukan field nama.",
     "Scope lewat label form, jangan .first()/.last() di atas OR-list.",
     "Sedang", "S", "Open"],

    ["QC-006", "Bug kode",
     "Semua POM — pageSizeDropdown",
     "[role=\"combobox\"] di-filter :contains(\"10\"), :contains(\"50\"), ...",
     "Kena juga combobox lain yang kebetulan memuat angka itu — misalnya tahun ajaran '2025/2026' atau nama instansi berangka.",
     "Scope ke [data-slot=\"data-grid-pagination\"].",
     "Sedang", "S", "Open"],

    ["QC-007", "Bug kode",
     "cypress/support/commands.js:32-35",
     "cy.session validate() hanya memastikan body TIDAK memuat 'Peran Belum Ditetapkan'.",
     "Halaman login juga tidak memuat teks itu. Jadi sesi yang sudah mati tetap dianggap valid — persis kebalikan dari tujuan validate().",
     "Ganti ke cek URL: cy.url().should('not.include', '/auth/login').",
     "Tinggi", "S", "Open"],

    ["QC-008", "Bug kode",
     "cypress/support/e2e.js:4-6",
     "beforeEach(() => cy.login()) dipasang GLOBAL, termasuk untuk spec yang menguji halaman login itu sendiri (authentications/input-email.cy.js).",
     "Spec login dijalankan dalam kondisi sudah login — kemungkinan besar langsung di-redirect, jadi yang diuji bukan alur login.",
     "Pindahkan cy.login() ke beforeEach masing-masing spec modul, jangan di support global.",
     "Tinggi", "S", "Open"],

    ["QC-009", "Data tidak konsisten",
     "InventoryCategoryPage.js:41 vs cypress/fixtures/inventoryCategoryData.json",
     "POM mencari /data inventaris tidak ditemukan/, fixture-nya menulis 'Tidak ada data yang ditemukan'.",
     "Salah satu pasti tidak pernah match. Test empty state jadi tidak bermakna.",
     "Samakan ke teks UI yang sebenarnya, dan baca dari fixture — jangan hardcode regex di POM.",
     "Sedang", "S", "Open"],

    ["QC-010", "Data tidak konsisten",
     "cypress/fixtures/violationTypeData.json",
     "Nilai seperti \"sudah digunakan|sudah ada\" dan \"Batas 100 karakter tercapai|maksimal 100\" — regex alternation disimpan sebagai string biasa.",
     "Kalau dipakai lewat contains() apa adanya, string dengan tanda pipe tidak akan pernah cocok dengan teks UI mana pun.",
     "Simpan sebagai array, atau bangun RegExp secara eksplisit di POM.",
     "Sedang", "S", "Open"],

    ["QC-011", "Route salah",
     "LegalityPage.js:49 vs cypress/fixtures/legalityData.json",
     "POM cy.visit('/setting/invoice/invoice-reminder'), fixture menulis '/setting/invoice/legality'.",
     "Dua nilai berbeda untuk halaman yang sama di repo yang sama. Salah satunya pasti salah.",
     "Tentukan route sebenarnya. Kalau modul ini memang modal-dari-sidebar, hapus konstanta route-nya sekalian.",
     "Tinggi", "S", "Open"],

    ["QC-012", "Route salah",
     "PermissionTimePage.js:42, 62 dan cypress/fixtures/permissionTimeData.json",
     "POM cy.visit('/setting/student-affairs/violation-type') — halaman Tipe Pelanggaran — untuk modul Waktu Perizinan. Fixture menyimpan DUA kandidat: permission-time dan permit-time.",
     "Jelas copy-paste dari ViolationTypePage. Menyimpan dua kandidat berarti route-nya memang tidak pernah dipastikan.",
     "Konfirmasi apakah modul ini punya halaman sendiri atau hanya modal dari sidebar, lalu buang kandidat yang salah.",
     "Tinggi", "S", "Open"],

    ["QC-013", "Kebersihan repo",
     "README.md",
     "Masih ada penanda konflik git '<<<<<<< HEAD' dan '=======' yang belum di-resolve. Contoh perintah di dalamnya juga menunjuk path lama (cypress/e2e/PGT-16_... padahal file ada di cypress/e2e/PGT/).",
     "Dokumen onboarding yang menyesatkan anggota baru sejak menit pertama.",
     "Resolve konflik, samakan contoh perintah dengan struktur folder yang sekarang.",
     "Sedang", "S", "Open"],

    ["QC-014", "Keamanan",
     ".env (ter-commit) + cypress.config.js",
     ".env berisi PASSWORD_EMAIL asli dan ikut masuk repo, PADAHAL cypress.config.js tidak pernah membacanya (tidak ada dotenv).",
     "File mati yang fungsinya cuma membocorkan kredensial. Tidak ada satu pun manfaat yang menyeimbangkan risikonya.",
     "Hapus dari repo, masukkan ke .gitignore, dan rotasi password akun tersebut.",
     "Tinggi", "S", "Open"],

    ["QC-015", "Config",
     "cypress.config.js:7 dan package.json",
     "allowCypressEnv: false bukan opsi Cypress yang valid (tidak berefek). Script 'test' di package.json masih stub bawaan npm: echo \"Error: no test specified\" && exit 1.",
     "Opsi palsu bikin orang mengira ada proteksi env. Script test yang selalu exit 1 bikin repo tidak bisa dipasang ke CI apa adanya.",
     "Hapus opsi yang tidak valid; isi script test dengan perintah cypress run yang sebenarnya.",
     "Rendah", "S", "Open"],

    ["QC-016", "Praktik berbahaya",
     "ViolationTypePage.js:146-160 (saveForm)",
     "Kalau modal masih terbuka dan tidak ada pesan error, kode otomatis mengklik tombol Simpan UNTUK KEDUA KALINYA.",
     "Dua dampak sekaligus: bisa membuat data ganda, dan menutupi bug 'FE diam' yang justru paling layak dilaporkan. Test jadi alat penyembunyi bug.",
     "Hapus klik kedua. Kalau modal tidak menutup dan tidak ada pesan, itu memang harus FAIL.",
     "Tinggi", "S", "Open"],

    ["QC-017", "Praktik berbahaya",
     "ensureDataExists() / ensureInactiveDataExists() di beberapa POM",
     "Membuat data dummy di tengah test kalau tabel kosong, dan mengubah status baris pertama jadi Tidak Aktif.",
     "Mencemari data staging dan membuat assertion jumlah baris tidak deterministik. Test berikutnya jalan di atas kondisi yang diubah test sebelumnya.",
     "Pakai data unik per-run (pola QA<timestamp><seq>) dan spec cleanup terpisah, bukan auto-create di dalam test.",
     "Tinggi", "M", "Open"],

    ["QC-018", "Praktik",
     "seluruh repo (ratusan kemunculan)",
     "cy.wait(<angka>) 500-3000 ms di mana-mana, plus afterEach(() => cy.wait(3000)) di 5 spec akademik.",
     "afterEach saja sudah sekitar 5 spec x ~50 test x 3 detik = kira-kira 12 menit terbuang tiap run penuh, tanpa menambah kepastian apa pun.",
     "Ganti ke cy.intercept + cy.wait('@alias'); hapus afterEach yang cuma menunggu.",
     "Sedang", "L", "Open"],

    ["QC-019", "Praktik",
     "seluruh POM",
     ".click({ force: true }) dipakai sebagai default, bukan sebagai pengecualian.",
     "Menyembunyikan bug nyata: elemen tertutup overlay, tombol disabled, atau modal yang belum siap — semua lolos diam-diam.",
     "Pakai .click() biasa; force hanya untuk kasus yang sudah dipahami dan diberi komentar alasannya.",
     "Sedang", "M", "Open"],

    ["QC-020", "Praktik",
     "seluruh POM",
     "Selector berupa OR-list 4-6 alternatif, mis. '.toast, [role=\"status\"], [class*=\"toast\"], [data-slot=\"toast\"], [data-sonner-toast]'.",
     "Kalau satu selector benar, sisanya cuma derau yang membuat kegagalan tidak bisa dibaca. Selector [data-slot=\"error\"], [data-slot=\"toast\"], [data-slot=\"dialog\"] tidak pernah berdiri sendiri di seluruh repo — indikasi kuat itu tebakan yang tidak pernah dibuktikan.",
     "Ambil HTML asli sekali, pilih SATU selector, buang sisanya.",
     "Sedang", "M", "Open"],

    ["QC-021", "Praktik",
     "cypress/e2e/settings/academic/school_year.cy.js:69 dst",
     "Sel tanggal dipilih dengan button[aria-label*=\"July 17\"].",
     "aria-label memakai nama bulan Inggris walaupun UI berbahasa Indonesia, dan operator *= bisa nyangkut ke tanggal lain. Pecah begitu locale berubah.",
     "Pakai td[data-day=\"YYYY-MM-DD\"] button — ISO dan exact. Ini yang dipakai repo kita.",
     "Sedang", "S", "Open"],

    ["QC-022", "Kebersihan repo",
     "cypress/e2e/settings/academic/jurusan.cy.js:1 dan struktur folder cypress/e2e/Subhan/",
     "describe('The School Year Page') padahal isinya modul Jurusan. Folder spec dinamai per orang (Subhan), bukan per modul.",
     "Judul salah bikin laporan mochawesome menyesatkan. Folder per-orang tidak terbaca oleh siapa pun di luar tim.",
     "Perbaiki judul describe; ubah struktur folder mengikuti modul.",
     "Rendah", "S", "Open"],
]


def build_sheet(ws, title, subtitle, rows):
    ws.title = title
    ws["A1"] = subtitle["title"]
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=NAVY)
    ws["A2"] = subtitle["sub"]
    ws["A2"].font = Font(name="Arial", size=9, color=SUBTLE)

    header_row = 4
    hfill = PatternFill("solid", fgColor=NAVY)
    for ci, (name, width) in enumerate(COLS, start=1):
        c = ws.cell(row=header_row, column=ci, value=name)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[chr(64 + ci)].width = width

    for ri, row in enumerate(rows, start=header_row + 1):
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(
                horizontal="center" if ci in CENTER else "left",
                vertical="top", wrap_text=True)
            c.border = BORDER
        prio = ws.cell(row=ri, column=7)
        fill = PRIO_FILL.get((prio.value or "").strip())
        if fill:
            prio.fill = PatternFill("solid", fgColor=fill)
            prio.font = Font(name="Arial", size=10, bold=True)
        eff = ws.cell(row=ri, column=8)
        eff.fill = PatternFill("solid", fgColor=EFFORT_FILL)
        eff.font = Font(name="Arial", size=10, bold=True)
        st = ws.cell(row=ri, column=9)
        sv=(st.value or "").strip()
        st.font = Font(name="Arial", size=10, bold=True, color=("548235" if sv=="Done" else "BF8F00" if sv=="Partial" else "C00000"))
        ws.row_dimensions[ri].height = 108

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False


wb = Workbook()

build_sheet(
    wb.active, "Repo Ini",
    {"title": "\U0001F527 ACTION ITEMS — stagingv3-automation",
     "sub": "Hasil audit silang dengan repo qa-cazh (app sama: v3.cazh.id) | 13 Agustus 2026 | "
            "Effort: S = di bawah 1 jam, M = setengah hari, L = lebih dari sehari | "
            "Narasi lengkap: docs/REFERENSI_ELEMEN.md"},
    REPO_INI)

build_sheet(
    wb.create_sheet(), "qa-cazh",
    {"title": "\U0001F50D TEMUAN — repo qa-cazh (tim lain)",
     "sub": "Bukan repo kita. Daftar ini untuk (a) alasan kenapa kita tidak meniru polanya, "
            "dan (b) bahan kalau mau disetor ke tim mereka. | 13 Agustus 2026"},
    QA_CAZH)

wb.save(OUT)
print(f"OK -> {OUT} (Repo Ini: {len(REPO_INI)} item, qa-cazh: {len(QA_CAZH)} item)")
