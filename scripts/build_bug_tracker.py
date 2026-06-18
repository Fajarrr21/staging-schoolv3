#!/usr/bin/env python3
"""
Build Bug_Tracker_QA_CARDS.xlsx dari bugs.csv.

Pakai:
    python scripts/build_bug_tracker.py
    python scripts/build_bug_tracker.py bugs.csv Bug_Tracker_QA_CARDS.xlsx

bugs.csv adalah single source of truth (di-commit ke git).
File .xlsx adalah artifact hasil generate — regenerate kapan aja mau setor laporan.
"""
import csv
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = sys.argv[1] if len(sys.argv) > 1 else "bugs.csv"
OUT = sys.argv[2] if len(sys.argv) > 2 else "Bug_Tracker_QA_CARDS.xlsx"

# --- palet warna (match template lama) ---
NAVY = "1F4E78"        # header bar + judul
SUBTLE = "595959"      # subtitle abu
PRIORITY_FILL = "D9E1F2"  # biru muda
SEV_FILL = {
    "High": "FFC7CE",     # merah muda
    "Medium": "FFEB9C",   # kuning
    "Low": "C6EFCE",      # hijau muda
}
STATUS_COLOR = {
    "Open": "C00000",     # merah
    "In Progress": "BF8F00",  # oranye gelap
    "Fixed": "548235",    # hijau
    "Closed": "548235",
    "Won't Fix": "808080",
}
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# kolom: (header, lebar)
COLS = [
    ("Bug ID", 10), ("Module", 18), ("Title", 34), ("Steps to Reproduce", 42),
    ("Expected Result", 42), ("Actual Result", 52), ("Severity", 11),
    ("Priority", 10), ("Reporter", 16), ("Status", 12), ("Date Found", 13),
]
CENTER_COLS = {1, 7, 8, 9, 10, 11}  # Bug ID, Severity, Priority, Reporter, Status, Date

with open(SRC, encoding="utf-8") as f:
    reader = csv.reader(f)
    header = next(reader)
    data = [row for row in reader if row and row[0].strip()]

wb = Workbook()
ws = wb.active
ws.title = "Bug Tracker"

# baris 1: judul
ws["A1"] = "\U0001F41B BUG TRACKER \u2014 QA Automation Findings"
ws["A1"].font = Font(name="Arial", bold=True, size=14, color=NAVY)

# baris 2: subtitle
ws["A2"] = ("Reporter: Fajar Ardiansyah  |  Env: staging-new-school.cazh.id  |  "
            "Tester: QA Automation (Cypress)")
ws["A2"].font = Font(name="Arial", size=9, color=SUBTLE)

# baris 4: header tabel
HEADER_ROW = 4
hfill = PatternFill("solid", fgColor=NAVY)
for ci, (name, width) in enumerate(COLS, start=1):
    c = ws.cell(row=HEADER_ROW, column=ci, value=name)
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill = hfill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    ws.column_dimensions[chr(64 + ci)].width = width

# data
start = HEADER_ROW + 1
for ri, row in enumerate(data, start=start):
    row = (row + [""] * len(COLS))[:len(COLS)]
    for ci, val in enumerate(row, start=1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(
            horizontal="center" if ci in CENTER_COLS else "left",
            vertical="top", wrap_text=True)
        c.border = BORDER
    # Severity (col 7)
    sev = ws.cell(row=ri, column=7)
    fill = SEV_FILL.get((sev.value or "").strip())
    if fill:
        sev.fill = PatternFill("solid", fgColor=fill)
        sev.font = Font(name="Arial", size=10, bold=True)
    # Priority (col 8)
    pr = ws.cell(row=ri, column=8)
    pr.fill = PatternFill("solid", fgColor=PRIORITY_FILL)
    pr.font = Font(name="Arial", size=10, bold=True)
    # Status (col 10)
    sc = STATUS_COLOR.get((ws.cell(row=ri, column=10).value or "").strip())
    if sc:
        ws.cell(row=ri, column=10).font = Font(name="Arial", size=10, bold=True, color=sc)
    ws.row_dimensions[ri].height = 120

ws.freeze_panes = "A5"
ws.sheet_view.showGridLines = False

wb.save(OUT)
print(f"OK -> {OUT} ({len(data)} bugs)")
