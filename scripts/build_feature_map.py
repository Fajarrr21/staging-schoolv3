#!/usr/bin/env python3
"""
Build docs/Peta_Fitur_qa-cazh.xlsx — peta fitur app CAZH v3 yang SUDAH digarap tim lain
(repo qa-cazh, dianalisis 13 Agustus 2026), dipakai untuk memilih modul mana yang kita garap.

Pakai:
    python scripts/build_feature_map.py

Angka TC = jumlah it() yang benar-benar ada di spec per-file (bukan file gabungan,
supaya tidak dobel hitung).
File .xlsx adalah artifact hasil generate — jangan diedit manual, ubah script ini.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = "docs/Peta_Fitur_qa-cazh.xlsx"

NAVY = "1F4E78"
SUBTLE = "595959"
SIAP_FILL = {"A": "C6EFCE", "B": "FFEB9C", "C": "D9D9D9"}
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [
    ("Kode", 10), ("Modul / Fitur", 26), ("Menu sidebar", 30), ("Route", 34),
    ("TC", 6), ("Aset dari qa-cazh", 40), ("Siap", 6),
    ("Status di repo kita", 20), ("Catatan penting", 48),
]
CENTER = {1, 5, 7}

# Siap: A = POM + fixture + judul TC  |  B = judul TC kaya, elemen mentah  |  C = sudah kita punya
ROWS = [
    # --- A: paling siap digarap -------------------------------------------------
    ["PGT-17", "Kategori Inventaris", "PENGATURAN > Inventaris > Kategori Inventaris",
     "/setting/inventory", 41,
     "InventoryCategoryPage.js + inventoryCategoryData.json (placeholder, pesan validasi, empty state)",
     "A", "Belum ada",
     "REKOMENDASI PERTAMA. Pola CRUD-nya kembar dengan Jurusan/Kamar/Tag yang sudah kita kuasai: form 2 field (Instansi + Nama), search, filter instansi, sorting, page size, hapus + konfirmasi. Route paling tidak ambigu di antara modul baru."],

    ["PGT-18", "Tipe Pelanggaran", "PENGATURAN > Kesiswaan > Tipe Pelanggaran",
     "/setting/student-affairs/violation-type", 57,
     "ViolationTypePage.js + violationTypeData.json (6 pesan validasi + aturan range poin)",
     "A", "Belum ada",
     "Paling kaya validasi: Min/Max poin, angka negatif, min>max, min=max, max>999, overlap range antar tipe, duplikat nama, limit 100 karakter. Ada integrasi ke fitur Buat Pelanggaran (/student-affairs/violation/create) — hanya tipe Aktif yang boleh muncul."],

    ["PGT-20", "Kategori Pengumuman", "PENGATURAN > Administrasi > Kategori Pengumuman",
     "/setting/administration/announcement-category", 48,
     "AnnouncementCategoryPage.js + announcementCategoryData.json (8 pesan validasi + 3 toast, teks lengkap)",
     "A", "Belum ada",
     "Satu-satunya yang tambah/edit-nya lewat HALAMAN, bukan modal (.../create). Teks validasinya paling lengkap & spesifik di antara semua fixture mereka. Ada soft delete + penolakan hapus kalau kategori masih dipakai pengumuman aktif."],

    ["DSH-1", "Dashboard", "Dashboard (halaman utama setelah login)",
     "/dashboard", 42,
     "DashboardPage.js (paling detail: 7 metric card, recharts, weak-PIN banner) + dashboardData.json (label ID+EN)",
     "A", "Belum ada",
     "KANDIDAT PILOT PALING AMAN: read-only, tidak bikin/ubah data sama sekali, jadi nol risiko mencemari staging. Bagus buat menguji dulu apakah pola POM kita cocok dipakai di modul non-CRUD (grafik recharts, scroll area, alert)."],

    ["KSW-1", "Progres Kegiatan", "KESISWAAN > Progres Kegiatan",
     "/student-affairs/progress", 87,
     "ProgressActivityPage.js (tipis — cuma 3 selector) + progressActivityData.json",
     "A", "Belum ada",
     "TC paling banyak (87) tapi asetnya paling tipis — POM-nya cuma 110 baris. Efektifnya mendekati kelas B. Jangan dijadikan modul pertama."],

    ["PGT-16", "Legalitas Bukti Bayar", "PENGATURAN > Tagihan > Legalitas Bukti Bayar",
     "KONFLIK: /setting/invoice/legality vs /invoice-reminder", 20,
     "LegalityPage.js + legalityData.json + 5 file gambar/PDF untuk uji upload",
     "A", "Belum ada",
     "Satu-satunya modul dengan UPLOAD FILE (PNG/JPG/JPEG < 2MB, tolak PDF & >2MB) — file ujinya sudah tersedia di fixture mereka, tinggal disalin. TAPI route-nya konflik dan modul ini kemungkinan modal-dari-sidebar. Butuh konfirmasi manual dulu."],

    ["PGT-19", "Waktu Perizinan", "PENGATURAN > Kesiswaan > Perizinan",
     "KONFLIK: permission-time vs permit-time", 13,
     "PermissionTimePage.js + permissionTimeData.json (helper text ON/OFF lengkap)",
     "A", "Belum ada",
     "PALING KECIL (13 TC) jadi menggoda buat pilot, TAPI dua hal berat: route-nya belum pasti, dan input jamnya React Aria [data-slot=datefield] per-segment — pola yang belum pernah kita tangani. Bagus sebagai latihan pola baru, buruk sebagai modul pertama."],

    ["AGT-6", "Log Aktivasi Alumni", "ANGGOTA > Alumni > Log Aktivasi",
     "/member/alumni/activation-log", 33,
     "AlumniLogAktivitasPage.js + alumniLogAktivitasData.json (8 nama kolom + 3 tab)",
     "A", "Belum ada",
     "Berbasis tab (Menunggu / Disetujui / Ditolak) + approval flow. Nama kolom tabelnya sudah tercatat lengkap, jadi TC struktur list bisa langsung disusun."],

    ["AGT-7", "PPDB Pengaturan Web", "ANGGOTA > SPMB/PPDB > Pengaturan Web",
     "/member/admission/setting", 40,
     "PpdbPengaturanWebPage.js + ppdbPengaturanWebData.json (7 tab + sample data)",
     "A", "Belum ada",
     "Form bertab 7 halaman (Profil, Beranda, Jadwal, Status Kustom, Formulir Standar, Formulir Kustom, Program Khusus). Paling besar per-layar; sebaiknya dipecah jadi beberapa modul TC."],

    # --- B: TC kaya, elemen mentah ---------------------------------------------
    ["PGT-14", "Jenis Tagihan", "PENGATURAN > Tagihan > Jenis Tagihan",
     "belum diketahui (mereka navigasi via klik sidebar)", 60,
     "TIDAK ada POM/fixture. TAPI judul it() memuat Skenario + Expected Result lengkap.",
     "B", "Belum ada",
     "Paling banyak logika bisnis: dropdown Pengulangan 8 opsi (Sekaligus Lunas, Setiap Minggu/Bulan/2/3/4/6 Bulan, Setiap Tahun) + Periode + cek duplikat kombinasi Nama+Instansi+Pengulangan+Periode. Judul TC-nya praktis sudah jadi 2 kolom TC sheet kita."],

    ["PGT-11", "Pengaturan Aplikasi", "PENGATURAN > Aplikasi (Halaman Utama, Partner, Banner, SPMB)",
     "belum diketahui", 48,
     "TIDAK ada POM/fixture. Judul it() memuat Skenario + Expected.",
     "B", "Belum ada",
     "Sebenarnya beberapa sub-fitur digabung jadi satu kode (Halaman Utama, Partner, Tambah Banner, SPMB). Kalau digarap, pecah dulu jadi beberapa modul TC terpisah — jangan ikut penggabungan mereka."],

    ["PGT-15", "Pengingat Tagihan", "PENGATURAN > Tagihan > Pengingat Tagihan",
     "belum diketahui", 50,
     "TIDAK ada POM/fixture. Judul it() memuat Skenario + Expected.",
     "B", "Belum ada",
     "Field banyak (Instansi, Jenis Tagihan, Judul, Pesan, Target, Tanggal, Jam, toggle WA) dan menyentuh push notification — sebagian expected-nya sulit diverifikasi lewat UI saja. Perlu dipilah mana yang benar-benar bisa diotomasi."],

    ["PGT-12", "Jenis Guru", "PENGATURAN > Kepegawaian > Jenis Guru",
     "belum diketahui", 43,
     "TIDAK ada POM/fixture. Judul it() memuat Skenario + Expected.",
     "B", "Belum ada",
     "CRUD paling sederhana di kelas B (Instansi + Nama + Status). Konversi ke TC sheet paling cepat karena expected-nya sudah tertulis. Cocok kalau mau target cepat setelah PGT-17."],

    ["PGT-13", "Jenis Staff", "PENGATURAN > Kepegawaian > Jenis Staff",
     "belum diketahui", 44,
     "TIDAK ada POM/fixture. Judul it() memuat Skenario + Expected.",
     "B", "Belum ada",
     "Kembar dengan PGT-12. Kalau PGT-12 digarap, modul ini praktis tinggal salin-sesuaikan — kerjakan berurutan, jangan terpisah jauh."],

    # --- C: sudah kita punya ----------------------------------------------------
    ["PGT-1", "Tahun Ajaran", "PENGATURAN > Akademik > Tahun Ajaran",
     "/setting/academic/school-year", 0,
     "Spec inline mereka (tanpa POM) — 241 it() total untuk 5 modul akademik",
     "C", "SUDAH ADA",
     "Jangan digarap ulang. Nilainya ada di GAP ANALYSIS: bandingkan skenario mereka vs TC sheet kita untuk menemukan kasus yang belum kita cover (sorting kolom, kombinasi filter+search, page size)."],

    ["PGT-2", "Jurusan", "PENGATURAN > Akademik > Jurusan",
     "/setting/academic/major", 0, "Spec inline mereka", "C", "SUDAH ADA",
     "Sama seperti PGT-1 — hanya untuk gap analysis."],

    ["PGT-3", "Tingkat", "PENGATURAN > Akademik > Tingkat",
     "/setting/academic/school-level", 0, "Spec inline mereka", "C", "SUDAH ADA",
     "Sama seperti PGT-1 — hanya untuk gap analysis."],

    ["PGT-4", "Kelas", "PENGATURAN > Akademik > Kelas",
     "/setting/academic/class", 0, "Spec inline mereka", "C", "SUDAH ADA",
     "Sama seperti PGT-1 — hanya untuk gap analysis."],

    ["PGT-5", "Kamar", "PENGATURAN > Akademik > Kamar",
     "/setting/academic/room", 0, "Spec inline mereka", "C", "SUDAH ADA",
     "Sama seperti PGT-1 — hanya untuk gap analysis."],

    ["AUTH", "Login / Localization", "Halaman login",
     "/auth/login", 3,
     "Teks login ID + EN lengkap (sudah masuk cypress/fixtures/login.json blok localization)",
     "C", "SUDAH ADA (36 TC)",
     "Spec login kita jauh lebih lengkap dari mereka. Yang kita ambil cuma data localization-nya — kandidat TC bilingual, tapi teksnya masih berstatus unverified."],
]

wb = Workbook()
ws = wb.active
ws.title = "Peta Fitur"

ws["A1"] = "\U0001F5FA PETA FITUR CAZH v3 — apa saja yang sudah digarap tim lain (qa-cazh)"
ws["A1"].font = Font(name="Arial", bold=True, size=14, color=NAVY)
ws["A2"] = ("Dianalisis 13 Agustus 2026  |  Kolom 'Siap':  A = ada POM + fixture + judul TC (paling siap)  ·  "
            "B = judul TC memuat Skenario+Expected tapi elemen mentah  ·  C = sudah kita punya, jangan diulang  |  "
            "Detail selector & route: cypress/fixtures/app.json + docs/REFERENSI_ELEMEN.md")
ws["A2"].font = Font(name="Arial", size=9, color=SUBTLE)

HEADER_ROW = 4
hfill = PatternFill("solid", fgColor=NAVY)
for ci, (name, width) in enumerate(COLS, start=1):
    c = ws.cell(row=HEADER_ROW, column=ci, value=name)
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill = hfill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    ws.column_dimensions[chr(64 + ci)].width = width

for ri, row in enumerate(ROWS, start=HEADER_ROW + 1):
    for ci, val in enumerate(row, start=1):
        if ci == 5 and val == 0:
            val = "—"
        c = ws.cell(row=ri, column=ci, value=val)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(
            horizontal="center" if ci in CENTER else "left",
            vertical="top", wrap_text=True)
        c.border = BORDER
    siap = ws.cell(row=ri, column=7)
    fill = SIAP_FILL.get((siap.value or "").strip())
    if fill:
        siap.fill = PatternFill("solid", fgColor=fill)
        siap.font = Font(name="Arial", size=10, bold=True)
    st = ws.cell(row=ri, column=8)
    st.font = Font(name="Arial", size=10, bold=True,
                   color=("808080" if "SUDAH" in (st.value or "") else "1F4E78"))
    ws.row_dimensions[ri].height = 96

ws.freeze_panes = "A5"
ws.sheet_view.showGridLines = False

total = sum(r[4] for r in ROWS if r[6] != "C")
wb.save(OUT)
print(f"OK -> {OUT} ({len(ROWS)} modul, {total} TC di modul yang belum kita punya)")
