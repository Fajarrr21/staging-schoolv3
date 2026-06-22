# -*- coding: utf-8 -*-
"""Generate Test Case sheet for Tambah Kalender Akademik (per CLAUDE.md columns)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = ["ID", "Kategori", "Skenario", "Pre-condition", "Langkah",
           "Test Data", "Expected", "Sumber", "Status", "Catatan"]

# Form Tambah Kalender Akademik (dari HTML real app + konfirmasi user):
#   a. Instansi*           (REQUIRED; Select Radix; default=placeholder "Pilih Instansi"; payload key=office)
#   b. Awal pekan dimulai* (REQUIRED; Select Radix; default=placeholder; options: Minggu, Senin; key=start_day)
#   c. Nama Pekan*         (REQUIRED; Select Radix; default=placeholder; options: Ahad, Minggu; key=weekend_name)
#   d. Header (Opsional)   (file upload; accept=.png/.jpg/.jpeg/.webp/.svg; max 2MB; dimensi sugesti 500x300px)
#
# Default form: SEMUA field wajib pakai placeholder (TIDAK ada preselect). Awal pekan & Nama Pekan
# adalah field wajib juga (bukan optional dengan default value seperti dugaan awal dari snapshot HTML
# yang ternyata diambil pada state form yang sudah pernah diisi).
#
# Pesan error duplikasi instansi (saat instansi sudah punya kalender):
#   -> TOAST notification varian error (portal global, bukan child dari modal/dialog-content).
#   -> Text exact: "Pengaturan kalender untuk office ini sudah ada"
#   -> Auto-dismiss; di Cypress: query global (mis. [data-sonner-toast] / [role="status"])
#      dan andalkan retry-ability cy.contains, BUKAN cy.wait fixed.
#
# Catatan PRD vs realita:
#   - PRD: "Awal minggu dimulai" -> realita: "Awal pekan dimulai"
#   - PRD: "Tampilan nama minggu" -> realita: "Nama Pekan"
#   - PRD: Awal pekan & Nama Pekan disebut "opsional" -> realita: keduanya WAJIB (per konfirmasi user)
#   - PRD: format file tidak disebut -> realita accept = .png/.jpg/.jpeg/.webp/.svg; tip UI hanya 3 (jpg/jpeg/png)
#   - PRD: pesan error 2MB tidak disebut -> realita: alert inline di modal,
#     title "Gagal mengunggah file", desc "File melebihi ukuran maksimal 2MB."
ROWS = [
    # ---------------- HAPPY ----------------
    ["TC-KLD-ADD-001", "Happy",
     "Tambah kalender field wajib minimum (3 field wajib terisi, TANPA upload header)",
     "Login admin; halaman List Kalender Akademik terbuka; ada instansi yang BELUM punya kalender",
     "1) Klik \"Tambah Kalender Akademik\"\n2) Pilih Instansi (yang belum punya kalender)\n3) Pilih Awal pekan dimulai = Minggu\n4) Pilih Nama Pekan = Ahad\n5) Biarkan area Header kosong (opsional)\n6) Klik \"Simpan\"",
     "Instansi=<instansi tanpa kalender>; Awal pekan=Minggu; Nama Pekan=Ahad; Header=(kosong)",
     "Toast sukses \"Kalender Akademik berhasil ditambahkan\"; modal tertutup; baris baru muncul di list dengan kolom: Nama instansi, badge \"Minggu\" (Awal pekan), badge \"Ahad\" (Nama pekan), kolom header pakai gambar default",
     "PRD", "Belum dijalankan",
     "Header default = PRD: \"Jika tidak upload foto maka akan menggunakan header default\". Teks toast = Asumsi (PRD tak sebut)"],

    ["TC-KLD-ADD-002", "Happy",
     "Tambah kalender lengkap (3 field wajib + upload foto <2MB)",
     "Login admin; ada instansi tanpa kalender; file `cypress/fixtures/kalender/header-valid.jpg` (<2MB) tersedia",
     "1) Klik \"Tambah Kalender Akademik\"\n2) Pilih Instansi\n3) Pilih Awal pekan dimulai = Senin\n4) Pilih Nama Pekan = Minggu\n5) Upload header-valid.jpg (<2MB)\n6) Tunggu preview file muncul (card dengan nama file)\n7) Klik \"Simpan\"",
     "Instansi=<instansi tanpa kalender>; Awal pekan=Senin; Nama Pekan=Minggu; Header=header-valid.jpg",
     "Toast sukses; modal tertutup; baris baru muncul dengan badge \"Senin\" + badge \"Minggu\" + thumbnail foto header yang diupload (bukan default)",
     "PRD", "Belum dijalankan",
     "Verifikasi persist via reload + cek img src kolom Header bukan default URL"],

    # ---------------- POSITIF ----------------
    ["TC-KLD-ADD-003", "Positif",
     "Form Tambah terbuka dengan default = semua field wajib placeholder (kosong, tanpa preselect)",
     "Login admin; List Kalender Akademik terbuka",
     "1) Klik \"Tambah Kalender Akademik\"\n2) Amati kondisi awal form (jangan ubah apapun)",
     "-",
     "Modal \"Tambah Kalender Akademik\" tampil; Instansi placeholder (\"Pilih Instansi\"); Awal pekan dimulai placeholder; Nama Pekan placeholder; area Header kosong (ikon upload + \"Pilih File\" + tip ukuran/format)",
     "PRD-ambigu", "Belum dijalankan",
     "PRD diam soal default values & status wajib Awal pekan/Nama Pekan. Per konfirmasi user: 3 field wajib, semua placeholder."],

    ["TC-KLD-ADD-004", "Positif",
     "Batal menutup form tanpa input apapun",
     "Modal Tambah Kalender Akademik terbuka",
     "1) Klik \"Tambah Kalender Akademik\"\n2) Klik \"Batal\"",
     "-",
     "Modal tertutup; kembali ke List Kalender Akademik; tidak ada data baru di list",
     "PRD", "Belum dijalankan", ""],

    ["TC-KLD-ADD-005", "Positif",
     "Isi form valid lalu Batal -> data tidak tersimpan",
     "Modal terbuka; ada instansi tanpa kalender; file valid tersedia",
     "1) Pilih Instansi + Awal pekan + Nama Pekan + upload foto\n2) Klik \"Batal\" (BUKAN Simpan)\n3) Reload list; cek instansi target",
     "Instansi=<instansi tanpa kalender>; Awal pekan=Senin; Nama Pekan=Minggu; Header=header-valid.jpg",
     "Modal tertutup; instansi target tetap TIDAK punya kalender (tidak muncul baris baru di list setelah reload)",
     "PRD", "Belum dijalankan", ""],

    ["TC-KLD-ADD-006", "Positif",
     "Klik tombol close X di pojok kanan atas modal -> modal tertutup",
     "Modal terbuka",
     "1) Klik \"Tambah Kalender Akademik\"\n2) Klik ikon X di pojok kanan atas modal",
     "-",
     "Modal tertutup; tidak ada data baru; tidak ada toast",
     "Asumsi", "Belum dijalankan",
     "PRD tak sebut tombol X; behavior konsisten modal Radix standar"],

    ["TC-KLD-ADD-007", "Positif",
     "Dependency Awal pekan vs Nama Pekan independent (kombinasi bebas)",
     "Modal terbuka",
     "1) Pilih Awal pekan = Senin; buka dropdown Nama Pekan; cek semua opsi (Ahad, Minggu) tersedia\n2) Pilih Awal pekan = Minggu; buka dropdown Nama Pekan; cek semua opsi tetap tersedia\n3) Sebaliknya: pilih Nama Pekan = Minggu; cek opsi Awal pekan tetap Minggu+Senin",
     "kombinasi Awal pekan x Nama Pekan",
     "Kedua dropdown independent; tidak ada opsi yang di-disable/auto-update berdasarkan pilihan field lain; user bebas pilih kombinasi apapun",
     "Asumsi", "Belum dijalankan",
     "Per request user: cek dependency. Kalau ada relasi (mis. Awal pekan=Senin auto-set Nama Pekan=Minggu) -> catat sebagai behavior, evaluasi apakah intended"],

    # ---------------- NEGATIF ----------------
    ["TC-KLD-ADD-008", "Negatif",
     "Simpan dengan Instansi kosong -> error \"Instansi wajib diisi\"",
     "Modal terbuka; Instansi belum dipilih",
     "1) Klik \"Tambah Kalender Akademik\"\n2) Isi Awal pekan + Nama Pekan; biarkan Instansi placeholder\n3) Klik \"Simpan\"",
     "Instansi=(kosong); Awal pekan=Minggu; Nama Pekan=Ahad",
     "Pesan error \"Instansi wajib diisi\" muncul di bawah field Instansi (form-message destructive); Simpan tidak bekerja; modal tetap terbuka; tanpa toast sukses",
     "PRD", "Belum dijalankan", ""],

    ["TC-KLD-ADD-009", "Negatif",
     "Simpan dengan Awal pekan dimulai kosong -> error required",
     "Modal terbuka",
     "1) Pilih Instansi + Nama Pekan; biarkan Awal pekan placeholder\n2) Klik \"Simpan\"",
     "Instansi=<valid>; Awal pekan=(kosong); Nama Pekan=Ahad",
     "Pesan error \"Awal pekan dimulai wajib diisi\" (atau pesan setara) muncul di bawah field; Simpan tidak bekerja; modal tetap terbuka",
     "PRD-ambigu", "Belum dijalankan",
     "PRD tidak eksplisit menyebut Awal pekan wajib (PRD bilang opsional). Per konfirmasi user: wajib. Teks pesan exact menunggu trigger di app"],

    ["TC-KLD-ADD-010", "Negatif",
     "Simpan dengan Nama Pekan kosong -> error required",
     "Modal terbuka",
     "1) Pilih Instansi + Awal pekan; biarkan Nama Pekan placeholder\n2) Klik \"Simpan\"",
     "Instansi=<valid>; Awal pekan=Minggu; Nama Pekan=(kosong)",
     "Pesan error \"Nama Pekan wajib diisi\" (atau pesan setara) muncul di bawah field; Simpan tidak bekerja; modal tetap terbuka",
     "PRD-ambigu", "Belum dijalankan",
     "Sama PRD-ambigu seperti Awal pekan: PRD bilang opsional, app wajib"],

    ["TC-KLD-ADD-011", "Negatif",
     "Simpan dengan SEMUA field wajib kosong -> 3 pesan error muncul bersamaan",
     "Modal terbuka",
     "1) Klik \"Tambah Kalender Akademik\"\n2) Langsung klik \"Simpan\" tanpa isi field apa pun",
     "Instansi=(kosong); Awal pekan=(kosong); Nama Pekan=(kosong); Header=(kosong)",
     "3 pesan error required muncul bersamaan (Instansi/Awal pekan/Nama Pekan); Simpan tidak bekerja; modal tetap terbuka; tanpa toast sukses; Header tidak memunculkan error (opsional)",
     "PRD-ambigu", "Belum dijalankan",
     "Cek FE validate semua field sekaligus, bukan stop di field pertama yang kosong"],

    ["TC-KLD-ADD-012", "Negatif",
     "Duplikasi: Tambah untuk instansi yang sudah punya kalender -> toast error global",
     "Login admin; pilih instansi yang sudah ada entry di list (mis. SMA Digital Indonesia)",
     "1) Klik \"Tambah Kalender Akademik\"\n2) Pilih Instansi yang sudah ada kalender-nya\n3) Pilih Awal pekan + Nama Pekan\n4) Klik \"Simpan\"\n5) Amati toast & list",
     "Instansi=SMA Digital Indonesia (atau instansi lain yang sudah ada di list)",
     "Toast error global muncul (di portal, BUKAN inline di modal): \"Pengaturan kalender untuk office ini sudah ada\". Tidak ada toast sukses; modal tetap terbuka (atau tertutup, catat aktual); tidak ada baris duplikat baru di list",
     "PRD-ambigu", "Belum dijalankan",
     "Per konfirmasi user: toast varian error, query global ([data-sonner-toast]/[role=status]), auto-dismiss -> assert harus quick & andalkan retry cy.contains, jangan cy.wait fixed"],

    ["TC-KLD-ADD-013", "Negatif",
     "Upload foto >2MB -> alert error inline di modal, file tidak ter-attach",
     "Modal terbuka; file `cypress/fixtures/kalender/header-oversize.jpg` (>2MB) tersedia",
     "1) Klik \"Tambah Kalender Akademik\"\n2) Klik area upload / \"Pilih File\"\n3) Pilih file >2MB\n4) Amati area upload & pesan",
     "Header=header-oversize.jpg (>2MB)",
     "Alert inline di dalam modal (bukan toast): title \"Gagal mengunggah file\" + desc \"File melebihi ukuran maksimal 2MB.\"; file TIDAK muncul di card preview; user masih bisa pilih file lain atau Simpan tanpa header",
     "PRD", "Belum dijalankan",
     "PRD eksplisit batas 2MB. Pesan exact dari HTML. Alert ada di [data-slot=\"alert\"] INSIDE dialog-content"],

    # ---------------- EDGE ----------------
    ["TC-KLD-ADD-014", "Edge",
     "Upload file non-image (mis. .pdf) -> ditolak browser sesuai accept attribute",
     "Modal terbuka; file `cypress/fixtures/kalender/sample.pdf` tersedia (atau pakai package.json)",
     "1) Klik \"Tambah Kalender Akademik\"\n2) Klik \"Pilih File\"\n3) Coba pilih .pdf (lewat dialog browser native pakai cy.selectFile force)",
     "Header=sample.pdf",
     "File .pdf TIDAK ter-attach: accept attribute restrict ke image (.png/.jpg/.jpeg/.webp/.svg). Card preview tidak muncul. Atau jika lolos accept (cypress force) -> ada validasi backend/FE tambahan",
     "PRD-ambigu", "Belum dijalankan",
     "PRD tak sebut tipe file. Catat behavior aktual; kalau lolos tanpa error -> log BUG"],

    ["TC-KLD-ADD-015", "Edge",
     "Setelah upload sukses, klik ikon trash di card preview -> file ter-detach",
     "Modal terbuka; sudah upload header-valid.jpg dan preview card muncul dengan nama file",
     "1) Upload header-valid.jpg\n2) Tunggu preview card muncul (nama file)\n3) Klik ikon trash (lucide-trash2) di card preview\n4) Amati area upload",
     "Header=header-valid.jpg -> klik trash",
     "Card preview hilang; area upload kembali ke kondisi awal (ikon cloud-upload + \"Pilih File\"); user bisa upload file lain atau Simpan tanpa header",
     "Asumsi", "Belum dijalankan",
     "Behavior reset upload tak disebut PRD. Penting untuk UX"],

    ["TC-KLD-ADD-016", "Edge",
     "Persistence: setelah Tambah sukses, reload halaman -> data tetap muncul",
     "Sudah Tambah kalender baru untuk instansi target (TC-001 atau TC-002)",
     "1) Setelah Tambah sukses & toast muncul\n2) Reload halaman List Kalender Akademik (F5)\n3) Cari baris instansi target di tabel",
     "Instansi yang baru di-Tambah",
     "Baris instansi target ada di list dengan semua kolom benar (Nama instansi, badge Awal pekan, badge Nama pekan, thumbnail header). Bukti data persist ke backend, bukan optimistic UI doang",
     "PRD", "Belum dijalankan",
     "Pola assertPersisted standar untuk semua modul"],
]

# Legend reference data
LEGEND = [
    ["Kolom Sumber", ""],
    ["PRD", "Skenario diturunkan langsung dari PRD Kalender Pendidikan"],
    ["PRD-ambigu", "PRD menyinggung tapi tidak jelas / app berbeda dari PRD -> perlu konfirmasi"],
    ["Asumsi", "Tidak ada di PRD; asumsi berbasis konvensi/HTML/modul lain"],
    ["", ""],
    ["Kolom Kategori", ""],
    ["Happy", "Alur utama sukses sesuai desain"],
    ["Positif", "Variasi valid non-utama"],
    ["Negatif", "Input/kondisi yang harus ditolak sistem"],
    ["Edge", "Batas/kondisi tepi & verifikasi persistence"],
    ["", ""],
    ["Kolom Status", ""],
    ["Belum dijalankan", "Sudah di-generate, menunggu eksekusi Cypress"],
    ["DITUNDA", "Sengaja belum di-run; menunggu prasyarat tersedia"],
    ["", ""],
    ["Catatan", "Tanpa kategori security/injection (sesuai kesepakatan)"],
    ["", "Bug ditandai inline: Status=FAIL + ref BUG-### di kolom Catatan"],
    ["", "Field wajib di app: Instansi*, Awal pekan dimulai*, Nama Pekan* (3 wajib semua, payload key Instansi=office)"],
    ["", "Default form: SEMUA field wajib placeholder (kosong, tanpa preselect)"],
    ["", "Header optional: accept .png/.jpg/.jpeg/.webp/.svg; max 2MB; tip dimensi 500x300px"],
    ["", "Pesan error >2MB (exact, alert INLINE di modal): \"Gagal mengunggah file\" / \"File melebihi ukuran maksimal 2MB.\""],
    ["", "Pesan error duplikat (exact, TOAST GLOBAL di portal, BUKAN inline modal): \"Pengaturan kalender untuk office ini sudah ada\""],
    ["", "Cypress hint utk toast: query global, andalkan retry-ability cy.contains, JANGAN cy.wait fixed (auto-dismiss)"],
]

# ---- styling ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CAT_FILL = {
    "Happy":   PatternFill("solid", fgColor="E2EFDA"),
    "Positif": PatternFill("solid", fgColor="DDEBF7"),
    "Negatif": PatternFill("solid", fgColor="FCE4D6"),
    "Edge":    PatternFill("solid", fgColor="FFF2CC"),
}
SRC_FONT = {
    "PRD":        Font(color="375623"),
    "PRD-ambigu": Font(color="BF8F00", bold=True),
    "Asumsi":     Font(color="C00000", bold=True),
}
STATUS_FILL = {
    "DITUNDA": PatternFill("solid", fgColor="D9D9D9"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPWRAP = Alignment(vertical="top", wrap_text=True)
WIDTHS = [16, 12, 38, 32, 44, 28, 50, 14, 16, 46]

wb = Workbook()
ws = wb.active
ws.title = "Tambah Kalender"

# Title row
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
t = ws.cell(row=1, column=1, value="Test Case - Tambah Kalender Akademik (CARDS School v3)  |  Modul: Pengaturan Akademik > Kalender Akademik  |  Sumber PRD: Kalender Pendidikan")
t.font = Font(bold=True, size=12)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22

# Header row (row 2)
for c, name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=2, column=c, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = BORDER

# Data rows
for r, row in enumerate(ROWS, start=3):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = TOPWRAP
        cell.border = BORDER
    cat = row[1]
    if cat in CAT_FILL:
        ws.cell(row=r, column=2).fill = CAT_FILL[cat]
    src = row[7]
    if src in SRC_FONT:
        ws.cell(row=r, column=8).font = SRC_FONT[src]
    status = row[8]
    if status in STATUS_FILL:
        ws.cell(row=r, column=9).fill = STATUS_FILL[status]

# Column widths
for c, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + len(ROWS)}"

# Legend sheet
ls = wb.create_sheet("Legenda")
ls.column_dimensions["A"].width = 18
ls.column_dimensions["B"].width = 72
for r, (a, b) in enumerate(LEGEND, 1):
    ca = ls.cell(row=r, column=1, value=a)
    cb = ls.cell(row=r, column=2, value=b)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    if a in ("Kolom Sumber", "Kolom Kategori", "Kolom Status", "Catatan"):
        ca.font = Font(bold=True)

out = "docs/test-cases/TC_Kalender_Tambah.xlsx"
wb.save(out)
print("saved:", out, "| rows:", len(ROWS))
