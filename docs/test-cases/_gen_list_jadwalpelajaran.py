# -*- coding: utf-8 -*-
"""Generate Test Case sheet for List Jadwal Pelajaran (per CLAUDE.md columns)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = ["ID", "Kategori", "Skenario", "Pre-condition", "Langkah",
           "Test Data", "Expected", "Sumber", "Status", "Catatan"]

# ============================================================================
# SCOPE: halaman LIST Jadwal Pelajaran (/setting/academic/course-schedule).
# Modul Tambah sudah di-cover di TC_JadwalPelajaran_Tambah.xlsx (Fase 1).
#
# ---------------------------------------------------------------------------
# ⚠️ KEJUJURAN SUMBER — BACA DULU SEBELUM ACC
# ---------------------------------------------------------------------------
# Belum ada dokumen PRD khusus halaman List Jadwal Pelajaran. Yang tersedia:
#   (a) Hasil element analysis modul TAMBAH (sudah terverifikasi) — route,
#       judul halaman, tombol Tambah, selector grid, ikon aksi baris.
#   (b) Fixture cypress/fixtures/jadwal_pelajaran.json blok `columns` yang
#       memetakan 10 kolom. TAPI peta ini BELUM PERNAH di-assert satu pun test:
#       spec Tambah cuma memakai countRows(), dan findRowByCombo() hanya
#       menyentuh kolom 0-4. Cleanup cuma memakai kolom 8 (createdAt).
#       Artinya kalau peta kolomnya meleset, tidak ada test yang menangkapnya.
#   (c) Pola List modul lain (Tag/Kamar/Kalender): sort default terbaru->terlama,
#       filter, search, empty state. Itu PRD MODUL LAIN — dipakai di sini
#       sebagai dugaan, bukan sebagai requirement.
#
# Maka kolom `Sumber` diisi:
#   PRD           -> berasal dari element analysis Tambah yang sudah terbukti
#   PRD-ambigu    -> requirement-nya ada tapi belum jelas jawabannya
#   Asumsi        -> ekstrapolasi dari modul lain / peta kolom yang belum terbukti
#
# TC ber-Status "BLOCKED" TIDAK boleh dikoding sebelum dijawab. Daftar
# pertanyaannya ada di blok OPEN QUESTIONS di bawah.
#
# ---------------------------------------------------------------------------
# OPEN QUESTIONS (perlu jawaban user / PO / element analysis List)
# ---------------------------------------------------------------------------
#  Q1. Apakah halaman List punya SEARCH? Fixture punya testData.noMatchSearch
#      dan timeouts.searchDebounce — tapi POM tidak punya elemen search sama
#      sekali. Kemungkinan itu sisa copy dari modul lain.
#  Q2. Apakah punya FILTER? Kalau ya, field apa saja (Instansi / Tingkat /
#      Tahun Ajaran / Semester / Status)?
#  Q3. Apakah punya PAGINATION + page-size selector?
#  Q4. Berapa kolom sebenarnya, dan apa urutannya? (peta di fixture: Instansi,
#      Tahun Ajaran, Semester, Tingkat, Kelas, Jurusan, Tag, Status, Dibuat Pada,
#      Aksi = 10 kolom — BELUM TERBUKTI)
#  Q5. Kolom "Status" isinya apa? Aktif/Tidak Aktif, atau status kelengkapan
#      jadwal (mis. Draft/Lengkap)? Ini penting karena BUG-036 menghasilkan
#      record dengan header ada tapi jadwal kosong.
#  Q6. Multi Kelas ON (mis. 2 kelas) menghasilkan BERAPA BARIS di list —
#      1 baris gabungan, atau 1 baris per kelas?
#  Q7. Sort default-nya apa? (dugaan: Dibuat Pada terbaru -> terlama)
#  Q8. Teks empty state persisnya apa? (dugaan pola app: "Data Jadwal
#      Pelajaran tidak ditemukan")
#  Q9. Klik baris / ikon Edit membuka apa — halaman detail
#      /course-schedule/{id}, atau modal?
#
# ---------------------------------------------------------------------------
# KNOWN BUGS yang menyentuh List (konvensi CLAUDE.md: assert BEHAVIOR YG BENAR,
# jadi TC-nya SENGAJA FAIL sampai bug diperbaiki)
# ---------------------------------------------------------------------------
#   BUG-036  Klik Simpan di modal langsung commit ke DB -> list bertambah
#            padahal jadwal belum diisi (ghost record).
#   BUG-035  Opsi Tag tidak di-scope per Instansi -> berdampak ke isi kolom Tag.
# ============================================================================

ROWS = [
    # ================= S-A — Akses & Struktur =================
    ["TC-JP-LST-001", "Happy",
     "Halaman List Jadwal Pelajaran berhasil dimuat",
     "Login sebagai admin",
     "1) Buka Pengaturan > Akademik > Jadwal Pelajaran\n"
     "2) Amati URL, judul halaman, dan tabel",
     "-",
     "URL = /setting/academic/course-schedule; judul halaman \"Jadwal Pelajaran\"; "
     "tabel data-grid tampil",
     "PRD", "Belum dijalankan",
     "Route & pageTitle sudah terverifikasi lewat element analysis modul Tambah"],

    ["TC-JP-LST-002", "Happy",
     "Tombol \"Tambah Jadwal Pelajaran\" tampil & bisa diklik",
     "Login sebagai admin; berada di halaman List",
     "1) Amati toolbar di atas tabel\n2) Klik tombol Tambah",
     "-",
     "Tombol \"Tambah Jadwal Pelajaran\" (ikon lucide-square-plus) tampil dan "
     "membuka modal Tambah",
     "PRD", "Belum dijalankan",
     "Selector sudah terverifikasi: button[data-slot=\"dialog-trigger\"]"],

    ["TC-JP-LST-003", "Happy",
     "Header tabel menampilkan kolom lengkap sesuai peta kolom",
     "Ada >=1 data jadwal",
     "1) Buka List Jadwal Pelajaran\n2) Amati header tabel\n"
     "3) Cocokkan urutan kolom dengan fixture `columns`",
     "-",
     "Kolom berurutan: Instansi, Tahun Ajaran, Semester, Tingkat, Kelas, Jurusan, "
     "Tag, Status, Dibuat Pada, Aksi (10 kolom)",
     "Asumsi", "BLOCKED",
     "Q4 — peta kolom di fixture BELUM PERNAH di-assert test mana pun. "
     "Konfirmasi dulu ke UI asli sebelum dikoding"],

    ["TC-JP-LST-004", "Happy",
     "Setiap baris punya aksi Edit & Hapus",
     "Ada >=1 data jadwal",
     "1) Buka List\n2) Amati kolom Aksi pada baris pertama",
     "-",
     "Tiap baris menampilkan ikon Edit (svg.lucide-square-pen) dan "
     "Hapus (svg.lucide-trash), keduanya enabled",
     "Asumsi", "BLOCKED",
     "Q9 — POM sudah punya getter editIcon/deleteIcon, tapi belum pernah dipakai "
     "satu spec pun. Perlu dipastikan keduanya memang ada di baris"],

    # ================= S-B — Konsistensi data dengan modul Tambah =================
    ["TC-JP-LST-005", "Happy",
     "Jadwal yang baru dibuat muncul di list dengan kombinasi yang sesuai input",
     "Login admin; tersedia Instansi/Tingkat/Kelas valid",
     "1) Catat baseline jumlah baris\n"
     "2) Tambah jadwal via modal (Instansi+Tingkat+Kelas)\n"
     "3) Selesaikan simpan di halaman edit\n"
     "4) Kembali ke List\n"
     "5) Cari baris dengan kombinasi Instansi+TA+Semester+Tingkat+Kelas",
     "Instansi=<master.instansi>; Tingkat=<master.tingkat>; Kelas=<master.kelasSingle>",
     "Baris baru ada; kolom Instansi/Tahun Ajaran/Semester/Tingkat/Kelas "
     "sama persis dengan input; jumlah baris = baseline + 1",
     "PRD", "Belum dijalankan",
     "Pola findRowByCombo() di POM sudah memakai kolom 0-4 — bagian ini paling "
     "aman dari peta kolom"],

    ["TC-JP-LST-006", "Positif",
     "Tahun Ajaran & Semester di baris = nilai read-only dari modal (TA aktif)",
     "Ada jadwal yang baru dibuat",
     "1) Buka modal Tambah, catat nilai Tahun Ajaran & Semester (read-only)\n"
     "2) Simpan\n3) Bandingkan dengan isi baris di List",
     "Tahun Ajaran=<TA aktif>; Semester=<GANJIL/GENAP aktif>",
     "Kolom Tahun Ajaran & Semester di baris sama dengan nilai read-only di modal",
     "PRD", "Belum dijalankan",
     "Nilai read-only ini sudah terverifikasi ada di modal Tambah"],

    ["TC-JP-LST-007", "Positif",
     "Kolom Jurusan terisi saat jadwal dibuat dengan Jurusan",
     "Ada Instansi/Tingkat/Kelas/Jurusan valid",
     "1) Tambah jadwal lengkap termasuk Jurusan\n2) Buka List\n"
     "3) Amati kolom Jurusan pada baris tersebut",
     "Jurusan=<master.jurusan>",
     "Kolom Jurusan menampilkan jurusan yang dipilih",
     "Asumsi", "BLOCKED",
     "Q4 — bergantung pada kepastian index kolom Jurusan"],

    ["TC-JP-LST-008", "Positif",
     "Kolom Tag menampilkan tag yang dipilih saat Tambah",
     "Ada tag terdaftar untuk Instansi tersebut",
     "1) Tambah jadwal dengan 1 tag\n2) Buka List\n3) Amati kolom Tag",
     "Tag=[<master.tag>]",
     "Kolom Tag menampilkan tag yang dipilih",
     "Asumsi", "BLOCKED",
     "Q4 + terkait BUG-035 (opsi Tag tidak di-scope per Instansi) — hasil kolom "
     "ini bisa menyesatkan selama bug itu ada"],

    ["TC-JP-LST-009", "Positif",
     "Jadwal Multi Kelas tampil di list sesuai aturan yang disepakati",
     "Ada >=2 kelas pada satu Tingkat",
     "1) Tambah jadwal dengan Multi Kelas ON, pilih 2 kelas\n"
     "2) Buka List\n3) Hitung baris yang terbentuk & amati isi kolom Kelas",
     "Kelas=<master.kelasMulti> (2 kelas)",
     "TBD — tergantung jawaban Q6: (a) 1 baris gabungan dengan kedua kelas di "
     "kolom Kelas, ATAU (b) 2 baris terpisah",
     "PRD-ambigu", "BLOCKED",
     "Q6 — ini pertanyaan requirement, bukan sekadar selector. Jawabannya "
     "mengubah Expected dan juga assertion jumlah baris di TC lain"],

    # ================= S-C — Sort & urutan =================
    ["TC-JP-LST-010", "Positif",
     "Urutan default: data terbaru di baris paling atas",
     "Login admin",
     "1) Tambah 1 jadwal baru\n2) Buka List tanpa filter/search\n"
     "3) Amati baris teratas",
     "Jadwal baru dengan kombinasi unik",
     "Baris teratas = jadwal yang baru saja dibuat (sort by Dibuat Pada, "
     "terbaru -> terlama)",
     "Asumsi", "BLOCKED",
     "Q7 — pola ini diambil dari PRD modul Tag/Kamar/Kalender, BELUM tentu "
     "berlaku untuk Jadwal Pelajaran"],

    # ================= S-D — Empty state =================
    ["TC-JP-LST-011", "Negatif",
     "Empty state saat belum ada data jadwal",
     "Tidak ada data jadwal untuk kondisi yang ditampilkan",
     "1) Pastikan list kosong (atau pakai kondisi yang tidak punya data)\n"
     "2) Amati area tabel",
     "-",
     "Tampil empty state; tidak ada baris data; tabel tidak error",
     "Asumsi", "BLOCKED",
     "Q8 — teks persisnya belum diketahui. Pola app: \"Data {Modul} tidak "
     "ditemukan\", jadi dugaan: \"Data Jadwal Pelajaran tidak ditemukan\""],

    # ================= S-E — Persistensi & integritas =================
    ["TC-JP-LST-012", "Positif",
     "Data list tetap konsisten setelah reload halaman",
     "Ada >=1 data jadwal",
     "1) Catat jumlah baris & isi baris pertama\n2) Reload halaman (F5)\n"
     "3) Bandingkan kembali",
     "-",
     "Jumlah baris & isi baris pertama sama seperti sebelum reload "
     "(data dari backend, bukan state lokal)",
     "PRD", "Belum dijalankan",
     "Konvensi repo: verifikasi persistensi lewat reload, bukan optimistic UI"],

    ["TC-JP-LST-013", "Edge",
     "Jadwal yang dibatalkan di modal TIDAK muncul di list",
     "Login admin",
     "1) Catat baseline jumlah baris\n2) Klik Tambah, isi form lengkap\n"
     "3) Klik Batal (jangan Simpan)\n4) Amati jumlah baris list",
     "Instansi+Tingkat+Kelas valid",
     "Jumlah baris = baseline; tidak ada baris baru",
     "PRD", "Belum dijalankan",
     "Perilaku modal Radix standar; sejalan dengan TC-JP-ADD-008"],

    ["TC-JP-LST-014", "Edge",
     "Ghost record: jadwal yang belum diselesaikan TIDAK muncul di list",
     "Login admin",
     "1) Catat baseline jumlah baris\n"
     "2) Klik Tambah, isi form, klik \"Simpan\" di MODAL\n"
     "3) Setelah redirect ke halaman edit, tekan Back tanpa menyimpan\n"
     "4) Kembali ke List, hitung baris",
     "Instansi+Tingkat+Kelas valid",
     "Jumlah baris = baseline (record belum boleh commit sebelum Simpan final "
     "di halaman edit)",
     "PRD", "FAIL",
     "BUG-036 — AKTUAL: baris LANGSUNG bertambah saat Simpan di modal. "
     "TC ini sengaja FAIL sampai bug diperbaiki (konvensi CLAUDE.md). "
     "Sejalan dengan TC-JP-ADD-015"],

    ["TC-JP-LST-015", "Edge",
     "Kolom Status membedakan jadwal lengkap vs belum lengkap",
     "Ada jadwal yang header-nya ada tapi isi jadwalnya kosong",
     "1) Buat jadwal lewat modal, jangan isi jadwal per hari\n"
     "2) Buka List\n3) Amati kolom Status pada baris tersebut",
     "-",
     "TBD — tergantung jawaban Q5. Kalau Status memang menandai kelengkapan, "
     "baris ini harus tampil sebagai belum lengkap/draft",
     "PRD-ambigu", "BLOCKED",
     "Q5 — kalau kolom Status ternyata cuma Aktif/Tidak Aktif, TC ini dibuang "
     "dan diganti TC status biasa"],

    # ================= S-F — Kandidat, tergantung fitur yang tersedia =================
    ["TC-JP-LST-016", "Positif",
     "[KANDIDAT] Search menemukan jadwal berdasarkan kata kunci",
     "Ada >=1 data jadwal",
     "1) Ketik kata kunci di kolom pencarian\n2) Amati hasil",
     "Keyword = nama kelas / instansi yang ada",
     "Hanya baris yang cocok dengan kata kunci yang tampil",
     "Asumsi", "BLOCKED",
     "Q1 — BELUM TENTU ADA. Fixture punya testData.noMatchSearch & "
     "timeouts.searchDebounce, TAPI POM tidak punya elemen search sama sekali "
     "— kemungkinan besar sisa copy dari modul lain. Konfirmasi dulu"],

    ["TC-JP-LST-017", "Negatif",
     "[KANDIDAT] Search tanpa hasil menampilkan empty state",
     "Ada data jadwal",
     "1) Ketik kata kunci yang pasti tidak ada\n2) Amati hasil",
     "Keyword = <testData.noMatchSearch>",
     "Tidak ada baris; empty state tampil; tabel tidak error",
     "Asumsi", "BLOCKED", "Q1 — sama seperti TC-016"],

    ["TC-JP-LST-018", "Positif",
     "[KANDIDAT] Filter mempersempit list sesuai kriteria",
     "Ada data dari >=2 instansi berbeda",
     "1) Terapkan filter (mis. Instansi)\n2) Amati semua baris hasil",
     "Instansi=<master.instansi>",
     "Semua baris yang tampil punya nilai kolom sesuai filter; "
     "baris di luar kriteria tidak tampil",
     "Asumsi", "BLOCKED",
     "Q2 — belum diketahui apakah List punya filter dan field apa saja"],

    ["TC-JP-LST-019", "Positif",
     "[KANDIDAT] Pagination & page size berfungsi",
     "Jumlah data melebihi 1 halaman",
     "1) Amati kontrol pagination\n2) Ganti page size\n3) Pindah halaman",
     "Page size = 10 / 25 / 50",
     "Jumlah baris per halaman sesuai page size; navigasi antar halaman "
     "menampilkan data berbeda",
     "Asumsi", "BLOCKED",
     "Q3 — belum diketahui apakah ada pagination"],

    ["TC-JP-LST-020", "Positif",
     "[KANDIDAT] Ikon Edit membuka halaman/modal edit jadwal yang benar",
     "Ada >=1 data jadwal",
     "1) Klik ikon Edit pada satu baris\n2) Amati tujuan & data yang tampil",
     "-",
     "TBD — tergantung jawaban Q9. Kalau halaman: URL "
     "/setting/academic/course-schedule/{id} dan datanya sesuai baris tersebut",
     "PRD-ambigu", "BLOCKED",
     "Q9 — modul Tambah sudah membuktikan ADA halaman /{id} setelah simpan, "
     "tapi belum tentu ikon Edit menuju ke sana"],
]

# ============================================================================
# Styling — mengikuti pola generator TC sheet lain di folder ini
# ============================================================================
NAVY = "1F4E78"
SUBTLE = "595959"
KATEGORI_FILL = {
    "Happy": "C6EFCE",
    "Positif": "DDEBF7",
    "Negatif": "FFEB9C",
    "Edge": "FCE4D6",
}
STATUS_COLOR = {
    "Belum dijalankan": "595959",
    "BLOCKED": "BF8F00",
    "FAIL": "C00000",
    "PASS": "548235",
    "SKIP": "808080",
}
SUMBER_COLOR = {
    "PRD": "548235",
    "PRD-ambigu": "BF8F00",
    "Asumsi": "C00000",
}
WIDTHS = [16, 10, 44, 34, 52, 34, 54, 12, 16, 46]

THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()
ws = wb.active
ws.title = "List Jadwal Pelajaran"

ws["A1"] = "TEST CASE — List Jadwal Pelajaran (Pengaturan > Akademik)"
ws["A1"].font = Font(name="Arial", bold=True, size=14, color=NAVY)
ws["A2"] = ("Route: /setting/academic/course-schedule  |  "
            "Sumber: PRD = dari element analysis modul Tambah (terbukti) · "
            "PRD-ambigu = requirement belum jelas · Asumsi = ekstrapolasi, belum terbukti  |  "
            "Status BLOCKED = JANGAN dikoding sebelum open question-nya dijawab")
ws["A2"].font = Font(name="Arial", size=9, color=SUBTLE)

HEADER_ROW = 4
hfill = PatternFill("solid", fgColor=NAVY)
for ci, name in enumerate(COLUMNS, start=1):
    c = ws.cell(row=HEADER_ROW, column=ci, value=name)
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill = hfill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    ws.column_dimensions[get_column_letter(ci)].width = WIDTHS[ci - 1]

for ri, row in enumerate(ROWS, start=HEADER_ROW + 1):
    for ci, val in enumerate(row, start=1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(
            horizontal="center" if ci in (1, 2, 8, 9) else "left",
            vertical="top", wrap_text=True)
        c.border = BORDER
    kat = ws.cell(row=ri, column=2)
    fill = KATEGORI_FILL.get((kat.value or "").strip())
    if fill:
        kat.fill = PatternFill("solid", fgColor=fill)
        kat.font = Font(name="Arial", size=10, bold=True)
    sm = ws.cell(row=ri, column=8)
    sm.font = Font(name="Arial", size=10, bold=True,
                   color=SUMBER_COLOR.get((sm.value or "").strip(), "000000"))
    st = ws.cell(row=ri, column=9)
    st.font = Font(name="Arial", size=10, bold=True,
                   color=STATUS_COLOR.get((st.value or "").strip(), "000000"))
    ws.row_dimensions[ri].height = 104

ws.freeze_panes = "A5"
ws.sheet_view.showGridLines = False

OUT = "TC_JadwalPelajaran_List.xlsx"
wb.save(OUT)

blocked = sum(1 for r in ROWS if r[8] == "BLOCKED")
siap = sum(1 for r in ROWS if r[8] in ("Belum dijalankan", "FAIL"))
print(f"OK -> {OUT} ({len(ROWS)} TC | siap dikoding: {siap} | BLOCKED: {blocked})")
