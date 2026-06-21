# -*- coding: utf-8 -*-
"""Generate Test Case sheet for Hapus Tag (per CLAUDE.md columns)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = ["ID", "Kategori", "Skenario", "Pre-condition", "Langkah",
           "Test Data", "Expected", "Sumber", "Status", "Catatan"]

# PRD Hapus Tag:
#   1. Aksi -> Hapus pada baris tag
#   2. Sistem menampilkan popup delete confirmation
#   3. Klik Hapus -> data terhapus + pesan sukses
#   4. Klik Batal -> tidak hapus, kembali ke list
#
# DEFERRED (per instruksi user): "Tag yang berhasil dihapus tidak dapat digunakan kembali
# untuk data diri siswa serta pengaturan presensi kegiatan, pembuatan tagihan."
#   -> butuh fitur konsumen (Data Siswa/Guru/Staff, Presensi, Tagihan) yg belum ada
#
# Catatan: PRD typo "list kamar" -> diasumsikan "list tag" (modul Tag, bukan Kamar).
ROWS = [
    # ---------------- HAPPY ----------------
    ["TC-TAG-HPS-001", "Happy",
     "Buka popup konfirmasi hapus via Aksi -> Hapus",
     "Login admin; minimal 1 tag di list",
     "1) Buka list Tag\n2) Pada baris target klik ikon Hapus (svg.lucide-trash) -> .closest('button')",
     "Tag existing apa saja",
     "Popup konfirmasi tampil dgn judul \"Hapus Tag\"; ada tombol \"Hapus\" dan \"Batal\"",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-HPS-002", "Happy",
     "Hapus tag -> sukses (toast + row hilang)",
     "Ada tag QA baru di list",
     "1) Seed tag QA\n2) Aksi -> Hapus pada baris target\n3) Klik \"Hapus\" di popup",
     "QA<6digit-ts><seq>",
     "Toast sukses \"Tag berhasil dihapus\"; popup tertutup; baris hilang dari list",
     "PRD", "Belum dijalankan",
     "Teks toast = Asumsi (PRD hanya sebut \"pesan sukses\"); assertion title-only (BUG-021 mitigasi)"],

    ["TC-TAG-HPS-003", "Happy",
     "Klik Batal di popup -> tidak terhapus, kembali ke list",
     "Popup konfirmasi terbuka utk baris target",
     "1) Seed tag QA\n2) Aksi -> Hapus\n3) Klik \"Batal\"",
     "QA<ts><seq>",
     "Popup tertutup; baris target TETAP ada di list; tidak ada toast sukses",
     "PRD", "Belum dijalankan", ""],

    # ---------------- POSITIF ----------------
    ["TC-TAG-HPS-004", "Positif",
     "Hapus tag pada instansi sekunder (Sekolah Alam)",
     "Ada tag QA di instansi sekunder",
     "1) Seed tag QA @ Sekolah Alam\n2) Aksi -> Hapus -> Hapus",
     "QA<ts><seq> @ Sekolah Alam",
     "Sukses terhapus; hanya baris instansi tsb hilang; tag instansi lain tidak terpengaruh",
     "Asumsi", "Belum dijalankan",
     "PRD tidak membahas multi-instansi pada hapus -> Asumsi konsistensi modul lain"],

    ["TC-TAG-HPS-005", "Positif",
     "Hapus tag setelah Cari by nama",
     "Ada tag QA yg bisa dicari",
     "1) Seed tag QA\n2) Search nama tag (debounce settle)\n3) Aksi -> Hapus -> Hapus",
     "QA<ts><seq>",
     "Baris hasil pencarian terhapus; toast sukses; list refresh tanpa baris itu",
     "Asumsi", "Belum dijalankan",
     "Pakai cy.wait(DEBOUNCE) konsisten POM (no intercept)"],

    # ---------------- EDGE ----------------
    ["TC-TAG-HPS-006", "Edge",
     "Tutup popup via Escape -> tidak terhapus",
     "Popup konfirmasi terbuka",
     "1) Seed tag QA\n2) Aksi -> Hapus\n3) Tekan Escape",
     "QA<ts><seq>",
     "Popup tertutup; tidak ada perubahan data; baris TETAP ada",
     "Asumsi", "Belum dijalankan",
     "PRD hanya sebut tombol Batal; perilaku Escape = Asumsi"],

    ["TC-TAG-HPS-007", "Edge",
     "Tutup popup via X close icon -> tidak terhapus",
     "Popup konfirmasi terbuka",
     "1) Seed tag QA\n2) Aksi -> Hapus\n3) Klik X close (lucide-x)",
     "QA<ts><seq>",
     "Popup tertutup; baris TETAP ada",
     "Asumsi", "Belum dijalankan",
     "Konfirmasi via element analysis apakah popup punya X close (konsisten modal Tambah/Edit)"],

    ["TC-TAG-HPS-008", "Edge",
     "Hapus baris terakhir di hasil search -> empty state",
     "Hasil search menyisakan tepat 1 baris",
     "1) Seed tag QA dgn nama unik\n2) Search nama tsb\n3) Aksi -> Hapus -> Hapus",
     "QA<ts><seq> tunggal di hasil search",
     "Setelah hapus, muncul empty state \"Data Tag tidak ditemukan\"",
     "Asumsi", "Belum dijalankan",
     "emptyStateTitle diambil dari fixture"],

    ["TC-TAG-HPS-009", "Edge",
     "Persistence: hapus sukses -> reload -> data tetap hilang",
     "TC-TAG-HPS-002 sukses",
     "1) Setelah hapus sukses, reload halaman (visit)\n2) Search nama tag yg dihapus",
     "QA<ts><seq>",
     "Tag TIDAK muncul setelah reload (persist di backend, bukan sekadar update optimistik)",
     "Asumsi", "Belum dijalankan",
     "Sesuai aturan persistence CLAUDE.md; pakai assertNotPersisted-style"],

    ["TC-TAG-HPS-010", "Edge",
     "Popup konfirmasi menampilkan nama tag yg akan dihapus",
     "Popup konfirmasi terbuka utk baris target",
     "1) Seed tag QA\n2) Aksi -> Hapus\n3) Baca isi teks popup",
     "QA<ts><seq>",
     "Teks popup memuat nama tag target (cth: \"Apakah Anda yakin ingin menghapus tag QA...?\")",
     "PRD-ambigu", "Belum dijalankan",
     "Perlu cek HTML asli apakah popup memuat nama tag; PRD tidak eksplisit"],

    # ---------------- DEFERRED (asosiasi Member) ----------------
    ["TC-TAG-HPS-011", "Positif",
     "Tag Tipe=Siswa yg dihapus tidak muncul di picker Data Diri Siswa",
     "Tag QA Tipe=Siswa pernah dipakai di Data Siswa",
     "1) Hapus tag QA Tipe=Siswa\n2) Buka form Data Diri Siswa\n3) Cek picker Tag -> tag QA tidak muncul",
     "Tag QA Tipe=Siswa",
     "Tag terhapus TIDAK muncul lagi di picker Data Siswa (sesuai PRD validasi asosiasi)",
     "PRD", "DITUNDA",
     "DEFERRED per instruksi user: asosiasi cross-fitur (Data Siswa) belum bisa diakses"],

    ["TC-TAG-HPS-012", "Positif",
     "Tag Tipe=Guru&Staff yg dihapus tidak muncul di picker Data Guru/Staff",
     "Tag QA Tipe=Guru&Staff pernah dipakai",
     "1) Hapus tag QA Tipe=Guru&Staff\n2) Buka form Data Guru/Staff\n3) Cek picker Tag -> tidak muncul",
     "Tag QA Tipe=Guru & Staff",
     "Tag terhapus TIDAK muncul lagi di picker Data Guru/Staff",
     "PRD", "DITUNDA",
     "DEFERRED per instruksi user: asosiasi cross-fitur belum bisa diakses"],

    ["TC-TAG-HPS-013", "Positif",
     "Tag yg dihapus tidak muncul di picker Presensi & Tagihan",
     "Tag QA pernah dipakai di Presensi/Tagihan",
     "1) Hapus tag QA\n2) Buka pengaturan Presensi Kegiatan -> picker Tag tidak menampilkan tag tsb\n3) Buka pembuatan Tagihan -> picker Tag tidak menampilkan tag tsb",
     "Tag QA",
     "Tag terhapus TIDAK muncul di picker Presensi maupun Tagihan",
     "PRD", "DITUNDA",
     "DEFERRED per instruksi user: butuh fitur Presensi/Tagihan yg belum tersedia"],
]

LEGEND = [
    ["Kolom Sumber", ""],
    ["PRD", "Skenario diturunkan langsung dari PRD Hapus Tag"],
    ["PRD-ambigu", "PRD menyinggung tapi tidak jelas/typo -> perlu konfirmasi via element analysis"],
    ["Asumsi", "Tidak ada di PRD; asumsi berbasis konvensi/fixture/modul lain"],
    ["", ""],
    ["Kolom Kategori", ""],
    ["Happy", "Alur utama sukses sesuai desain"],
    ["Positif", "Variasi valid non-utama"],
    ["Negatif", "Input/kondisi yang harus ditolak sistem"],
    ["Edge", "Batas/kondisi tepi & verifikasi persistence"],
    ["", ""],
    ["Kolom Status", ""],
    ["Belum dijalankan", "Sudah di-generate, menunggu eksekusi Cypress"],
    ["DITUNDA", "Sengaja belum di-run; menunggu fitur konsumen (Data Siswa / Guru-Staff / Presensi / Tagihan) tersedia"],
    ["FAIL", "Gagal saat run; lihat ref BUG-### di Catatan"],
    ["", ""],
    ["Catatan", "Tanpa kategori security/injection (sesuai kesepakatan)"],
    ["", "PRD typo \"list kamar\" -> diasumsikan \"list tag\" (modul Tag)"],
    ["", "PRD typo \"hapus data tag\"/\"hapus tag\" -> konsisten konteks Hapus Tag"],
    ["", "DEFERRED: asosiasi cross-fitur (Data Siswa/Guru/Staff, Presensi, Tagihan) butuh modul konsumen yg belum ada"],
    ["", "Teks toast = Asumsi (PRD hanya \"pesan sukses\"); assertion title-only (mitigasi BUG-021)"],
]

# ---- styling (mirror modul lain) ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CAT_FILL = {
    "Happy":   PatternFill("solid", fgColor="E2EFDA"),
    "Positif": PatternFill("solid", fgColor="DDEBF7"),
    "Negatif": PatternFill("solid", fgColor="FCE4D6"),
    "Edge":    PatternFill("solid", fgColor="FFF2CC"),
}
SRC_FONT = {
    "PRD":        Font(color="375623"),
    "PRD-ambigu": Font(color="BF8F00", bold=True),
    "Asumsi":     Font(color="C00000", bold=True),
}
STATUS_FILL = {
    "DITUNDA": PatternFill("solid", fgColor="D9D9D9"),
    "FAIL":    PatternFill("solid", fgColor="FFC7CE"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPWRAP = Alignment(vertical="top", wrap_text=True)
WIDTHS = [16, 10, 38, 30, 44, 28, 50, 12, 16, 46]

wb = Workbook()
ws = wb.active
ws.title = "Hapus Tag"

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
t = ws.cell(row=1, column=1, value="Test Case - Hapus Tag (CARDS School v3)  |  Modul: Pengaturan Akademik > Tag  |  Sumber PRD: Hapus Tag")
t.font = Font(bold=True, size=12)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22

for c, name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=2, column=c, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = BORDER

for r, row in enumerate(ROWS, start=3):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = TOPWRAP
        cell.border = BORDER
    cat = row[1]
    if cat in CAT_FILL:
        ws.cell(row=r, column=2).fill = CAT_FILL[cat]
    src = row[7]
    if src in SRC_FONT:
        ws.cell(row=r, column=8).font = SRC_FONT[src]
    status = row[8]
    if status in STATUS_FILL:
        ws.cell(row=r, column=9).fill = STATUS_FILL[status]

for c, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + len(ROWS)}"

ls = wb.create_sheet("Legenda")
ls.column_dimensions["A"].width = 28
ls.column_dimensions["B"].width = 72
for r, (a, b) in enumerate(LEGEND, 1):
    ca = ls.cell(row=r, column=1, value=a)
    cb = ls.cell(row=r, column=2, value=b)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    if a in ("Kolom Sumber", "Kolom Kategori", "Kolom Status", "Catatan"):
        ca.font = Font(bold=True)

out = "docs/test-cases/TC_Tag_Hapus.xlsx"
wb.save(out)
print("saved:", out, "| rows:", len(ROWS))
