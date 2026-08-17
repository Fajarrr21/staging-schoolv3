# -*- coding: utf-8 -*-
"""Generate Test Case sheet for List Kalender Akademik (per CLAUDE.md columns).

Catatan PRD:
  PRD Kalender Pendidikan HANYA membahas form Edit per instansi (Foto header, Awal minggu,
  Tampilan nama minggu). PRD TIDAK menyebut list/filter/search/pagination sama sekali.
  Maka mayoritas TC list bersumber 'Asumsi' (pola konsisten dengan modul Akademik lain:
  Mapel/Jurusan/Tingkat/Kelas/Kamar/Tag) atau 'PRD-ambigu' (turunan tidak langsung).

Field tabel (5 kolom visual = 6 td HTML, dikonfirmasi user 25 Juni 2026):
  col-0 (td eq 0) = Instansi - span.font-medium
  col-1 (td eq 1) = Awal pekan dimulai - span[data-slot=badge]
  col-2 (td eq 2) = Nama Pekan - span[data-slot=badge]
  col-3 (td eq 3) = Header - img[alt=header] kalau ada upload, atau text "-" kalau kosong
  col-4 (td eq 4) = Edit button - button[data-slot=dialog-trigger] + svg.lucide-square-pen
  col-5 (td eq 5) = Hapus button - button[data-slot=dialog-trigger] + svg.lucide-trash

Header tabel (4 header dengan dropdown-menu-trigger + svg.lucide-chevrons-up-down):
  Semua header punya dropdown menu = SORT (Asc/Desc) per kolom.
  Filter Instansi adalah BUTTON TERPISAH di toolbar atas tabel, BUKAN
  bagian dari dropdown header kolom Instansi.

Toolbar:
  - Tombol "Tambah Kalender Akademik" (button[data-slot=dialog-trigger] + svg.lucide-square-plus)
  - Filter by Instansi (button terpisah, dropdown list instansi + opsi "Semua")
  - Pagination - default 10 baris per halaman, di div[data-slot=data-grid-pagination]
  - TIDAK ada search tabel (search icon di global header app saja)

Known bugs ter-cover di TC List:
  BUG-027 - opsi "Semua" muncul 2x di dropdown filter Instansi (confusing UX)
  BUG-028 - inkonsistensi naming "Akademik" vs "Pendidikan" di tombol vs empty state
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = ["ID", "Kategori", "Skenario", "Pre-condition", "Langkah",
           "Test Data", "Expected", "Sumber", "Status", "Catatan"]

ROWS = [
    # ---------------- HAPPY ----------------
    ["TC-KLD-LST-001", "Happy",
     "Buka halaman List Kalender Akademik -> tampil tabel 5 kolom visual (6 td HTML) + toolbar lengkap",
     "Login admin; ada minimal 1 data kalender existing di sistem",
     "1) Navigate ke Pengaturan > Akademik > Kalender Akademik\n2) Tunggu halaman load\n3) Amati tabel & toolbar",
     "-",
     "Tabel render 5 kolom visual (Instansi, Awal pekan dimulai, Nama Pekan, Header, Action) = 6 td di HTML (Edit & Hapus terpisah cell); tombol \"Tambah Kalender Akademik\" visible (kanan atas); button filter Instansi visible di toolbar; header kolom punya sort dropdown (svg.lucide-chevrons-up-down); pagination control (baris per halaman, default 10); minimal 1 row data muncul",
     "Asumsi", "Belum dijalankan",
     "Pola list konsisten dengan modul Akademik lain"],

    ["TC-KLD-LST-002", "Happy",
     "Row data menampilkan field sesuai input form (Instansi + badge Awal pekan + badge Nama Pekan)",
     "Login admin; ada data kalender (mis. instansi=SMA, Awal pekan=Senin, Nama Pekan=Minggu)",
     "1) Buka List Kalender Akademik\n2) Cari baris instansi target\n3) Inspect 3 kolom pertama (Nama, Awal pekan badge, Nama Pekan badge)",
     "Instansi=<existing>; Awal pekan=Senin; Nama Pekan=Minggu",
     "Kol-0 menampilkan nama instansi exact (span.font-medium); kol-1 menampilkan badge text \"Senin\"; kol-2 menampilkan badge text \"Minggu\"",
     "PRD-ambigu", "Belum dijalankan",
     "PRD describe form fields; rendering di list adalah turunan logis"],

    ["TC-KLD-LST-003", "Happy",
     "Kol-3 Header: row dengan upload header menampilkan <img>; tanpa upload menampilkan \"-\"",
     "Login admin; ada 2 row existing di list: 1 dengan header image (existing data) + 1 tanpa upload",
     "1) Buka List Kalender Akademik\n2) Cari row WITH header image existing -> inspect kol-3\n3) Cari row TANPA upload header -> inspect kol-3",
     "Row A: existing dengan image (instansi yang dah punya kalender ber-header); Row B: tanpa upload",
     "Row A: kol-3 berisi <img alt=\"header\" src=\"...\">; Row B: kol-3 berisi teks \"-\" (placeholder, no img element)",
     "PRD-ambigu", "Belum dijalankan",
     "Pakai existing data (user maintain). BUG-026 tidak terkait — itu issue Tambah baru, bukan list rendering"],

    # ---------------- POSITIF ----------------
    ["TC-KLD-LST-004", "Positif",
     "Filter Instansi -> list ter-filter ke instansi yang dipilih saja",
     "Login admin; ada minimal 2 data kalender di 2 instansi berbeda",
     "1) Buka List Kalender Akademik\n2) Klik button filter Instansi di toolbar (dropdown muncul)\n3) Pilih salah satu instansi\n4) Amati tabel",
     "Filter Instansi = <instansi A>",
     "Tabel hanya menampilkan row dengan kol-0 = <instansi A>; row instansi lain tidak muncul; label trigger button berubah jadi nama instansi yang dipilih",
     "Asumsi", "Belum dijalankan",
     "Pola filter konsisten dengan modul lain (mis. Kelas/Kamar)"],

    ["TC-KLD-LST-005", "Positif",
     "Reset filter Instansi -> list kembali ke semua data",
     "Filter Instansi sedang aktif (state dari TC-004)",
     "1) Klik button filter Instansi\n2) Pilih opsi reset (dari opsi 'Semua' ke-2 yang me-reset trigger label balik ke 'Instansi' - lihat BUG-027)\n3) Amati tabel",
     "Filter Instansi = (clear)",
     "Tabel kembali menampilkan semua data kalender existing (sama dengan state TC-001); trigger label balik ke 'Instansi'",
     "Asumsi", "Belum dijalankan",
     "BUG-027 - 2 opsi 'Semua' confusing; test pakai opsi reset yang benar"],

    ["TC-KLD-LST-006", "Positif",
     "Reload halaman -> data tetap tampil (backend persistence)",
     "Login admin; ada data kalender existing di list",
     "1) Buka List Kalender Akademik\n2) Catat jumlah & isi row\n3) Reload halaman (F5)\n4) Bandingkan row sebelum & sesudah reload",
     "-",
     "Setelah reload, jumlah row & isi tiap row identik dengan sebelum reload (data ter-persist di backend, bukan optimistic state)",
     "Asumsi", "Belum dijalankan",
     "Pola assertPersisted standar untuk semua modul list"],

    ["TC-KLD-LST-007", "Positif",
     "Badge Awal pekan & Nama Pekan render dengan warna/styling konsisten (per data-slot=\"badge\")",
     "Login admin; ada data kalender dengan berbagai kombinasi (Senin/Minggu, Ahad/Minggu)",
     "1) Buka List Kalender Akademik\n2) Inspect badge di kol-1 & kol-2 multiple row\n3) Bandingkan styling (warna, ukuran, border)",
     "-",
     "Semua badge di kol-1 & kol-2 punya class konsisten (data-slot=\"badge\"); warna primary-accent (biru) sama di semua row; tidak ada badge yang \"plain text\"",
     "Asumsi", "Belum dijalankan",
     "Visual consistency check"],

    ["TC-KLD-LST-008", "Positif",
     "Klik icon Edit (svg.lucide-square-pen) di row -> modal Edit terbuka",
     "Login admin; ada data kalender existing di list",
     "1) Buka List Kalender Akademik\n2) Klik icon Edit di salah satu row\n3) Amati apakah modal terbuka",
     "Row instansi target",
     "Modal Edit terbuka (dialog-content dengan title \"Edit Kalender Akademik\"); form ter-prefill dengan data row yang dipilih (Instansi readonly?; Awal pekan; Nama Pekan; preview header kalau ada)",
     "Asumsi", "Belum dijalankan",
     "Trigger check only; flow edit detail di modul Edit"],

    ["TC-KLD-LST-009", "Positif",
     "Klik icon Hapus (svg.lucide-trash) di row -> dialog konfirmasi terbuka",
     "Login admin; ada data kalender existing di list",
     "1) Buka List Kalender Akademik\n2) Klik icon Hapus di salah satu row\n3) Amati apakah dialog konfirmasi muncul",
     "Row instansi target",
     "Dialog konfirmasi terbuka (mis. \"Apakah Anda yakin ingin menghapus...\") dengan tombol Hapus + Batal; row TIDAK terhapus sebelum konfirmasi",
     "Asumsi", "Belum dijalankan",
     "Trigger check only; flow hapus detail di modul Hapus"],

    # ---------------- NEGATIF ----------------
    ["TC-KLD-LST-010", "Negatif",
     "Filter Instansi yang BELUM punya kalender -> empty state muncul (naming inkonsisten - BUG-028)",
     "Login admin; ada instansi X yang belum punya kalender",
     "1) Buka List Kalender Akademik\n2) Klik button filter Instansi\n3) Pilih instansi X (yang belum punya kalender)\n4) Amati tabel & empty state title/CTA",
     "Filter Instansi = <instansi X tanpa kalender>",
     "Tabel kosong; muncul empty state dengan title 'Data Kalender Akademik tidak ditemukan' (KONSISTEN dengan naming modul); inline CTA 'Tambah Kalender Akademik'; tombol Tambah Kalender Akademik di toolbar tetap visible. NOTE: aktual saat ini menampilkan 'Kalender Pendidikan' di title/CTA - BUG-028",
     "Asumsi", "FAIL",
     "BUG-028 - Empty state pakai 'Pendidikan' bukan 'Akademik' (inkonsisten dgn tombol & breadcrumb)"],

    # ---------------- EDGE ----------------
    ["TC-KLD-LST-011", "Positif",
     "Sort kolom Instansi ascending -> descending via dropdown header",
     "Login admin; ada minimal 3 data kalender (variasi nama instansi)",
     "1) Buka List Kalender Akademik\n2) Klik header kolom 'Instansi' (button dropdown-menu-trigger)\n3) Dropdown menu muncul, pilih 'Sort Asc' (atau setara)\n4) Amati urutan row -> harus ascending A->Z\n5) Klik header lagi, pilih 'Sort Desc'\n6) Amati urutan row -> harus descending Z->A",
     "-",
     "Klik header -> dropdown menu (bukan toggle langsung). Pilih ascending: row terurut A-Z by Instansi. Pilih descending: row terurut Z-A. Sort icon di header bisa berubah indicator (up/down)",
     "Asumsi", "Belum dijalankan",
     "Pattern: dropdown-menu-trigger di header (Radix DropdownMenu), bukan native sort"],

    ["TC-KLD-LST-012", "Edge",
     "Pagination -> navigate ke page 2 + ubah baris per halaman",
     "Login admin; ada > 10 data kalender di sistem (saat ini hanya 5 - butuh seed)",
     "1) Buka List Kalender Akademik\n2) Scroll ke bawah cari div[data-slot='data-grid-pagination']\n3) Klik next page (svg.lucide-move-right) -> amati row di page 2\n4) Ubah select-trigger 'Baris Per Halaman' (mis. 10 -> 25)\n5) Amati count row & indicator '1 - N Dari M' update",
     "Default page size = 10",
     "Page 2 menampilkan row 11+ (sesuai page size 10); indicator '1 - 10 Dari N' update jadi '11 - X Dari N'; ubah page size mengubah jumlah row yang ditampilkan dan recount total page",
     "Asumsi", "DITUNDA",
     "Defer sampai data > 10 (saat ini total 5: '1 - 5 Dari 5'). Tunggu seed atau real data growth"],

    ["TC-KLD-LST-013", "Edge",
     "Persistence row baru -> setelah Tambah dari modal, row baru langsung muncul di list TANPA reload",
     "Login admin; halaman List Kalender Akademik terbuka",
     "1) Klik \"Tambah Kalender Akademik\"\n2) Isi form valid (instansi baru + Awal pekan + Nama Pekan)\n3) Klik Simpan\n4) Modal close\n5) Cari row baru di tabel TANPA reload",
     "Instansi baru; Awal pekan=Senin; Nama Pekan=Minggu",
     "Row baru muncul instan di tabel setelah toast sukses (optimistic UI atau auto-refetch). Tidak perlu reload manual untuk lihat data baru",
     "Asumsi", "Belum dijalankan",
     "Bridges Tambah & List — verify integration tanpa reload"],
]

LEGEND = [
    ["Kolom Sumber", ""],
    ["PRD", "Skenario diturunkan langsung dari PRD Kalender Pendidikan"],
    ["PRD-ambigu", "PRD menyinggung tapi tidak detail; turunan logis (mis. rendering field di list)"],
    ["Asumsi", "PRD tidak menyebut sama sekali; berbasis pola konsisten modul Akademik lain"],
    ["", ""],
    ["Kolom Kategori", ""],
    ["Happy", "Alur utama sukses sesuai desain"],
    ["Positif", "Variasi valid non-utama (filter, reload, render konsistensi, trigger Edit/Hapus)"],
    ["Negatif", "Input/kondisi yang harus menampilkan empty state / error"],
    ["Edge", "Pagination + persistence tepi"],
    ["", ""],
    ["Kolom Status", ""],
    ["Belum dijalankan", "Sudah di-generate, menunggu eksekusi Cypress"],
    ["DITUNDA", "Defer sampai element analysis konfirmasi UI element terkait"],
    ["FAIL", "Diketahui FAIL karena bug app (lihat kolom Catatan untuk BUG-### ref)"],
    ["", ""],
    ["Catatan", "PRD Kalender HANYA membahas form Edit per instansi; list/filter/search/pagination 100% asumsi konvensi"],
    ["", "Modul referensi pattern: Kamar/Kelas/Tingkat/Tag (semua punya List sheet serupa)"],
    ["", "Element analysis SETELAH ACC TC sheet (memory rule: STOP & tunggu ACC sebelum element analysis)"],
    ["", "Tanpa kategori security/injection (CLAUDE.md kesepakatan)"],
    ["", "Sort + Pagination DIKONFIRMASI ada (25 Juni 2026); Search tabel TIDAK ada — yang ada cuma filter Instansi"],
]

# ---- styling (copy pattern dari _gen_tambah_kalender.py) ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CAT_FILL = {
    "Happy":   PatternFill("solid", fgColor="E2EFDA"),
    "Positif": PatternFill("solid", fgColor="DDEBF7"),
    "Negatif": PatternFill("solid", fgColor="FCE4D6"),
    "Edge":    PatternFill("solid", fgColor="FFF2CC"),
}
SRC_FONT = {
    "PRD":        Font(color="375623"),
    "PRD-ambigu": Font(color="BF8F00", bold=True),
    "Asumsi":     Font(color="C00000", bold=True),
}
STATUS_FILL = {
    "DITUNDA": PatternFill("solid", fgColor="D9D9D9"),
    "FAIL":    PatternFill("solid", fgColor="F4B084"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPWRAP = Alignment(vertical="top", wrap_text=True)
WIDTHS = [16, 12, 42, 36, 48, 28, 56, 14, 16, 50]

wb = Workbook()
ws = wb.active
ws.title = "List Kalender"

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
t = ws.cell(row=1, column=1, value="Test Case - List Kalender Akademik (CARDS School v3)  |  Modul: Pengaturan Akademik > Kalender Akademik  |  Sumber: Asumsi konvensi (PRD tidak cover list)")
t.font = Font(bold=True, size=12)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22

for c, name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=2, column=c, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = BORDER

for r, row in enumerate(ROWS, start=3):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = TOPWRAP
        cell.border = BORDER
    cat = row[1]
    if cat in CAT_FILL:
        ws.cell(row=r, column=2).fill = CAT_FILL[cat]
    src = row[7]
    if src in SRC_FONT:
        ws.cell(row=r, column=8).font = SRC_FONT[src]
    status = row[8]
    if status in STATUS_FILL:
        ws.cell(row=r, column=9).fill = STATUS_FILL[status]

for c, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + len(ROWS)}"

ls = wb.create_sheet("Legenda")
ls.column_dimensions["A"].width = 18
ls.column_dimensions["B"].width = 78
for r, (a, b) in enumerate(LEGEND, 1):
    ca = ls.cell(row=r, column=1, value=a)
    cb = ls.cell(row=r, column=2, value=b)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    if a in ("Kolom Sumber", "Kolom Kategori", "Kolom Status", "Catatan"):
        ca.font = Font(bold=True)

out = "docs/test-cases/TC_Kalender_List.xlsx"
wb.save(out)
print("saved:", out, "| rows:", len(ROWS))
