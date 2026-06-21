# -*- coding: utf-8 -*-
"""Generate Test Case sheet for Edit Kamar (per CLAUDE.md columns)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = ["ID", "Kategori", "Skenario", "Pre-condition", "Langkah",
           "Test Data", "Expected", "Sumber", "Status", "Catatan"]

# Form Edit Kamar (PRD):
#   a. Instansi*   (required, select)
#   b. Nama Kamar* (required, text)
#   c. PIC         (optional, select -> guru & staff instansi terpilih)
#   d. Lokasi      (optional, text)
#   e. Status*     (required: Aktif / Tidak Aktif)
# SKIP: skenario downstream Status Tidak Aktif (hidden di data diri/presensi) - belum bisa diuji

ROWS = [
    # ---------------- HAPPY ----------------
    ["TC-KMR-EDT-001", "Happy",
     "Edit nama kamar berhasil → perubahan tersimpan & tampil di list",
     "Login admin; ada kamar di list",
     "1) Seed 1 kamar (Tambah)\n2) Klik ikon Edit pada baris kamar tsb\n3) Ubah Nama Kamar menjadi nama baru unik\n4) Klik \"Simpan\"",
     "Nama baru=QA<ts><seq>-EDIT",
     "Toast sukses \"Kamar berhasil diperbarui\"; modal tertutup; kolom Nama Kamar di list ter-update ke nama baru; reload → nama baru tetap ada",
     "PRD", "Belum dijalankan",
     "Teks toast = Asumsi (PRD hanya sebut \"menyimpan perubahan\"). Verifikasi persist via reload+search"],

    ["TC-KMR-EDT-002", "Happy",
     "Edit PIC berhasil → kolom PIC ter-update di list",
     "Login admin; ada kamar tanpa PIC di list; instansi punya minimal 1 guru/staff",
     "1) Seed kamar tanpa PIC\n2) Buka Edit\n3) Pilih PIC pertama dari dropdown\n4) Klik \"Simpan\"",
     "Instansi=SDIT; PIC=<guru/staff pertama>",
     "Toast sukses; modal tutup; kolom PIC di baris kamar menampilkan nama PIC yang dipilih",
     "PRD", "Belum dijalankan",
     "Verifikasi nama PIC muncul di kolom PIC tabel (td ke-4)"],

    ["TC-KMR-EDT-003", "Happy",
     "Edit lokasi berhasil → kolom Lokasi ter-update di list",
     "Login admin; ada kamar tanpa lokasi",
     "1) Seed kamar tanpa lokasi\n2) Buka Edit\n3) Isi Lokasi\n4) Klik \"Simpan\"",
     "Lokasi=LOK<ts><seq>",
     "Toast sukses; modal tutup; kolom Lokasi di baris kamar menampilkan lokasi baru",
     "PRD", "Belum dijalankan",
     "Verifikasi kolom Lokasi (td ke-3) ter-update"],

    ["TC-KMR-EDT-004", "Happy",
     "Edit status Aktif → Tidak Aktif → badge di list berubah",
     "Login admin; ada kamar berstatus Aktif",
     "1) Seed kamar (default Aktif)\n2) Buka Edit\n3) Ubah Status ke \"Tidak Aktif\"\n4) Klik \"Simpan\"",
     "Status awal=Aktif; Status baru=Tidak Aktif",
     "Toast sukses; modal tutup; badge status kamar di list berubah menjadi \"Tidak Aktif\"",
     "PRD", "Belum dijalankan",
     "Verifikasi badge (data-slot=badge) di baris kamar setelah save"],

    ["TC-KMR-EDT-005", "Happy",
     "Edit status Tidak Aktif → Aktif → badge di list berubah",
     "Login admin; ada kamar berstatus Tidak Aktif di staging (dikonfirmasi ada)",
     "1) Cari kamar berstatus Tidak Aktif (search nama)\n2) Buka Edit\n3) Ubah Status ke \"Aktif\"\n4) Klik \"Simpan\"",
     "Status awal=Tidak Aktif; Status baru=Aktif",
     "Toast sukses; modal tutup; badge status kamar di list berubah menjadi \"Aktif\"",
     "PRD", "Belum dijalankan",
     "Pre-condition bergantung data Tidak Aktif yang ada di staging. Bisa dibuat via TC-004 jika perlu isolation"],

    # ---------------- POSITIF ----------------
    ["TC-KMR-EDT-006", "Positif",
     "Modal Edit terbuka dengan data pre-populated sesuai data kamar yang diedit",
     "Login admin; ada kamar lengkap (Instansi + Nama + PIC + Lokasi + Status)",
     "1) Seed kamar lengkap\n2) Klik ikon Edit\n3) Amati isi semua field di modal",
     "Kamar dengan semua field terisi",
     "Modal Edit terbuka; semua field terisi sesuai data kamar: Instansi, Nama, PIC, Lokasi, dan Status ter-pre-populate dengan benar",
     "PRD", "Belum dijalankan",
     "Pastikan nilai di setiap field input/select match dengan data yang ada di baris list"],

    ["TC-KMR-EDT-007", "Positif",
     "Ganti Instansi di modal Edit → dropdown PIC ter-reset ke placeholder [Asumsi]",
     "Login admin; ada minimal 2 instansi berbeda di sistem",
     "1) Seed kamar di Instansi A dengan PIC\n2) Buka Edit\n3) Ganti Instansi ke Instansi B\n4) Amati dropdown PIC",
     "Instansi A=SDIT; Instansi B=Sekolah Alam",
     "Dropdown PIC ter-reset ke placeholder \"Tidak ada PIC\"; opsi PIC berubah ke daftar guru/staff Instansi B",
     "Asumsi", "Belum dijalankan",
     "PRD: PIC = guru & staff instansi terpilih. Reset PIC saat ganti instansi = asumsi (dikonfirmasi sama seperti perilaku modal Tambah)"],

    ["TC-KMR-EDT-008", "Positif",
     "Hapus PIC (set ke tidak ada) → tersimpan, kolom PIC kosong di list",
     "Login admin; ada kamar dengan PIC terisi",
     "1) Seed kamar + PIC\n2) Buka Edit\n3) Ubah PIC ke \"Tidak ada PIC\" (placeholder)\n4) Klik \"Simpan\"",
     "PIC dihapus/dikosongkan",
     "Toast sukses; modal tutup; kolom PIC di list kosong / tidak menampilkan nama",
     "PRD", "Belum dijalankan",
     "PIC optional -> bisa di-clear ke kosong. Perlu konfirmasi apakah dropdown PIC punya opsi kosong/placeholder"],

    ["TC-KMR-EDT-009", "Positif",
     "Hapus Lokasi (clear) → tersimpan, kolom Lokasi kosong di list",
     "Login admin; ada kamar dengan Lokasi terisi",
     "1) Seed kamar + Lokasi\n2) Buka Edit\n3) Clear field Lokasi\n4) Klik \"Simpan\"",
     "Lokasi dihapus (field dikosongkan)",
     "Toast sukses; modal tutup; kolom Lokasi di list kosong",
     "PRD", "Belum dijalankan",
     "Lokasi optional -> boleh dikosongkan. Verifikasi td ke-3 kosong setelah save"],

    ["TC-KMR-EDT-010", "Positif",
     "Klik Simpan tanpa mengubah apapun → data tetap sama",
     "Login admin; ada kamar di list",
     "1) Seed kamar\n2) Buka Edit\n3) Jangan ubah field apapun\n4) Klik \"Simpan\"",
     "Semua field tidak diubah",
     "Toast sukses; modal tutup; data kamar di list tetap sama persis (nama, PIC, lokasi, status tidak berubah)",
     "PRD", "Belum dijalankan",
     "Simpan tanpa perubahan harus tetap valid (PRD tidak sebut kondisi 'wajib ada perubahan')"],

    ["TC-KMR-EDT-011", "Positif",
     "Klik Batal → modal tutup, data kamar tidak berubah",
     "Login admin; ada kamar di list",
     "1) Seed kamar\n2) Buka Edit\n3) Ubah Nama Kamar (jangan Simpan)\n4) Klik \"Batal\"",
     "Nama baru diketik tapi tidak disimpan",
     "Modal tertutup; data kamar di list tetap menampilkan nama awal (perubahan di-discard)",
     "PRD", "Belum dijalankan",
     "PRD: 'Klik Batal → kembali ke halaman list kamar'. Perubahan unsaved harus di-discard"],

    # ---------------- NEGATIF ----------------
    ["TC-KMR-EDT-012", "Negatif",
     "Submit tanpa Instansi → error validasi 'Instansi wajib diisi'",
     "Login admin; modal Edit Kamar terbuka",
     "1) Buka Edit kamar manapun\n2) Clear / kosongkan field Instansi\n3) Klik \"Simpan\"",
     "Instansi dikosongkan",
     "Muncul pesan error di field Instansi (misal 'Instansi wajib diisi'); data tidak tersimpan",
     "PRD", "DITUNDA",
     "Element analysis: select Instansi di modal Edit tidak punya opsi kosong -> UI MENCEGAH pengosongan. TC tidak bisa dieksekusi. Validasi ini sudah dicover di TC Tambah (TC-KMR-ADD-006). Spec: it.skip"],

    ["TC-KMR-EDT-013", "Negatif",
     "Submit dengan Nama Kamar kosong → error validasi 'Kamar wajib diisi'",
     "Login admin; modal Edit Kamar terbuka",
     "1) Buka Edit kamar manapun\n2) Clear field Nama Kamar\n3) Klik \"Simpan\"",
     "Nama Kamar dikosongkan",
     "Muncul pesan error di field Nama Kamar (misal 'Kamar wajib diisi'); data tidak tersimpan",
     "PRD", "Belum dijalankan",
     "Nama Kamar* = required. Verifikasi form-message muncul & modal tetap terbuka"],

    ["TC-KMR-EDT-014", "Negatif",
     "Edit nama kamar jadi duplikat (nama + instansi sama) → sistem menampilkan pesan error [BUG-019]",
     "Login admin; ada minimal 2 kamar di instansi yang sama",
     "1) Seed 2 kamar di instansi yang sama (Nama A & Nama B)\n2) Buka Edit Nama B\n3) Ubah nama menjadi Nama A\n4) Klik \"Simpan\"\n5) Monitor UI & Network (XHR)",
     "Nama B diubah menjadi Nama A (sudah ada di instansi yang sama)",
     "Muncul toast/inline error berisi pesan dari API; data tidak tersimpan (Nama B tetap di list)",
     "PRD", "FAIL",
     "BUG-019: FE silent — backend balik {\"message\":\"Nama kamar sudah ada di instansi ini\"} tapi FE tidak tampilkan apapun. Pola BUG-008/010. Dikonfirmasi manual saat element analysis. assertNotSilent SENGAJA fail sampai fixed"],

    # ---------------- EDGE ----------------
    ["TC-KMR-EDT-015", "Edge",
     "Edit nama kamar dengan input >255 karakter → FE wajib beri feedback (tidak boleh silent) [Kandidat pola BUG-015]",
     "Login admin; ada kamar di list",
     "1) Buka Edit kamar manapun\n2) Clear Nama Kamar lalu isi ~300 karakter\n3) Klik \"Simpan\"\n4) Monitor UI & Network (XHR)",
     "Nama=300 karakter 'A' (overflow batas DB 255 char)",
     "FE membatasi input (maxlength) ATAU menampilkan pesan error ramah; tidak boleh silent. Backend tidak boleh balikkan raw SQL error ke client",
     "PRD", "Belum dijalankan",
     "Pola BUG-015 (Kamar Tambah) & BUG-013 (Jurusan Tambah): FE silent + raw SQL leak saat >255 char. Assert correct -> TC fail jika silent (cek assertNotSilent). Data tidak boleh corrupt"],
]

LEGEND = [
    ["Kolom Sumber", ""],
    ["PRD", "Skenario diturunkan langsung dari PRD Edit Kamar"],
    ["PRD-ambigu", "PRD menyinggung tapi tidak jelas/typo -> perlu konfirmasi"],
    ["Asumsi", "Tidak ada di PRD; asumsi berbasis konvensi/fixture/modul lain"],
    ["", ""],
    ["Kolom Kategori", ""],
    ["Happy", "Alur utama sukses sesuai desain"],
    ["Positif", "Variasi valid non-utama"],
    ["Negatif", "Input/kondisi yang harus ditolak / mengembalikan error"],
    ["Edge", "Batas/kondisi tepi"],
    ["", ""],
    ["Kolom Status", ""],
    ["Belum dijalankan", "Sudah di-generate, menunggu eksekusi Cypress"],
    ["FAIL", "Gagal saat run; lihat ref BUG-### di kolom Catatan"],
    ["", ""],
    ["Skenario SKIP (tidak ada di sheet ini)", ""],
    ["SKIP", "Downstream Status Tidak Aktif (hidden di data diri siswa & presensi) — fitur terkait belum bisa diuji"],
    ["", ""],
    ["Kandidat Bug", ""],
    ["BUG-019", "Edit Kamar duplikat nama (instansi sama) -> FE silent, backend error tidak ditampilkan. Pola BUG-008/010. Lihat TC-014"],
    ["BUG-015 pola", "Tambah Kamar >255 char silent (BUG-015). TC-015 assert correct behavior -> fail jika Edit Kamar >255 char juga silent"],
    ["", ""],
    ["Catatan", "Tanpa kategori security/injection (sesuai kesepakatan)"],
    ["", "Bug ditandai inline: Status=FAIL + ref BUG-### di kolom Catatan"],
    ["", "Field PRD: Instansi* + Nama Kamar* + Status* (required), PIC + Lokasi (optional)"],
    ["", "PIC = dropdown guru & staff pada instansi yang dipilih; reset saat instansi diganti"],
]

# ---- styling ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CAT_FILL = {
    "Happy":   PatternFill("solid", fgColor="E2EFDA"),
    "Positif": PatternFill("solid", fgColor="DDEBF7"),
    "Negatif": PatternFill("solid", fgColor="FCE4D6"),
    "Edge":    PatternFill("solid", fgColor="FFF2CC"),
}
SRC_FONT = {
    "PRD":        Font(color="375623"),
    "PRD-ambigu": Font(color="BF8F00", bold=True),
    "Asumsi":     Font(color="C00000", bold=True),
}
STATUS_FILL = {
    "FAIL":    PatternFill("solid", fgColor="FF0000"),
    "DITUNDA": PatternFill("solid", fgColor="D9D9D9"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPWRAP = Alignment(vertical="top", wrap_text=True)
WIDTHS = [16, 10, 40, 34, 46, 30, 52, 12, 16, 48]

wb = Workbook()
ws = wb.active
ws.title = "Edit Kamar"

# Title row
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
t = ws.cell(row=1, column=1,
            value="Test Case - Edit Kamar (CARDS School v3)  |  Modul: Pengaturan Akademik > Kamar  |  Sumber PRD: Edit Kamar")
t.font = Font(bold=True, size=12)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22

# Header row
for c, name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=2, column=c, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = BORDER

# Data rows
for r, row in enumerate(ROWS, start=3):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = TOPWRAP
        cell.border = BORDER
    cat = row[1]
    if cat in CAT_FILL:
        ws.cell(row=r, column=2).fill = CAT_FILL[cat]
    src = row[7]
    if src in SRC_FONT:
        ws.cell(row=r, column=8).font = SRC_FONT[src]
    status = row[8]
    if status in STATUS_FILL:
        ws.cell(row=r, column=9).fill = STATUS_FILL[status]

# Column widths
for c, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + len(ROWS)}"

# Legend sheet
ls = wb.create_sheet("Legenda")
ls.column_dimensions["A"].width = 26
ls.column_dimensions["B"].width = 72
for r, (a, b) in enumerate(LEGEND, 1):
    ca = ls.cell(row=r, column=1, value=a)
    cb = ls.cell(row=r, column=2, value=b)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    if a in ("Kolom Sumber", "Kolom Kategori", "Kolom Status",
             "Skenario SKIP (tidak ada di sheet ini)", "Kandidat Bug", "Catatan"):
        ca.font = Font(bold=True)

out = "docs/test-cases/TC_Kamar_Edit.xlsx"
wb.save(out)
print("saved:", out, "| rows:", len(ROWS))
