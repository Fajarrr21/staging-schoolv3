# -*- coding: utf-8 -*-
"""Generate Test Case sheet for List Tag (per CLAUDE.md columns)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = ["ID", "Kategori", "Skenario", "Pre-condition", "Langkah",
           "Test Data", "Expected", "Sumber", "Status", "Catatan"]

# PRD List Tag:
#   Kolom: Instansi | Nama Tag | Kode Tag | Tipe Member | Status | Dibuat Pada | Aksi (Edit & Hapus) = 8 sel/baris
#   Sort default: waktu pembuatan terbaru -> terlama
#   Default status saat Tambah = Aktif
#   Admin dapat ubah status & data via halaman Edit (modul Edit, di-cover spec terpisah)
#   Filter: Instansi ; Status (Semua/Aktif/Tidak Aktif) ; Tipe Member (Semua/Siswa/Guru&Staff)
#   Search: Instansi, Nama Tag, Kode Tag (Tipe Member & Status BUKAN kriteria search)
#     CATATAN: search by Instansi TIDAK dibuatkan TC karena sudah dicover oleh filter Instansi
#              (redundansi -- per keputusan user)
#   Empty: no-data / no-filter-result / no-search-result -> halaman kosong
#
# Pre-existing bug-pola yang sengaja dicari kemunculannya di Tag:
#   - BUG-016 (Kamar): filter Status tunggal balik empty padahal data ada -> TC-007 [TIDAK ter-reproduce di Tag]
#   - BUG-009/014: filter kombinasi Status+filter lain balik empty -> TC-013/015 [TER-REPRODUCE -> BUG-023]
#
# PRD validasi (e) "Tipe member ... dapat digunakan utk tipe member siswa atau guru dan staff":
#   asosiasi cross-fitur -> DITUNDA (butuh modul Data Diri Siswa / Data Guru-Staff yg belum ada)
ROWS = [
    # ---------------- HAPPY ----------------
    ["TC-TAG-LST-001", "Happy",
     "List menampilkan kolom lengkap sesuai PRD",
     "Login admin; ada >=1 data tag (seed via spec Tambah)",
     "1) Buka Pengaturan > Akademik > Tag\n2) Amati header & isi tabel",
     "Seed 1 tag (Instansi=SDIT, Tipe=Semua, Status=Aktif)",
     "Tampil 8 sel/baris: Instansi, Nama Tag, Kode Tag, Tipe Member (badge), Status (badge), Dibuat Pada, Aksi Edit, Aksi Hapus",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-LST-002", "Happy",
     "Setiap baris punya aksi Edit & Hapus",
     "Ada >=1 data tag",
     "1) Buka List Tag\n2) Cek ikon aksi pada baris pertama",
     "-",
     "Tiap baris menampilkan ikon Edit (pensil/lucide-square-pen) & Hapus (trash/lucide-trash)",
     "PRD", "Belum dijalankan", ""],

    # ---------------- SORT & DEFAULT ----------------
    ["TC-TAG-LST-003", "Positif",
     "Urutan default terbaru -> terlama (by Dibuat Pada)",
     "Login admin",
     "1) Tambah tag baru (nama unik)\n2) Buka List Tag tanpa filter/search\n3) Amati baris teratas",
     "Instansi=SDIT; Nama=QA<ts><seq>",
     "Tag yang baru dibuat tampil di BARIS PALING ATAS (newest first)",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-LST-004", "Positif",
     "Default status tag baru = Aktif (badge Aktif di list)",
     "Login admin",
     "1) Tambah tag baru (tanpa pilih status di form -- PRD: default Aktif)\n2) Buka List Tag\n3) Cari nama tsb\n4) Amati badge Status",
     "Instansi=SDIT; Nama=QA<ts><seq>",
     "Badge Status pada baris tag baru = \"Aktif\" (PRD validasi c)",
     "PRD", "Belum dijalankan", ""],

    # ---------------- FILTER ----------------
    ["TC-TAG-LST-005", "Happy",
     "Filter by Instansi menampilkan hanya tag instansi tsb",
     "Ada tag di Instansi SDIT",
     "1) Seed tag SDIT\n2) Set filter Instansi=SDIT\n3) Amati semua baris",
     "Instansi=SDIT",
     "Semua baris ber-Instansi = SDIT; minimal seed muncul",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-LST-006", "Positif",
     "Ganti filter Instansi -> list ter-update",
     "Ada tag di SDIT & Sekolah Alam",
     "1) Seed tag SDIT & Sekolah Alam\n2) Filter Instansi=SDIT (amati)\n3) Ganti Instansi=Sekolah Alam (amati)",
     "Instansi A=SDIT; Instansi B=Sekolah Alam",
     "Setelah ganti, semua baris ber-Instansi = instansi yang dipilih terakhir",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-LST-007", "Positif",
     "Filter Status=Aktif tunggal -> hanya tag Aktif",
     "Ada tag berstatus Aktif (default tambah = Aktif)",
     "1) Seed tag (default Aktif)\n2) Filter Status=Aktif\n3) Amati badge status tiap baris",
     "Status=Aktif",
     "Semua baris berstatus Aktif; minimal seed muncul",
     "PRD", "Belum dijalankan",
     "KANDIDAT BUG-016 (Kamar): filter Status tunggal di modul lain balik empty padahal data ada. Cek apakah Tag juga kena pola yg sama."],

    ["TC-TAG-LST-008", "Positif",
     "Filter Status=Tidak Aktif tunggal -> hanya tag Tidak Aktif",
     "Ada tag berstatus Tidak Aktif (status diubah via Edit / manual)",
     "1) Filter Status=Tidak Aktif\n2) Amati list & Network (XHR)",
     "Status=Tidak Aktif",
     "Semua baris berstatus Tidak Aktif (data Tidak Aktif ADA di staging)",
     "PRD", "Belum dijalankan",
     "KANDIDAT BUG-016. Butuh >=1 tag Tidak Aktif (siapkan via Edit setelah modul Edit Tag siap)"],

    ["TC-TAG-LST-009", "Positif",
     "Filter Status=Semua -> tampil semua status",
     "Ada data tag",
     "1) Set filter Status=Aktif\n2) Ganti Status=Semua\n3) Amati list",
     "Status=Semua",
     "List menampilkan semua tag tanpa batasan status",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-LST-010", "Happy",
     "Filter Tipe Member=Siswa -> hanya tag Tipe=Siswa",
     "Ada tag Tipe=Siswa",
     "1) Seed tag Tipe=Siswa\n2) Filter Tipe Member=Siswa\n3) Amati badge Tipe tiap baris",
     "Tipe Member=Siswa",
     "Semua baris badge Tipe = Siswa; minimal seed muncul",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-LST-011", "Positif",
     "Filter Tipe Member=Guru & Staff -> hanya tag Tipe=Guru&Staff",
     "Ada tag Tipe=Guru & Staff",
     "1) Seed tag Tipe=Guru&Staff\n2) Filter Tipe Member=Guru & Staff\n3) Amati badge Tipe tiap baris",
     "Tipe Member=Guru & Staff",
     "Semua baris badge Tipe = Guru & Staff; minimal seed muncul",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-LST-012", "Positif",
     "Filter Tipe Member=Semua (opsi reset)",
     "Ada data tag berbagai tipe",
     "1) Filter Tipe Member=Siswa\n2) Ganti Tipe Member=Semua\n3) Amati list",
     "Tipe Member=Semua",
     "List menampilkan semua tag tanpa batas tipe (reset filter)",
     "Asumsi", "Belum dijalankan",
     "PRD tak eksplisit sebut opsi 'Semua' di filter Tipe, tapi konsisten dgn filter Status. Konfirmasi UI"],

    ["TC-TAG-LST-013", "Positif",
     "Filter kombinasi: Instansi + Status=Aktif",
     "Ada tag SDIT berstatus Aktif",
     "1) Seed tag SDIT (Aktif)\n2) Filter Instansi=SDIT\n3) Tambah filter Status=Aktif\n4) Amati list & Network (XHR)",
     "Instansi=SDIT; Status=Aktif",
     "Tampil tag yang cocok KEDUA filter (minimal seed)",
     "PRD", "FAIL",
     "BUG-023: kombinasi Instansi+Status balik empty padahal data ada (pola BUG-009/014 ter-reproduce di Tag). Status TUNGGAL & Instansi+Tipe (tanpa Status) tetap jalan; hanya Status di kombinasi yg rusak."],

    ["TC-TAG-LST-014", "Positif",
     "Filter kombinasi: Instansi + Tipe Member",
     "Ada tag SDIT Tipe=Siswa",
     "1) Seed tag SDIT Tipe=Siswa\n2) Filter Instansi=SDIT + Tipe Member=Siswa\n3) Amati list",
     "Instansi=SDIT; Tipe Member=Siswa",
     "Tampil tag yang cocok KEDUA filter",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-LST-015", "Positif",
     "Filter kombinasi 3: Instansi + Status + Tipe Member",
     "Ada tag SDIT Aktif Tipe=Semua",
     "1) Seed tag SDIT (default Aktif) Tipe=Semua\n2) Filter Instansi=SDIT + Status=Aktif + Tipe=Semua\n3) Amati list",
     "Instansi=SDIT; Status=Aktif; Tipe Member=Semua",
     "Tampil tag yang cocok KETIGA filter (minimal seed)",
     "PRD", "FAIL",
     "BUG-023: kombinasi 3 ikut kena karena Status digabung dgn filter lain (lihat TC-013)."],

    # ---------------- SEARCH ----------------
    ["TC-TAG-LST-016", "Happy",
     "Cari by Nama Tag",
     "Ada tag dgn nama unik",
     "1) Seed tag nama unik\n2) Ketik nama di Cari\n3) Amati hasil",
     "Nama=QA<ts><seq>",
     "List ter-filter menampilkan tag dgn nama tsb",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-LST-017", "Happy",
     "Cari by Kode Tag",
     "Ada tag dgn kode unik",
     "1) Seed tag kode unik\n2) Ketik kode di Cari\n3) Amati hasil",
     "Kode=QA<ts><seq>",
     "List ter-filter menampilkan tag dgn kode tsb",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-LST-018", "Positif",
     "Clear search -> list kembali tampil penuh",
     "Sudah melakukan search",
     "1) Seed tag\n2) Search nama tsb\n3) Kosongkan kolom Cari\n4) Amati list",
     "Nama=QA<ts><seq>",
     "Setelah dikosongkan, list kembali menampilkan banyak baris (tidak ter-filter)",
     "PRD", "Belum dijalankan", ""],

    # ---------------- EDGE & NEGATIF ----------------
    ["TC-TAG-LST-019", "Edge",
     "Cari case-insensitive (huruf kecil) menemukan data",
     "Ada tag dgn nama unik",
     "1) Seed tag nama unik (uppercase QA...)\n2) Ketik versi huruf kecil di Cari\n3) Amati hasil",
     "Nama=QA<ts><seq>; keyword=qa<ts><seq>",
     "Search tidak case-sensitive -> tag tetap ditemukan",
     "Asumsi", "Belum dijalankan",
     "PRD tak eksplisit soal case-insensitive -> asumsi konvensi modul lain"],

    # NOTE: TC-020 lama (Cari teks Tipe Member -> TIDAK menemukan, PRD strict) DIHAPUS.
    #       Alasan: staging punya data lama dgn Nama/Kode mengandung "Siswa" -> false positive
    #       (assertion PRD-strict tidak bisa dibedain antara bug & data pollution).
    #       Coverage search PRD sudah cukup di TC-016/017/019.

    ["TC-TAG-LST-020", "Negatif",
     "Cari keyword tidak ada -> halaman kosong",
     "List Tag terbuka",
     "1) Ketik keyword acak yang pasti tak match\n2) Amati hasil",
     "Keyword=ZZZQA000NOTEXIST",
     "Tidak ada baris; tampil halaman kosong (empty state \"Data Tag tidak ditemukan\")",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-LST-021", "Negatif",
     "Filter kombinasi tanpa data cocok -> halaman kosong",
     "List Tag terbuka",
     "1) Pilih kombinasi Instansi+Status+Tipe yang pasti tidak punya data (mis. Instansi Sekolah Alam + Status Tidak Aktif + Tipe=Guru&Staff)\n2) Amati hasil",
     "Filter tanpa hasil",
     "Tampil halaman kosong (empty state); list tidak crash",
     "PRD", "Belum dijalankan",
     "Pastikan kombinasi yang dipilih memang tak ada data (perlu cek manual di staging)"],

    # ---------------- DEFERRED (asosiasi Member) ----------------
    ["TC-TAG-LST-022", "Positif",
     "Tag Tipe Member=Siswa dipakai di picker Tag pada Data Diri Siswa",
     "Tag QA Tipe=Siswa sudah dibuat & list menampilkan badge Tipe=Siswa",
     "1) Buka form Data Diri Siswa\n2) Buka picker Tag\n3) Cek tag QA muncul (sesuai PRD validasi e)",
     "Tag QA Tipe=Siswa",
     "Tag muncul & dapat dipilih pada form Data Diri Siswa, TIDAK muncul di picker Data Guru/Staff",
     "PRD", "DITUNDA",
     "DEFERRED per instruksi user: asosiasi ke Data Siswa belum bisa diakses. Aktifkan setelah modul Member/Data Siswa tersedia"],

    ["TC-TAG-LST-023", "Positif",
     "Tag Tipe Member=Guru & Staff dipakai di picker Tag pada Data Guru/Staff",
     "Tag QA Tipe=Guru&Staff sudah dibuat",
     "1) Buka form Data Guru/Staff\n2) Buka picker Tag\n3) Cek tag QA muncul (sesuai PRD validasi e)",
     "Tag QA Tipe=Guru & Staff",
     "Tag muncul & dapat dipilih pada form Data Guru/Staff, TIDAK muncul di picker Data Siswa",
     "PRD", "DITUNDA",
     "DEFERRED per instruksi user: asosiasi ke Data Guru/Staff belum bisa diakses"],

    ["TC-TAG-LST-024", "Positif",
     "Tag Tipe Member=Semua dipakai di semua picker (Siswa + Guru/Staff)",
     "Tag QA Tipe=Semua sudah dibuat",
     "1) Cek picker Tag di form Data Siswa -> tag QA muncul\n2) Cek picker Tag di form Data Guru/Staff -> tag QA muncul",
     "Tag QA Tipe=Semua",
     "Tag muncul di kedua picker member",
     "PRD", "DITUNDA",
     "DEFERRED per instruksi user: bergantung pada fitur Member yang belum tersedia"],
]

LEGEND = [
    ["Kolom Sumber", ""],
    ["PRD", "Skenario diturunkan langsung dari PRD List Tag"],
    ["PRD-ambigu", "PRD menyinggung tapi tidak jelas -> perlu konfirmasi"],
    ["Asumsi", "Tidak ada di PRD; asumsi berbasis konvensi/fixture/modul lain"],
    ["", ""],
    ["Kolom Kategori", ""],
    ["Happy", "Alur utama sukses sesuai desain"],
    ["Positif", "Variasi valid non-utama"],
    ["Negatif", "Input/kondisi yang harus mengembalikan kosong / tidak menemukan"],
    ["Edge", "Batas/kondisi tepi"],
    ["", ""],
    ["Kolom Status", ""],
    ["Belum dijalankan", "Sudah di-generate, menunggu eksekusi Cypress"],
    ["DITUNDA", "Sengaja belum di-run; menunggu fitur konsumen (Data Siswa / Data Guru-Staff / Member) tersedia"],
    ["FAIL", "Gagal saat run; lihat ref BUG-### di Catatan"],
    ["", ""],
    ["Bug ter-konfirmasi setelah run", ""],
    ["BUG-023 (Tag List/Filter)", "Status di kombinasi (Instansi+Status, Instansi+Status+Tipe) balik empty -> TC-013 & TC-015 FAIL. Pola BUG-009/014. Status TUNGGAL & Instansi+Tipe tetap jalan."],
    ["BUG-016 (Kamar)", "TIDAK reproduce di Tag: filter Status TUNGGAL di Tag (TC-007) tetap jalan."],
    ["", ""],
    ["Catatan", "Tanpa kategori security/injection (sesuai kesepakatan)"],
    ["", "Field PRD list: Instansi | Nama Tag | Kode Tag | Tipe Member | Status | Dibuat Pada | Aksi (Edit/Hapus) = 8 sel/baris"],
    ["", "Filter PRD: Instansi, Status (Semua/Aktif/Tidak Aktif), Tipe Member"],
    ["", "Search PRD: Instansi, Nama Tag, Kode Tag -> SEARCH BY INSTANSI TIDAK dibuat TC (sudah dicover filter Instansi, per keputusan user)"],
    ["", "Empty state: tidak ada data / filter tanpa hasil / search tanpa hasil"],
]

# ---- styling ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CAT_FILL = {
    "Happy":   PatternFill("solid", fgColor="E2EFDA"),
    "Positif": PatternFill("solid", fgColor="DDEBF7"),
    "Negatif": PatternFill("solid", fgColor="FCE4D6"),
    "Edge":    PatternFill("solid", fgColor="FFF2CC"),
}
SRC_FONT = {
    "PRD":        Font(color="375623"),
    "PRD-ambigu": Font(color="BF8F00", bold=True),
    "Asumsi":     Font(color="C00000", bold=True),
}
STATUS_FILL = {
    "DITUNDA": PatternFill("solid", fgColor="D9D9D9"),
    "FAIL":    PatternFill("solid", fgColor="FFC7CE"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPWRAP = Alignment(vertical="top", wrap_text=True)
WIDTHS = [16, 10, 38, 30, 44, 28, 50, 12, 16, 46]

wb = Workbook()
ws = wb.active
ws.title = "List Tag"

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
t = ws.cell(row=1, column=1, value="Test Case - List Tag (CARDS School v3)  |  Modul: Pengaturan Akademik > Tag  |  Sumber PRD: List Tag")
t.font = Font(bold=True, size=12)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22

for c, name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=2, column=c, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = BORDER

for r, row in enumerate(ROWS, start=3):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = TOPWRAP
        cell.border = BORDER
    cat = row[1]
    if cat in CAT_FILL:
        ws.cell(row=r, column=2).fill = CAT_FILL[cat]
    src = row[7]
    if src in SRC_FONT:
        ws.cell(row=r, column=8).font = SRC_FONT[src]
    status = row[8]
    if status in STATUS_FILL:
        ws.cell(row=r, column=9).fill = STATUS_FILL[status]

for c, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + len(ROWS)}"

ls = wb.create_sheet("Legenda")
ls.column_dimensions["A"].width = 28
ls.column_dimensions["B"].width = 72
for r, (a, b) in enumerate(LEGEND, 1):
    ca = ls.cell(row=r, column=1, value=a)
    cb = ls.cell(row=r, column=2, value=b)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    if a in ("Kolom Sumber", "Kolom Kategori", "Kolom Status",
             "Kandidat bug-pola (cek saat run)", "Catatan"):
        ca.font = Font(bold=True)

out = "docs/test-cases/TC_Tag_List.xlsx"
wb.save(out)
print("saved:", out, "| rows:", len(ROWS))
