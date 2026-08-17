# -*- coding: utf-8 -*-
"""Generate Test Case sheet for Tambah Jadwal Pelajaran (per CLAUDE.md columns)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = ["ID", "Kategori", "Skenario", "Pre-condition", "Langkah",
           "Test Data", "Expected", "Sumber", "Status", "Catatan"]

# ============================================================================
# SCOPE: TC file ini fokus di FASE 1 (Modal Tambah Jadwal Pelajaran + redirect).
#
# STATUS APP vs REFACTOR SPEC:
#   - Staging saat ini = PRE-REFACTOR. Label tombol modal = "Simpan"; klik
#     "Simpan" LANGSUNG commit record ke DB -> potensi ghost record.
#   - Refactor spec dari PO: modal button HANYA navigasi + bawa data ke
#     halaman edit; commit sebenarnya di Simpan halaman edit (fase 2).
#   - Konvensi CLAUDE.md: TC assert BEHAVIOR YG BENAR (post-refactor).
#     TC akan FAIL sampai bug fixed. Bug tracking di bugs.csv:
#       BUG-032: Tag picker layout overlap
#       BUG-033: Toast duplikasi pakai warning (harusnya destructive)
#       BUG-034: Tag tidak reset saat ganti Instansi
#       BUG-035: Tag picker prematur aktif + tidak scoped per Instansi
#       BUG-036: Klik Simpan modal langsung commit ke DB (ghost record risk)
#
# FLOW aktual di staging:
#   1) List Jadwal Pelajaran -> klik "Tambah Jadwal Pelajaran"
#   2) MODAL "Tambah Jadwal Pelajaran" tampil, form dgn field:
#      - Tahun Ajaran (readonly text, mis. "2026/2027")
#      - Semester (readonly badge, mis. "GANJIL")
#      - Instansi* (Select Radix, placeholder "Pilih Instansi")
#      - Tingkat* (Select Radix, DISABLED sampai Instansi dipilih)
#      - Multi Kelas (Switch toggle, default OFF)
#      - Kelas* (Select Radix, DISABLED sampai Tingkat dipilih;
#                single-select saat Multi OFF; multi-select saat Multi ON)
#      - Jurusan (Opsional; Select Radix, DISABLED sampai Kelas dipilih;
#                hide saat Multi Kelas ON)
#      - Tag (Opsional; inline dropdown checkbox multi-select, no search)
#      - Btn Batal + Btn "Simpan" (label aktual) + X (dialog-close)
#   3) Klik "Simpan" -> validasi FE (termasuk cek duplikasi Instansi+
#      Tingkat+Kelas+TA+Semester dgn toast warning) -> REDIRECT ke:
#      /setting/academic/course-schedule/{id} (mis. /90)
#      AKTUAL: record LANGSUNG masuk DB di titik ini (BUG-036).
#      BENAR: record TIDAK masuk di titik ini; commit di halaman edit.
#   4) Halaman edit: section Data Jadwal Pelajaran auto-fill (read-only);
#      user set jadwal per hari + toggle jam pelajaran + Simpan final
#      (INI FASE 2 - di TC file terpisah nanti)
#
# Verified selectors (dari element analysis):
#   - Modal: [data-slot="dialog-content"] + role="dialog"
#   - Btn tambah (dialog-trigger): svg.lucide-square-plus + text "Tambah Jadwal Pelajaran"
#   - Select trigger: [data-slot="select-trigger"] scoped by formItem
#   - Select content: [data-slot="select-content"] -> DIPORTAL ke <body> (di luar modal)
#     -> JANGAN .within('[data-slot="dialog-content"]'), query global.
#   - Select item: [data-slot="select-item"] role="option"
#   - Switch Multi Kelas: [role="switch"] name="combined_schedule"
#   - Tag checkbox item: <button role="checkbox" data-slot="checkbox">
#     -> TOGGLE via klik label ([data-slot="label"]), BUKAN cy.check()
#     -> pakai regex exact /^tagname$/ (hindari substring collision QA...)
#   - Chip terpilih: badge di atas field Tag
#   - Btn utama footer: text "Simpan" (aktual staging; refactor spec = "Lanjutkan")
#   - Btn Batal: text "Batal"; X close: [data-slot="dialog-close"] pojok kanan atas
#   - Form message error: [data-slot="form-message"] class text-destructive
#   - Toast Sonner: [data-sonner-toast][data-type="success|warning|error"]
#     - Success: title "Jadwal pelajaran berhasil ditambahkan"
#     - Warning duplikasi: text "Jadwal pelajaran sudah terdaftar!" (BUG-033)
#   - Tabel list: table[data-slot="data-grid-table"] > tbody > tr
#     - Kolom aktual (10): Instansi | Tahun Ajaran | Semester | Tingkat |
#       Kelas | Jurusan | Tag | Status | Dibuat Pada | Aksi
#     - Multi Kelas ON -> 1 row (chip pertama + badge N+ dalam 1 <td>)
#
# Konvensi Bahasa Indonesia (CLAUDE.md). Kategori: Happy / Positif / Negatif / Edge.
# ============================================================================

ROWS = [
    # ================= HAPPY =================
    ["TC-JP-ADD-001", "Happy",
     "Modal minimum (Instansi+Tingkat+Kelas, Multi OFF, tanpa Jurusan & Tag) -> redirect ke halaman edit; list belum bertambah entry",
     "Login admin; halaman List Jadwal Pelajaran terbuka; ada Instansi+Tingkat+Kelas yg belum punya jadwal di Tahun Ajaran+Semester aktif",
     "1) Catat jumlah baris tbody tr (baseline)\n2) Klik btn \"Tambah Jadwal Pelajaran\"\n3) Modal tampil\n4) Pilih Instansi\n5) Pilih Tingkat (setelah enable)\n6) Pilih Kelas (setelah enable)\n7) Biarkan Jurusan & Tag kosong; Multi Kelas OFF\n8) Klik \"Simpan\"\n9) Amati URL & halaman berikutnya\n10) Kembali ke List (back / navigasi menu) -> cek jumlah baris tbody tr",
     "Instansi=<valid>; Tingkat=<valid>; Kelas=<valid>; Jurusan=(kosong); Tag=(kosong); Multi Kelas=OFF",
     "Modal tertutup; redirect ke URL `/setting/academic/course-schedule/{id}`; halaman edit terbuka; section Data Jadwal Pelajaran auto-fill sesuai input; Tahun Ajaran+Semester = periode aktif. PENTING (per refactor spec): jumlah baris list = baseline (belum bertambah) - record baru muncul setelah Simpan final di halaman edit",
     "PRD", "Belum dijalankan",
     "AKTUAL: list langsung bertambah 1 entry di titik ini -> BUG-036 (klik Simpan modal langsung commit DB). TC akan FAIL sampai BUG-036 fixed. Konfirmasi format id di URL"],

    ["TC-JP-ADD-002", "Happy",
     "Modal lengkap dgn Jurusan + Tag (Multi Kelas OFF) -> redirect; list belum bertambah",
     "Modal terbuka; instansi+tingkat+kelas+jurusan valid tersedia; ada tag terdaftar utk Instansi tsb",
     "1) Catat baseline jumlah baris list\n2) Klik \"Tambah\"\n3) Pilih Instansi -> Tingkat -> Kelas\n4) Pilih Jurusan\n5) Klik picker Tag; pilih 1 tag via klik label (regex exact); verifikasi chip muncul\n6) Klik \"Simpan\"\n7) Amati URL & kembali ke list -> cek jumlah baris",
     "Instansi=<valid>; Tingkat=<valid>; Kelas=<valid>; Jurusan=<valid>; Tag=[<tag>]; Multi Kelas=OFF",
     "Modal tertutup; redirect ke halaman edit; section Data Jadwal Pelajaran auto-fill Instansi/Tingkat/Kelas/Jurusan sesuai input; chip tag terlihat; jumlah baris list = baseline",
     "PRD", "Belum dijalankan",
     "AKTUAL: list langsung bertambah -> BUG-036. Tag picker: klik label bukan cy.check(); regex exact"],

    ["TC-JP-ADD-003", "Happy",
     "Modal Multi Kelas ON (min 2 kelas, Jurusan hide) -> redirect; list belum bertambah",
     "Modal terbuka; ada 2+ kelas di tingkat yg sama, keduanya belum punya jadwal periode berjalan",
     "1) Catat baseline jumlah baris list\n2) Klik \"Tambah\"\n3) Pilih Instansi + Tingkat\n4) Aktifkan Multi Kelas\n5) Verifikasi field Jurusan ter-hide dari modal\n6) Verifikasi dropdown Kelas jadi multi-select\n7) Pilih 2 Kelas\n8) Klik \"Simpan\"\n9) Kembali ke list -> cek jumlah baris",
     "Instansi=<valid>; Tingkat=<valid>; Kelas=[Kelas A, Kelas B]; Multi Kelas=ON",
     "Modal tertutup; redirect ke halaman edit; section Data Jadwal Pelajaran menampilkan kedua kelas (chip pertama + N+); field Jurusan tetap hide di halaman edit; jumlah baris list = baseline. Multi Kelas ON di list = 1 ENTRY GABUNGAN (chip pertama + badge N+ dalam 1 row), bukan 2 row terpisah",
     "PRD", "Belum dijalankan",
     "AKTUAL: list langsung bertambah -> BUG-036. Assert render list: 1 row dgn chip pertama + counter N+ (bukan 2 row)"],

    # ================= POSITIF =================
    ["TC-JP-ADD-004", "Positif",
     "Form default state saat modal Tambah pertama kali dibuka",
     "Login admin; List Jadwal Pelajaran terbuka",
     "1) Klik btn \"Tambah Jadwal Pelajaran\"\n2) Amati kondisi awal modal (tanpa ubah apapun)",
     "-",
     "Modal \"Tambah Jadwal Pelajaran\" tampil; Tahun Ajaran preselect readonly text; Semester preselect badge readonly; Instansi placeholder \"Pilih Instansi\"; Tingkat placeholder + DISABLED; Kelas placeholder + DISABLED; Jurusan (Opsional) placeholder + DISABLED; Tag (Opsional) placeholder \"Pilih Tag\"; Multi Kelas toggle OFF; btn utama footer bertuliskan \"Simpan\" + btn Batal + X",
     "PRD", "Belum dijalankan",
     "Label footer button aktual = \"Simpan\". Refactor spec ke depan mengusulkan \"Lanjutkan\" - kalau/saat di-deploy, update TC ini"],

    ["TC-JP-ADD-005", "Positif",
     "Dependency chain: Instansi -> Tingkat -> Kelas -> Jurusan enable berurutan; ganti Instansi -> child RESET",
     "Modal terbuka; state default (Tingkat/Kelas/Jurusan disabled)",
     "1) Amati Tingkat/Kelas/Jurusan disabled\n2) Pilih Instansi\n3) Amati Tingkat enable; Kelas/Jurusan masih disabled\n4) Pilih Tingkat\n5) Amati Kelas enable; Jurusan masih disabled\n6) Pilih Kelas\n7) Amati Jurusan enable\n8) Ubah Instansi ke nilai lain -> amati Tingkat/Kelas/Jurusan RESET ke placeholder",
     "-",
     "Chain berjalan: Instansi -> Tingkat enable; Tingkat -> Kelas enable; Kelas -> Jurusan enable. Ubah Instansi -> Tingkat/Kelas/Jurusan reset ke placeholder (verified manual)",
     "PRD", "Belum dijalankan",
     "Verified via manual test. Reset Tag saat ganti Instansi = TIDAK reset (BUG-034) - lihat TC-018"],

    ["TC-JP-ADD-006", "Positif",
     "Multi Kelas toggle ON/OFF -> Jurusan hide/show; Kelas single/multi select",
     "Modal terbuka; sudah pilih Instansi + Tingkat",
     "1) Amati field Jurusan visible + Kelas single-select (Multi OFF)\n2) Aktifkan toggle Multi Kelas\n3) Amati field Jurusan ter-hide dari modal\n4) Amati dropdown Kelas berubah jadi multi-select (chip / checklist)\n5) Pilih 2 kelas\n6) Nonaktifkan Multi Kelas\n7) Amati Jurusan kembali visible; Kelas kembali single-select",
     "-",
     "Multi ON: Jurusan hide, Kelas multi-select; Multi OFF: Jurusan visible + enabled, Kelas single-select; transisi tidak error",
     "PRD", "Belum dijalankan",
     "Behavior reset Kelas saat toggle Multi Kelas -> catat aktual"],

    ["TC-JP-ADD-007", "Positif",
     "Pilih Tag (opsional) - inline dropdown checkbox multi-select; chip terpilih render di atas field",
     "Modal terbuka; Instansi/Tingkat/Kelas terisi; ada tag terdaftar (mis. \"mobil\")",
     "1) Klik button picker \"Pilih Tag\"\n2) Amati dropdown INLINE (bukan popup portal) terbuka dgn daftar checkbox 2 kolom + tombol \"+\" tambah tag baru\n3) Pilih 1 tag via klik LABEL (regex exact /^mobil$/); verifikasi state <button role=\"checkbox\"> jadi data-state=\"checked\"\n4) Amati chip \"mobil\" muncul di atas field Tag\n5) Pilih 1 tag lain -> chip berjejer\n6) Klik label tag terpilih lagi -> state kembali unchecked; chip hilang",
     "Tag=[\"mobil\"]",
     "Dropdown inline (bukan modal/portal); item pakai Radix Checkbox (klik label, bukan input tersembunyi); chip render di atas field; multi-select via chip berjejer; toggle via klik label kembali",
     "PRD", "Belum dijalankan",
     "TERKAIT BUG-032 (chip nabrak checkbox saat banyak dipilih). Automation: klik LABEL bukan cy.check(); regex exact /^...$/ (bloat data QA...)"],

    ["TC-JP-ADD-008", "Positif",
     "Batal / X close modal tanpa klik Simpan -> tidak redirect, list tidak berubah",
     "Modal terbuka; sudah isi Instansi + Tingkat + Kelas",
     "Case 1: klik btn \"Batal\" di footer modal\nCase 2: klik ikon X di pojok kanan atas modal\nCase 3: tekan Esc\nAmati modal & list setelah tiap case",
     "-",
     "Modal tertutup; tetap di halaman List Jadwal Pelajaran (tidak redirect); jumlah baris list tidak berubah; modal open ulang -> form reset ke default",
     "PRD-ambigu", "Belum dijalankan",
     "Behavior modal Radix standar. Case 3 (Esc) = asumsi Radix default"],

    # ================= NEGATIF =================
    ["TC-JP-ADD-009", "Negatif",
     "Klik Simpan dgn Instansi kosong -> error required inline",
     "Modal terbuka; state default",
     "1) Klik \"Tambah\"\n2) Biarkan Instansi placeholder (Tingkat/Kelas/Jurusan otomatis disabled)\n3) Klik \"Simpan\"",
     "Instansi=(kosong); field lain tidak reachable krn disabled",
     "Pesan error \"Instansi wajib diisi\" muncul di [data-slot=\"form-message\"] class text-destructive di bawah field Instansi; Simpan tidak bekerja; modal tetap terbuka; tidak redirect; list tidak berubah",
     "PRD", "Belum dijalankan",
     "Selector error: [data-slot=\"form-message\"] text-destructive"],

    ["TC-JP-ADD-010", "Negatif",
     "Klik Simpan dgn Tingkat kosong (Instansi terisi) -> error required",
     "Modal terbuka; sudah pilih Instansi",
     "1) Pilih Instansi\n2) Biarkan Tingkat placeholder\n3) Klik \"Simpan\"",
     "Instansi=<valid>; Tingkat=(kosong)",
     "Pesan error \"Tingkat wajib diisi\" muncul di form-message text-destructive di bawah field Tingkat; Simpan tidak bekerja; modal tetap terbuka; tidak redirect",
     "PRD", "Belum dijalankan", ""],

    ["TC-JP-ADD-011", "Negatif",
     "Klik Simpan dgn Kelas kosong (Multi OFF; Instansi+Tingkat terisi) -> error required",
     "Modal terbuka; Instansi + Tingkat terisi; Multi Kelas OFF",
     "1) Pilih Instansi + Tingkat\n2) Biarkan Kelas placeholder\n3) Klik \"Simpan\"",
     "Instansi=<valid>; Tingkat=<valid>; Kelas=(kosong); Multi=OFF",
     "Pesan error \"Kelas wajib diisi\" muncul di form-message text-destructive di bawah field Kelas; Simpan tidak bekerja",
     "PRD", "Belum dijalankan", ""],

    ["TC-JP-ADD-012", "Negatif",
     "Klik Simpan dgn Multi Kelas ON tapi Kelas kosong / hanya 1 kelas -> error min 2",
     "Modal terbuka; Instansi + Tingkat terisi; Multi Kelas ON",
     "Case 1: Multi ON, Kelas=(kosong) -> Simpan\nCase 2: Multi ON, pilih 1 Kelas saja -> Simpan\nAmati pesan error tiap case",
     "Case 1: Kelas=(kosong); Case 2: Kelas=[Kelas A]",
     "Pesan error muncul (mis. \"Minimal 2 kelas harus dipilih\" atau \"Kelas wajib diisi\") di form-message text-destructive; Simpan tidak bekerja; modal tetap terbuka; tidak redirect",
     "PRD", "Belum dijalankan",
     "PRD: \"Minimal untuk pilih 2 kelas jika hanya 1 kelas maka akan menampilkan pesan error\""],

    ["TC-JP-ADD-013", "Negatif",
     "Klik Simpan dgn seluruh field wajib kosong -> pesan error muncul",
     "Modal terbuka; state default",
     "1) Klik \"Tambah\"\n2) Langsung klik \"Simpan\" tanpa isi apapun",
     "Semua field wajib=(kosong)",
     "Pesan error required muncul untuk Instansi (mungkin hanya Instansi karena Tingkat/Kelas/Jurusan disabled belum bisa di-validate); Simpan tidak bekerja; modal tetap terbuka; Jurusan & Tag TIDAK memunculkan error (opsional); tidak redirect",
     "PRD-ambigu", "Belum dijalankan",
     "Catat aktual: apakah FE validate semua field sekaligus atau stop di parent chain pertama"],

    # ================= EDGE =================
    ["TC-JP-ADD-014", "Edge",
     "Duplikasi di modal: Instansi+Tingkat+Kelas+TA+Semester kombinasi sama -> toast warning, modal tetap terbuka",
     "Sudah ada entry jadwal untuk Instansi=X, Tingkat=Y, Kelas=Z, Tahun Ajaran+Semester aktif",
     "1) Klik \"Tambah\"\n2) Pilih Instansi=X, Tingkat=Y, Kelas=Z (kombinasi persis sama dgn existing)\n3) Klik \"Simpan\"\n4) Amati toast di bottom-right & state modal",
     "Kombinasi Instansi+Tingkat+Kelas identik dgn entry existing di periode aktif",
     "Validasi duplikasi berjalan di modal SEBELUM navigasi; toast [data-sonner-toast][data-type=\"warning\"] muncul dgn teks \"Jadwal pelajaran sudah terdaftar!\"; modal tetap terbuka (bukan redirect); tidak ada inline form-message; tidak ada entry duplikat di list",
     "PRD", "SKIP",
     "PINDAH KE FASE 2 -> aktual: modal Simpan LANGSUNG commit tanpa validate duplikat (kedua submit sukses & redirect), konsisten BUG-036. Toast warning \"sudah terdaftar\" kemungkinan muncul di flow halaman edit save. Bikin TC baru di TC sheet fase 2."],

    ["TC-JP-ADD-015", "Edge",
     "Ghost record prevention: user Simpan di modal lalu tinggalkan halaman edit tanpa Simpan final -> list tetap kosong",
     "Sudah selesai modal (Instansi+Tingkat+Kelas) & redirect ke halaman edit; belum klik Simpan final",
     "Baseline: catat jumlah baris list sebelum mulai\n\nCase A - Back tanpa Simpan:\n1) Setelah redirect ke halaman edit\n2) Klik back browser (atau navigasi ke menu lain)\n3) Kembali ke List Jadwal Pelajaran\n4) Cek jumlah baris & cari kombinasi input\n\nCase B - Close tab tanpa Simpan:\n1) Tutup tab saat di halaman edit\n2) Buka tab baru -> login -> masuk List Jadwal Pelajaran\n3) Cek jumlah baris & cari kombinasi input\n\nCase C - Reload halaman edit tanpa Simpan:\n1) Reload halaman edit\n2) Amati: apakah auto-fill data hilang (form kosong lagi) & user harus tambah ulang lewat modal",
     "Kombinasi Instansi+Tingkat+Kelas yg baru saja di-Simpan di modal (tapi belum Simpan final di halaman edit)",
     "SEMUA case: entry TIDAK ada di list (jumlah baris list = baseline); TIDAK ada ghost record (header tanpa jadwal). Case C: auto-fill hilang setelah reload = perilaku diterima (data belum ke-commit)",
     "PRD", "Belum dijalankan",
     "AKTUAL: entry SUDAH masuk list di titik modal Simpan (BUG-036). TC ini akan FAIL sampai BUG-036 fixed - itulah tujuannya (assert correct expected behavior, per CLAUDE.md)"],

    ["TC-JP-ADD-016", "Edge",
     "Halaman edit setelah redirect: verifikasi URL, auto-fill Data Jadwal Pelajaran, & list belum bertambah",
     "Sudah selesai TC-001 (modal Simpan sukses & redirect)",
     "1) Setelah redirect, verifikasi URL format `/setting/academic/course-schedule/{id}`\n2) Verifikasi title / breadcrumb halaman edit\n3) Verifikasi section \"Data Jadwal Pelajaran\" auto-fill sesuai input modal\n4) Verifikasi field auto-fill READ-ONLY (tidak bisa diubah dari halaman edit)\n5) Verifikasi button \"Tampilkan\" TIDAK ada (dihapus per refactor); section Jadwal Pelajaran langsung tampil siap diisi\n6) Buka tab baru ke List -> jumlah baris list = baseline",
     "Entry hasil TC-001 (belum di-Simpan final di halaman edit)",
     "URL cocok format dgn ID valid; halaman edit render section Data Jadwal Pelajaran; Tahun Ajaran/Semester/Instansi/Tingkat/Kelas/Jurusan/Tag terisi sesuai input modal; field auto-fill = read-only; button \"Tampilkan\" tidak ada; section Jadwal Pelajaran langsung visible; jumlah baris list = baseline (record baru muncul setelah Simpan final di fase 2)",
     "PRD", "Belum dijalankan",
     "AKTUAL: baris list bertambah setelah modal Simpan -> BUG-036. Fase 2 (Simpan final + jadwal per hari) di file TC terpisah"],

    ["TC-JP-ADD-017", "Positif",
     "Label tombol utama: modal = \"Simpan\", halaman edit = \"Simpan\"",
     "Login admin; List Jadwal Pelajaran terbuka",
     "1) Klik \"Tambah\" -> modal tampil\n2) Verifikasi label tombol utama footer modal = \"Simpan\"\n3) Isi Instansi+Tingkat+Kelas valid & klik Simpan\n4) Setelah redirect ke halaman edit, cari tombol commit\n5) Verifikasi label tombol commit di halaman edit = \"Simpan\"",
     "-",
     "Modal: tombol utama footer bertuliskan \"Simpan\" (aktual staging); Halaman edit: tombol commit bertuliskan \"Simpan\"",
     "PRD", "Belum dijalankan",
     "Refactor spec ke depan mengusulkan label modal jadi \"Lanjutkan\" (biar jelas beda dgn commit di halaman edit). Sekarang TC assert aktual \"Simpan\"; kalau refactor deploy, update TC ini + reference sesuai"],

    ["TC-JP-ADD-018", "Edge",
     "Ganti Instansi -> Tag TIDAK ikut reset (inkonsisten dgn child lain)",
     "Modal terbuka; state default",
     "1) Pilih Instansi = X\n2) Klik picker Tag; pilih 1 tag (mis. \"mobil\" via regex exact); chip muncul di button Tag\n3) Ubah Instansi ke Y (nilai berbeda)\n4) Amati state Tingkat/Kelas/Jurusan\n5) Amati state Tag (chip masih ada?)",
     "Instansi awal=X; Tag=[\"mobil\"]; Instansi baru=Y",
     "Tingkat/Kelas/Jurusan RESET ke placeholder (behavior benar); Tag JUGA reset ke placeholder (chip hilang) - konsisten dgn child lain karena tag di-scope per Instansi",
     "PRD", "Belum dijalankan",
     "AKTUAL: Tag TIDAK reset saat ganti Instansi (BUG-034) - chip yg dipilih dari Instansi X tetap ter-select di Instansi Y. TC akan FAIL sampai BUG-034 fixed"],

    ["TC-JP-ADD-019", "Edge",
     "Tag picker: dependency chain seharusnya Instansi -> Tag; picker disabled sampai Instansi dipilih; opsi tag scoped per Instansi",
     "Modal terbuka; state default (Instansi belum dipilih)",
     "1) Amati picker Tag (button \"Pilih Tag\") saat Instansi belum dipilih\n2) (Coba) Klik picker Tag\n3) Pilih Instansi = X\n4) Buka picker Tag; buka DevTools Network (XHR); cek response API tag\n5) Bandingkan daftar tag dgn tag milik Instansi X saja",
     "-",
     "(a) Picker Tag DISABLED sampai Instansi dipilih (konsisten dgn dependency chain Tingkat/Kelas/Jurusan); (b) Setelah Instansi dipilih, daftar tag hanya tag milik Instansi tersebut (query API di-scope instansi_id)",
     "PRD", "Belum dijalankan",
     "AKTUAL: (a) Picker Tag aktif meski Instansi belum dipilih; (b) Daftar tag load SEMUA tag lintas-instansi (ratusan QA... dari test data) -> BUG-035. TC akan FAIL sampai BUG-035 fixed. Kritis untuk mengurangi sumber flakiness automation"],
]

# Legend reference data
LEGEND = [
    ["Kolom Sumber", ""],
    ["PRD", "Skenario diturunkan langsung dari PRD Jadwal Pelajaran / refactor spec / verified manual"],
    ["PRD-ambigu", "PRD menyinggung tapi tidak jelas / app berbeda -> perlu konfirmasi"],
    ["Asumsi", "Tidak ada di PRD; asumsi berbasis konvensi/HTML/modul lain"],
    ["", ""],
    ["Kolom Kategori", ""],
    ["Happy", "Alur utama sukses sesuai desain"],
    ["Positif", "Variasi valid non-utama"],
    ["Negatif", "Input/kondisi yg harus ditolak sistem"],
    ["Edge", "Batas/kondisi tepi & verifikasi persistence + redirect"],
    ["", ""],
    ["Kolom Status", ""],
    ["Belum dijalankan", "Sudah di-generate, menunggu eksekusi Cypress"],
    ["DITUNDA", "Sengaja belum di-run; menunggu prasyarat tersedia"],
    ["", ""],
    ["Catatan", "Tanpa kategori security/injection (sesuai kesepakatan)"],
    ["", "Bug ditandai inline: Status=FAIL + ref BUG-### di kolom Catatan"],
    ["", "TC yg assert behavior BENAR (post-refactor) sengaja FAIL sampai bug fixed"],
    ["", "  -> TC-001/002/003/015/016 -> BUG-036 (modal langsung commit)"],
    ["", "  -> TC-018 -> BUG-034 (Tag gak reset saat ganti Instansi)"],
    ["", "  -> TC-019 -> BUG-035 (Tag picker prematur + tidak scoped)"],
    ["", ""],
    ["SCOPE TC ini", ""],
    ["", "TC file ini FOKUS di FASE 1: Modal \"Tambah Jadwal Pelajaran\" + redirect verify."],
    ["", "Fase 2 (halaman edit /course-schedule/{id}: set jadwal per hari, toggle jam pelajaran,"],
    ["", "  auto-generate, validasi bentrok waktu/guru, Simpan final) -> file TC terpisah setelah HTML"],
    ["", "  halaman edit lengkap di-share."],
    ["", ""],
    ["Status refactor spec", ""],
    ["", "Staging saat ini = PRE-REFACTOR. Modal button label = \"Simpan\"; klik Simpan LANGSUNG"],
    ["", "  commit record ke DB (BUG-036 - risiko ghost record)."],
    ["", "Refactor spec dari PO: modal button HANYA navigasi + bawa data ke halaman edit;"],
    ["", "  commit sebenarnya di Simpan halaman edit (fase 2)."],
    ["", "TC assert BEHAVIOR YG BENAR (post-refactor) - konvensi CLAUDE.md. TC akan FAIL"],
    ["", "  sampai bug fixed (itu tujuannya)."],
    ["", ""],
    ["Field modal", ""],
    ["", "Tahun Ajaran (readonly text, preselect aktif, mis. \"2026/2027\")"],
    ["", "Semester (readonly badge, preselect aktif, mis. \"GANJIL\")"],
    ["", "Instansi* (Select Radix, placeholder \"Pilih Instansi\")"],
    ["", "Tingkat* (Select Radix, DISABLED sampai Instansi dipilih)"],
    ["", "Multi Kelas (Switch toggle, default OFF; ON -> hide Jurusan, Kelas jadi multi-select)"],
    ["", "Kelas* (Select Radix, DISABLED sampai Tingkat dipilih)"],
    ["", "Jurusan (Opsional; Select Radix, DISABLED sampai Kelas dipilih)"],
    ["", "Tag (Opsional; inline dropdown checkbox multi-select, no search)"],
    ["", "Btn Batal + Btn \"Simpan\" (aktual) + X (dialog-close)"],
    ["", ""],
    ["Selector hint (verified)", ""],
    ["", "Modal: [data-slot=\"dialog-content\"] + role=\"dialog\""],
    ["", "Title: [data-slot=\"dialog-title\"] > \"Tambah Jadwal Pelajaran\""],
    ["", "Btn tambah trigger: [data-slot=\"dialog-trigger\"] + svg.lucide-square-plus"],
    ["", "Form item: [data-slot=\"form-item\"] scoped by label [data-slot=\"form-label\"]"],
    ["", "Select trigger: [data-slot=\"select-trigger\"] + role=\"combobox\""],
    ["", "Select content: [data-slot=\"select-content\"] -> DIPORTAL ke <body>, JANGAN .within(dialog)"],
    ["", "Select item: [data-slot=\"select-item\"] role=\"option\""],
    ["", "Switch Multi Kelas: [role=\"switch\"] name=\"combined_schedule\""],
    ["", "Tag checkbox item: <button role=\"checkbox\" data-slot=\"checkbox\"> + <label data-slot=\"label\">"],
    ["", "Tag toggle: klik LABEL (regex exact /^tagname$/), BUKAN cy.check()"],
    ["", "Btn utama: type=\"submit\" data-slot=\"button\" text=\"Simpan\""],
    ["", "Btn Batal: data-slot=\"dialog-close\" text=\"Batal\" (footer)"],
    ["", "X close: data-slot=\"dialog-close\" (pojok kanan atas)"],
    ["", "Form error: [data-slot=\"form-message\"] class text-destructive"],
    ["", "Toast: [data-sonner-toast][data-type=\"success|warning|error\"]"],
    ["", ""],
    ["Selector list (verified)", ""],
    ["", "Tabel: table[data-slot=\"data-grid-table\"] > tbody > tr"],
    ["", "Kolom aktual (10): Instansi | Tahun Ajaran | Semester | Tingkat | Kelas |"],
    ["", "  Jurusan | Tag | Status | Dibuat Pada | Aksi (pinned kanan)"],
    ["", "Multi Kelas ON -> 1 row (chip pertama + badge N+ dalam 1 <td>), BUKAN N row"],
    ["", ""],
    ["URL redirect", ""],
    ["", "Setelah Simpan modal sukses -> /setting/academic/course-schedule/{id}"],
]

# ---- styling ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CAT_FILL = {
    "Happy":   PatternFill("solid", fgColor="E2EFDA"),
    "Positif": PatternFill("solid", fgColor="DDEBF7"),
    "Negatif": PatternFill("solid", fgColor="FCE4D6"),
    "Edge":    PatternFill("solid", fgColor="FFF2CC"),
}
SRC_FONT = {
    "PRD":        Font(color="375623"),
    "PRD-ambigu": Font(color="BF8F00", bold=True),
    "Asumsi":     Font(color="C00000", bold=True),
}
STATUS_FILL = {
    "DITUNDA": PatternFill("solid", fgColor="D9D9D9"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPWRAP = Alignment(vertical="top", wrap_text=True)
WIDTHS = [16, 12, 42, 34, 48, 30, 54, 14, 16, 46]

wb = Workbook()
ws = wb.active
ws.title = "Tambah Jadwal Pelajaran"

# Title row
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
t = ws.cell(row=1, column=1, value="Test Case - Tambah Jadwal Pelajaran (CARDS School v3)  |  Modul: Pengaturan Akademik > Jadwal Pelajaran  |  Scope: FASE 1 (Modal Tambah + redirect)  |  Assert: behavior BENAR (post-refactor)")
t.font = Font(bold=True, size=12)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22

# Header row (row 2)
for c, name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=2, column=c, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = BORDER

# Data rows
for r, row in enumerate(ROWS, start=3):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = TOPWRAP
        cell.border = BORDER
    cat = row[1]
    if cat in CAT_FILL:
        ws.cell(row=r, column=2).fill = CAT_FILL[cat]
    src = row[7]
    if src in SRC_FONT:
        ws.cell(row=r, column=8).font = SRC_FONT[src]
    status = row[8]
    if status in STATUS_FILL:
        ws.cell(row=r, column=9).fill = STATUS_FILL[status]

# Column widths
for c, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + len(ROWS)}"

# Legend sheet
ls = wb.create_sheet("Legenda")
ls.column_dimensions["A"].width = 24
ls.column_dimensions["B"].width = 82
for r, (a, b) in enumerate(LEGEND, 1):
    ca = ls.cell(row=r, column=1, value=a)
    cb = ls.cell(row=r, column=2, value=b)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    if a in ("Kolom Sumber", "Kolom Kategori", "Kolom Status", "Catatan",
             "SCOPE TC ini", "Status refactor spec", "Field modal",
             "Selector hint (verified)", "Selector list (verified)",
             "URL redirect"):
        ca.font = Font(bold=True)

out = "docs/test-cases/TC_JadwalPelajaran_Tambah.xlsx"
wb.save(out)
print("saved:", out, "| rows:", len(ROWS))
