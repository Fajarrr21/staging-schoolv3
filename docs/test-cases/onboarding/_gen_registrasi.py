# -*- coding: utf-8 -*-
"""Generate Test Case sheet for Onboarding Partner — Form Registrasi (PRD 1 sub-flow A).

Revisi v2 (post element analysis):
  - Hapus TC "all empty -> 4 error required": app TIDAK punya validasi "wajib diisi";
    field kosong cuma bikin tombol Submit tetap disabled (sudah cover di TC-003).
  - Update strength label: Weak/Medium/Strong -> Lemah/Sedang/Kuat (Bahasa Indonesia).
  - Update Google OAuth: tombol DISABLED + tooltip "Login Google segera hadir" (not clickable).
  - Update helper text Tips: text app real "Tips: Password minimal 8 karakter dengan huruf besar, kecil, dan angka."
  - Update link: text "Masuk" (bukan "Log In"), href="/masuk" (bukan "/login").
  - Update text exact pesan error: 4 messages real dari HTML (em-dash "10–13").
  - Update happy stop point: redirect /verifikasi-otp (NO toast, app tidak punya toast di halaman ini).
  - Update submit text loading: "Membuat akun..." + spinner.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = ["ID", "Kategori", "Skenario", "Pre-condition", "Langkah",
           "Test Data", "Expected", "Sumber", "Status", "Catatan"]

# Form Registrasi (per PRD 1 + element analysis HTML asli):
# URL: https://staging.cards.co.id/daftar-akun
# Selector strategy: ID-based (#reg-name, #reg-phone, #reg-email, #reg-pw) — app pakai Tailwind biasa,
#   BUKAN shadcn/Radix (tidak ada data-slot).
#
# Fields:
#   a. Nama*      (input id=reg-name, name=name)              -> min 3 char
#   b. Nomor HP*  (input id=reg-phone, name=phoneNumber,      -> 10-13 digit, numerik
#                  type=tel, inputmode=numeric)
#   c. Email*     (input id=reg-email, name=email, type=email)-> format valid
#   d. Password*  (input id=reg-pw, name=password,            -> min 8 + besar/kecil/angka
#                  type=password) + tombol eye toggle
#   e. Tombol Google (DISABLED + title "Login Google segera hadir" — coming soon)
#   f. Tombol "Buat Akun" type=submit (disabled saat field invalid; loading: "Membuat akun..." + spinner)
#   g. Link "Sudah punya akun? Masuk" (a href="/masuk")
#   h. Page H1: "Daftar"
#
# Pesan error real (dari HTML):
#   - Nama:     "Nama minimal 3 karakter"
#   - No HP:    "Nomor HP tidak valid (10–13 digit angka)"   <- em-dash, BUKAN strip
#   - Email:    "Format email tidak valid"
#   - Password: "Password minimal 8 karakter dan harus mengandung kombinasi huruf besar, kecil, serta angka."
#
# Visual error state:
#   - input.aria-invalid="true" + class "border-pink-400"
#   - icon info pink: svg.text-pink-500 (pointer-events-none)
#   - message: p.text-pink-600
#
# Strength Indicator (real-time, container <div class="mt-2">):
#   - Label: <p>Kekuatan password: <span class="text-{red-500|yellow-600|green-600}">Lemah|Sedang|Kuat</span></p>
#   - 3 bars: <span class="h-1 flex-1 ..."> dengan bg color berbeda per level
#
# Trigger post-register sukses:
#   - Redirect ke https://staging.cards.co.id/verifikasi-otp
#   - TIDAK ada toast di halaman ini (confirmed user)
#   - OTP & PIN dikirim ke email (background, tidak terverifikasi di TC ini)

ROWS = [
    # ---------------- HAPPY ----------------
    ["TC-ONB-REG-001", "Happy",
     "Registrasi lengkap valid -> redirect ke halaman /verifikasi-otp",
     "Halaman https://staging.cards.co.id/daftar-akun terbuka; email + no HP belum pernah terdaftar",
     "1) Isi Nama (min 3 char)\n2) Isi Nomor HP (10-13 digit, numerik)\n3) Isi Email (format valid + unique)\n4) Isi Password (8+ char, kombinasi besar/kecil/angka)\n5) Tunggu tombol \"Buat Akun\" berubah enabled\n6) Klik \"Buat Akun\"",
     "Nama=Qa Cards Test; No HP=0812<rand6>; Email=qa.cards.school+<ts>@gmail.com; Password=Qwerty12",
     "Tombol \"Buat Akun\" loading state aktif (text berubah ke \"Membuat akun...\" + spinner); request submit dikirim; URL berubah ke https://staging.cards.co.id/verifikasi-otp",
     "PRD", "Belum dijalankan",
     "TIDAK ada toast (app tidak punya toast di halaman registrasi). Stop point happy = URL redirect ke /verifikasi-otp. OTP & PIN dikirim ke email tapi tidak diverifikasi di TC ini (= scope PRD 2)"],

    ["TC-ONB-REG-002", "Happy",
     "Registrasi dengan Password Kuat (12+ char + simbol)",
     "Halaman daftar-akun terbuka; email + no HP belum terdaftar",
     "1) Isi semua field valid\n2) Pakai password Kuat: 16 char + huruf besar/kecil/angka + simbol\n3) Amati strength indicator -> Kuat (hijau, 3 bar terisi)\n4) Klik \"Buat Akun\"",
     "Password=CardsSchool2026! (16 char + simbol)",
     "Strength indicator label \"Kuat\" warna hijau (text-green-600) + 3 bar bg-green-500; submit sukses; redirect ke /verifikasi-otp",
     "PRD", "Belum dijalankan", ""],

    # ---------------- POSITIF ----------------
    ["TC-ONB-REG-003", "Positif",
     "Halaman Registrasi terbuka -> H1 \"Daftar\", semua field kosong, 2 tombol disabled",
     "Belum login; navigate ke /daftar-akun",
     "1) Buka https://staging.cards.co.id/daftar-akun\n2) Amati kondisi awal halaman",
     "-",
     "H1 = \"Daftar\"; subtitle \"Sudah punya akun? Masuk\" (link Masuk visible); 4 field (Nama/Nomor HP/Email/Password) kosong dengan placeholder; tombol \"Buat Akun\" DISABLED (bg-blue-300); tombol \"Daftar dengan Google\" DISABLED (cursor-not-allowed) + tooltip title=\"Login Google segera hadir\"; helper text Tips visible di bawah field Password",
     "PRD", "Belum dijalankan",
     "Default state: kedua tombol disabled. Field kosong TIDAK memunculkan error (app tidak punya validasi \"wajib diisi\" — error hanya muncul saat input pendek/format salah)"],

    ["TC-ONB-REG-004", "Positif",
     "Real-time validation: tombol \"Buat Akun\" enable HANYA setelah semua wajib valid",
     "Halaman daftar-akun terbuka",
     "1) Isi Nama valid (3+ char) -> tombol masih disabled\n2) Tambah Nomor HP valid -> masih disabled\n3) Tambah Email valid -> masih disabled\n4) Tambah Password valid -> tombol berubah enabled (bg-blue-600)",
     "Inkremental fill 4 field valid",
     "Tombol \"Buat Akun\" tetap disabled selama ada 1+ field invalid/kosong. Begitu semua 4 field valid, attribute disabled hilang, tombol jadi clickable (bg-blue-600). Tidak perlu klik submit untuk ngecek state ini",
     "PRD", "Belum dijalankan", ""],

    ["TC-ONB-REG-005", "Positif",
     "Show/Hide password toggle (ikon Mata) berfungsi",
     "Halaman daftar-akun; field Password diisi",
     "1) Isi Password = \"Qwerty12\"\n2) Amati: field default masking (titik/bintang, type=password)\n3) Klik ikon mata (button aria-label=\"Tampilkan password\")\n4) Amati: text plain visible (type=text)\n5) Klik ikon mata lagi",
     "Password=Qwerty12",
     "Default masking (input type=password); klik mata pertama -> type berubah ke text, password plain visible; klik mata kedua -> kembali masking",
     "PRD", "Belum dijalankan",
     "Selector eye toggle: button[aria-label=\"Tampilkan password\"] di samping #reg-pw. Catat: kalau aria-label berubah ke \"Sembunyikan password\" saat aktif, assert label switch"],

    ["TC-ONB-REG-006", "Positif",
     "Link \"Masuk\" mengarahkan ke halaman Login (/masuk)",
     "Halaman daftar-akun terbuka",
     "1) Klik link \"Masuk\" di subtitle (anchor a[href=\"/masuk\"])\n2) Amati URL & halaman",
     "-",
     "URL berubah ke https://staging.cards.co.id/masuk; halaman Login muncul (form Email + Password)",
     "PRD", "Belum dijalankan", ""],

    ["TC-ONB-REG-007", "Positif",
     "Tombol \"Daftar dengan Google\" tampil tapi DISABLED (coming soon)",
     "Halaman daftar-akun terbuka",
     "1) Amati tombol di bawah divider \"Atau\"\n2) Hover tombol untuk lihat tooltip\n3) Coba klik tombol",
     "-",
     "Tombol \"Daftar dengan Google\" visible dengan ikon Google; attribute disabled=true; class cursor-not-allowed; title attribute = \"Login Google segera hadir\" (muncul sebagai tooltip native browser saat hover); klik tidak trigger event apapun",
     "PRD", "Belum dijalankan",
     "Coming soon feature. Full Google OAuth flow tetap DEFERRED (TC-031/032). Cypress: assert .should('be.disabled') + have.attr('title', '...')"],

    ["TC-ONB-REG-008", "Positif",
     "Password Strength Indicator: Lemah (8 char + complexity minimum)",
     "Halaman daftar-akun terbuka",
     "1) Klik field Password\n2) Ketik \"Qwerty12\" (8 char, 1 besar + kecil + angka)\n3) Amati indicator container <div class=\"mt-2\"> di bawah helper text",
     "Password=Qwerty12 (8 char)",
     "Indicator muncul: text \"Kekuatan password: Lemah\" dengan <span> warna merah (text-red-500); 1 bar bg-red-400 (penuh) + 2 bar bg-slate-200 (kosong)",
     "PRD", "Belum dijalankan", ""],

    ["TC-ONB-REG-009", "Positif",
     "Password Strength Indicator: Sedang (10+ char + complexity)",
     "Halaman daftar-akun terbuka",
     "1) Ketik Password = \"Qwerty1234\" (10 char, kombinasi minimum)\n2) Amati indicator",
     "Password=Qwerty1234 (10 char)",
     "Indicator: text \"Kekuatan password: Sedang\" warna kuning (text-yellow-600); 2 bar bg-yellow-400 + 1 bar bg-slate-200",
     "PRD", "Belum dijalankan", ""],

    ["TC-ONB-REG-010", "Positif",
     "Password Strength Indicator: Kuat (12+ char + complexity + simbol)",
     "Halaman daftar-akun terbuka",
     "1) Ketik Password = \"CardsSchool2026!\" (16 char + simbol)\n2) Amati indicator",
     "Password=CardsSchool2026! (16 char + simbol)",
     "Indicator: text \"Kekuatan password: Kuat\" warna hijau (text-green-600); 3 bar bg-green-500 (semua penuh)",
     "PRD", "Belum dijalankan", ""],

    ["TC-ONB-REG-011", "Positif",
     "Helper text Tips muncul permanent di bawah field Password",
     "Halaman daftar-akun terbuka (sebelum input apapun)",
     "1) Buka halaman daftar-akun\n2) Amati helper text di bawah field Password",
     "-",
     "Helper text always-visible: \"Tips: Password minimal 8 karakter dengan huruf besar, kecil, dan angka.\" (p.text-slate-400)",
     "PRD", "Belum dijalankan",
     "Text REAL dari HTML (berbeda dari PRD literal). PRD bilang \"gunakan passphrase\" tapi app pakai versi simpler. Always-visible (tidak depend focus)"],

    ["TC-ONB-REG-012", "Positif",
     "Tombol \"Buat Akun\" loading state saat submit (anti double-click)",
     "Halaman daftar-akun; semua field valid",
     "1) Isi semua field valid + unik\n2) Klik \"Buat Akun\" tepat 1x\n3) Amati tombol DURING request (network throttle bila perlu)",
     "data valid",
     "Setelah klik: tombol text berubah dari \"Buat Akun\" ke \"Membuat akun...\"; muncul span.animate-spin (spinner border); tombol tidak responsif klik kedua selama loading; setelah response sukses -> redirect /verifikasi-otp",
     "PRD", "Belum dijalankan",
     "Cypress: cy.intercept request, assert button contains \"Membuat akun...\" + has span.animate-spin during request"],

    # ---------------- NEGATIF ----------------
    ["TC-ONB-REG-013", "Negatif",
     "Nama < 3 karakter -> error \"Nama minimal 3 karakter\"",
     "Halaman daftar-akun; field lain valid",
     "1) Isi Nama = \"Ab\" (2 char)\n2) Blur field (Tab atau klik field lain)\n3) Amati state field & error message",
     "Nama=Ab (2 char)",
     "Input #reg-name dapat class border-pink-400 + aria-invalid=\"true\"; ikon pink info muncul di kanan input; pesan error: \"Nama minimal 3 karakter\" (p.text-pink-600); tombol Buat Akun tetap disabled",
     "PRD", "Belum dijalankan", ""],

    ["TC-ONB-REG-014", "Negatif",
     "Email format invalid (tanpa @) -> error \"Format email tidak valid\"",
     "Halaman daftar-akun; field lain valid",
     "1) Isi Email = \"qa.cards.school\" (tanpa @)\n2) Blur field\n3) Amati error",
     "Email=qa.cards.school (no @)",
     "Input #reg-email aria-invalid=\"true\" + border-pink; pesan error \"Format email tidak valid\"; tombol Buat Akun disabled",
     "PRD", "Belum dijalankan", ""],

    ["TC-ONB-REG-015", "Negatif",
     "Email format invalid (tanpa domain) -> error \"Format email tidak valid\"",
     "Halaman daftar-akun; field lain valid",
     "1) Isi Email = \"qa.cards@\"\n2) Blur field\n3) Amati error",
     "Email=qa.cards@ (no domain)",
     "Pesan error \"Format email tidak valid\"; submit ditolak",
     "PRD", "Belum dijalankan", ""],

    ["TC-ONB-REG-016", "Negatif",
     "Email sudah terdaftar (uniqueness check) -> error backend",
     "BLOCKED — butuh akun seed yang sudah terdaftar di staging",
     "1) Isi Email = <email akun seed yang sudah terdaftar>\n2) Isi field lain valid\n3) Klik \"Buat Akun\"",
     "Email=<seed email>",
     "Submit dikirim; response backend balas error duplikat; pesan error muncul (text exact dan lokasi UI menunggu observasi saat seed tersedia — kemungkinan inline di field Email atau alert)",
     "PRD", "PENDING",
     "BLOCKED: butuh seed account. User akan maintain. Aktifkan TC setelah credential seed tersedia + observe UI"],

    ["TC-ONB-REG-017", "Negatif",
     "Nomor HP < 10 digit -> error \"Nomor HP tidak valid (10–13 digit angka)\"",
     "Halaman daftar-akun; field lain valid",
     "1) Isi Nomor HP = \"081234567\" (9 digit)\n2) Blur field\n3) Amati error",
     "No HP=081234567 (9 digit)",
     "Pesan error: \"Nomor HP tidak valid (10–13 digit angka)\" (em-dash); border-pink + aria-invalid; tombol Buat Akun disabled",
     "PRD", "Belum dijalankan",
     "PERHATIAN: \"10–13\" pakai em-dash (–), BUKAN strip biasa (-). Pastikan fixture pakai char yang sama"],

    ["TC-ONB-REG-018", "Negatif",
     "Nomor HP > 13 digit -> input dibatasi atau error",
     "Halaman daftar-akun; field lain valid",
     "1) Coba ketik Nomor HP = \"08123456789012\" (14 digit)\n2) Amati: apakah input dibatasi (maxLength) atau muncul error \"Nomor HP tidak valid\"",
     "No HP=08123456789012 (14 digit)",
     "Behavior salah satu: (a) input dibatasi ke 13 char client-side; ATAU (b) error \"Nomor HP tidak valid (10–13 digit angka)\". Yang penting: tidak bisa submit dengan >13 digit",
     "PRD", "Belum dijalankan",
     "Catat behavior aktual saat run"],

    ["TC-ONB-REG-019", "Negatif",
     "Nomor HP mengandung huruf/simbol -> filter input (numeric only)",
     "Halaman daftar-akun",
     "1) Coba ketik di field Nomor HP: \"081abc-2345!@\" (mix)\n2) Amati: karakter non-numerik tidak masuk",
     "input attempt: 081abc-2345!@",
     "Hanya digit yang masuk ke value field (\"0812345\"); huruf/simbol auto-filter (inputmode=numeric + type=tel). Cypress: cy.get('#reg-phone').type('081abc-2345!@').should('have.value', '0812345')",
     "PRD", "Belum dijalankan",
     "Note: type=tel di browser umumnya NOT filter — kemungkinan ada JS filter di app. Verify behavior"],

    ["TC-ONB-REG-020", "Negatif",
     "Nomor HP sudah terdaftar (uniqueness check) -> error backend",
     "BLOCKED — butuh akun seed dengan No HP yang sudah terdaftar",
     "1) Isi Nomor HP = <no HP akun seed>\n2) Isi field lain valid + email unik\n3) Klik \"Buat Akun\"",
     "No HP=<seed no hp>",
     "Submit dikirim; backend balas error duplikat; pesan error muncul (lokasi & text exact menunggu observasi saat seed tersedia)",
     "PRD", "PENDING",
     "BLOCKED: butuh seed account"],

    ["TC-ONB-REG-021", "Negatif",
     "Password < 8 karakter -> error complexity",
     "Halaman daftar-akun; field lain valid",
     "1) Isi Password = \"Qwer123\" (7 char)\n2) Blur field\n3) Amati error",
     "Password=Qwer123 (7 char)",
     "Pesan error: \"Password minimal 8 karakter dan harus mengandung kombinasi huruf besar, kecil, serta angka.\"; border-pink + aria-invalid; tombol disabled",
     "PRD", "Belum dijalankan", ""],

    ["TC-ONB-REG-022", "Negatif",
     "Password tanpa huruf besar -> error complexity",
     "Halaman daftar-akun; field lain valid",
     "1) Isi Password = \"qwerty12\" (8 char, tanpa huruf besar)\n2) Blur field\n3) Amati error",
     "Password=qwerty12 (no uppercase)",
     "Pesan error sama: \"Password minimal 8 karakter dan harus mengandung kombinasi huruf besar, kecil, serta angka.\"; tombol disabled",
     "PRD", "Belum dijalankan", ""],

    ["TC-ONB-REG-023", "Negatif",
     "Password tanpa huruf kecil -> error complexity",
     "Halaman daftar-akun; field lain valid",
     "1) Isi Password = \"QWERTY12\" (8 char, tanpa huruf kecil)\n2) Blur field\n3) Amati error",
     "Password=QWERTY12 (no lowercase)",
     "Pesan error complexity; tombol disabled",
     "PRD", "Belum dijalankan", ""],

    ["TC-ONB-REG-024", "Negatif",
     "Password tanpa angka -> error complexity",
     "Halaman daftar-akun; field lain valid",
     "1) Isi Password = \"Qwertyab\" (8 char, tanpa angka)\n2) Blur field\n3) Amati error",
     "Password=Qwertyab (no digit)",
     "Pesan error complexity; tombol disabled",
     "PRD", "Belum dijalankan", ""],

    # ---------------- EDGE ----------------
    ["TC-ONB-REG-025", "Edge",
     "Nama tepat 3 karakter (min boundary) -> diterima",
     "Halaman daftar-akun; field lain valid + unik",
     "1) Isi Nama = \"Fai\" (3 char tepat)\n2) Isi field lain valid + unik\n3) Klik \"Buat Akun\"",
     "Nama=Fai (3 char)",
     "Tidak ada error di field Nama; tombol enable; submit sukses; redirect /verifikasi-otp. Boundary minimum diterima (>= 3, bukan > 3)",
     "PRD", "Belum dijalankan", ""],

    ["TC-ONB-REG-026", "Edge",
     "Nomor HP tepat 10 digit (min boundary) -> diterima",
     "Halaman daftar-akun; field lain valid + unik",
     "1) Isi Nomor HP = \"0812345678\" (10 digit tepat)\n2) Field lain valid\n3) Klik \"Buat Akun\"",
     "No HP=0812345678 (10 digit)",
     "Tidak ada error No HP; submit sukses; redirect /verifikasi-otp",
     "PRD", "Belum dijalankan", ""],

    ["TC-ONB-REG-027", "Edge",
     "Nomor HP tepat 13 digit (max boundary) -> diterima",
     "Halaman daftar-akun; field lain valid + unik",
     "1) Isi Nomor HP = \"0812345678901\" (13 digit tepat)\n2) Field lain valid\n3) Klik \"Buat Akun\"",
     "No HP=0812345678901 (13 digit)",
     "Submit sukses; redirect /verifikasi-otp",
     "PRD", "Belum dijalankan", ""],

    ["TC-ONB-REG-028", "Edge",
     "Password tepat 8 karakter (min boundary) dengan complexity minimum -> diterima sebagai Lemah",
     "Halaman daftar-akun; field lain valid + unik",
     "1) Isi Password = \"Qwerty12\" (8 char tepat)\n2) Amati strength indicator -> Lemah\n3) Field lain valid + unik\n4) Klik \"Buat Akun\"",
     "Password=Qwerty12",
     "Strength indicator: Lemah (merah, 1 bar); tidak ada error di field Password; tombol enable; submit DITERIMA (Lemah tetap submittable per PRD); redirect /verifikasi-otp",
     "PRD", "Belum dijalankan",
     "Verify: Lemah TIDAK blocker submit, hanya informational"],

    ["TC-ONB-REG-029", "Edge",
     "Nama whitespace-only (3+ spasi) -> error/ditolak",
     "Halaman daftar-akun; field lain valid",
     "1) Isi Nama = \"   \" (3+ spasi saja)\n2) Blur field\n3) Amati error",
     "Nama=(3 spasi)",
     "Behavior salah satu: (a) trim auto -> length 0 -> error \"Nama minimal 3 karakter\"; ATAU (b) accepted apa adanya (3 char termasuk spasi). Catat behavior aktual",
     "PRD-ambigu", "Belum dijalankan",
     "PRD tidak eksplisit soal trim. Test asumsi standar: whitespace-only dianggap kosong/length 0"],

    ["TC-ONB-REG-030", "Edge",
     "Email dengan leading/trailing space -> trim atau error format",
     "Halaman daftar-akun; field lain valid; email unik",
     "1) Isi Email = \"  qa.cards.school+<ts>@gmail.com  \" (spasi di awal & akhir)\n2) Blur field\n3) Amati error vs submit",
     "Email=\"  qa@gmail.com  \"",
     "Behavior salah satu: (a) trim auto + submit sukses; ATAU (b) error \"Format email tidak valid\" karena ada whitespace. Catat actual",
     "PRD-ambigu", "Belum dijalankan", ""],

    # ---------------- DEFERRED ----------------
    ["TC-ONB-REG-031", "Positif",
     "Registrasi via Google OAuth sukses -> auto-fill Nama+Email",
     "DEFERRED — tombol Google DISABLED di app + Cypress tidak bisa cross-origin Google",
     "1) Klik tombol \"Daftar dengan Google\"\n2) Login di popup Google\n3) Allow akses\n4) Cek form auto-fill",
     "Google account valid",
     "(Saat fitur sudah aktif) Form Nama+Email auto-filled dari Google; user lengkapi No HP+Password; submit -> redirect /verifikasi-otp",
     "PRD", "DITUNDA",
     "DOUBLE BLOCKED: (1) tombol di-app disable \"Login Google segera hadir\"; (2) Cypress can't cross-origin Google. Aktifkan kalau (a) tombol enable + (b) ada stub strategy"],

    ["TC-ONB-REG-032", "Negatif",
     "Registrasi via Google dibatalkan -> kembali ke form tanpa error",
     "DEFERRED — sama TC-031",
     "1) Klik tombol Google\n2) Klik Cancel di popup Google",
     "Cancel di Google popup",
     "Sistem kembali ke /daftar-akun; tidak ada error mengganggu; form tetap kosong (default)",
     "PRD", "DITUNDA",
     "DEFERRED: same blockers"],

    ["TC-ONB-REG-033", "Positif",
     "PIN 6-digit dikirim ke email setelah registrasi sukses",
     "DEFERRED — butuh akses inbox / mail catcher",
     "1) Register sukses (TC-001)\n2) Cek inbox email akun\n3) Cari email subject \"PIN Akun Cards School Anda\"",
     "akun yang baru terdaftar",
     "Ada email dari sistem dengan subject \"PIN Akun Cards School Anda\", berisi PIN 6-digit unique + reminder simpan",
     "PRD", "DITUNDA",
     "DEFERRED: butuh akses inbox / mail catcher. Aktifkan setelah strategi mail tersedia"],
]

# Legend reference data
LEGEND = [
    ["Modul", "Onboarding Partner — PRD 1 sub-flow A (Form Registrasi)"],
    ["URL", "https://staging.cards.co.id/daftar-akun"],
    ["URL post-success", "https://staging.cards.co.id/verifikasi-otp"],
    ["Total TC", f"{len(ROWS)} TC ({sum(1 for r in ROWS if r[1]=='Happy')} Happy, {sum(1 for r in ROWS if r[1]=='Positif')} Positif, {sum(1 for r in ROWS if r[1]=='Negatif')} Negatif, {sum(1 for r in ROWS if r[1]=='Edge')} Edge)"],
    ["", ""],
    ["Kolom Sumber", ""],
    ["PRD", "Skenario diturunkan langsung dari PRD 1 (Form Registrasi)"],
    ["PRD-ambigu", "PRD menyinggung tapi tidak jelas / detail belum ada -> butuh observasi saat run"],
    ["Asumsi", "Tidak ada di PRD; asumsi berbasis konvensi standar"],
    ["", ""],
    ["Kolom Kategori", ""],
    ["Happy", "Alur utama sukses sesuai desain PRD"],
    ["Positif", "Variasi valid non-utama (state form, link, strength indicator, etc.)"],
    ["Negatif", "Input invalid / kondisi yang harus ditolak sistem"],
    ["Edge", "Batas/kondisi tepi (boundary value)"],
    ["", ""],
    ["Kolom Status", ""],
    ["Belum dijalankan", "Sudah di-generate, menunggu eksekusi Cypress"],
    ["PENDING", "BLOCKED — butuh akun seed yang sudah terdaftar (TC duplikat email/HP)"],
    ["DITUNDA", "DEFERRED — keterbatasan tools (Google OAuth, PIN email)"],
    ["", ""],
    ["Catatan implementasi", ""],
    ["", "Selector strategy: ID-based (#reg-name, #reg-phone, #reg-email, #reg-pw). App pakai Tailwind biasa, TIDAK ada data-slot (beda dari modul Pengaturan)"],
    ["", "Tanpa kategori security/injection (sesuai kesepakatan)"],
    ["", "Bug ditandai inline: Status=FAIL + ref BUG-### di kolom Catatan"],
    ["", "Test data Email: qa.cards.school+<6digit-ts>@gmail.com (Gmail plus addressing, anti-duplicate)"],
    ["", "Test data No HP: prefix 0812 + 6 digit random (10 digit total) — verify uniqueness saat run"],
    ["", "Stop point happy flow: URL berubah ke /verifikasi-otp. TIDAK ada toast (confirmed)"],
    ["", "Error message exact (untuk fixture):"],
    ["", "  Nama: \"Nama minimal 3 karakter\""],
    ["", "  No HP: \"Nomor HP tidak valid (10–13 digit angka)\"  <- em-dash"],
    ["", "  Email: \"Format email tidak valid\""],
    ["", "  Password: \"Password minimal 8 karakter dan harus mengandung kombinasi huruf besar, kecil, serta angka.\""],
    ["", "Helper text Tips (real): \"Tips: Password minimal 8 karakter dengan huruf besar, kecil, dan angka.\""],
    ["", "Strength labels (BUKAN Weak/Medium/Strong): Lemah / Sedang / Kuat"],
    ["", "Tombol Google: DISABLED dengan title=\"Login Google segera hadir\" (coming soon)"],
    ["", "Submit loading text: \"Membuat akun...\" + span.animate-spin"],
    ["", "Link login: a[href=\"/masuk\"] dengan text \"Masuk\""],
    ["", "Akun seed (TC-016 & TC-020 PENDING): user akan maintain manual"],
]

# ---- styling ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CAT_FILL = {
    "Happy":   PatternFill("solid", fgColor="E2EFDA"),
    "Positif": PatternFill("solid", fgColor="DDEBF7"),
    "Negatif": PatternFill("solid", fgColor="FCE4D6"),
    "Edge":    PatternFill("solid", fgColor="FFF2CC"),
}
SRC_FONT = {
    "PRD":        Font(color="375623"),
    "PRD-ambigu": Font(color="BF8F00", bold=True),
    "Asumsi":     Font(color="C00000", bold=True),
}
STATUS_FILL = {
    "DITUNDA": PatternFill("solid", fgColor="D9D9D9"),
    "PENDING": PatternFill("solid", fgColor="FFE699"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPWRAP = Alignment(vertical="top", wrap_text=True)
WIDTHS = [16, 12, 40, 36, 46, 30, 52, 14, 18, 50]

wb = Workbook()
ws = wb.active
ws.title = "Registrasi"

# Title row
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
t = ws.cell(row=1, column=1, value="Test Case - Onboarding Partner: Form Registrasi (PRD 1 sub-flow A)  |  URL: https://staging.cards.co.id/daftar-akun  |  v2 (post element analysis)")
t.font = Font(bold=True, size=12)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22

# Header row (row 2)
for c, name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=2, column=c, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = BORDER

# Data rows
for r, row in enumerate(ROWS, start=3):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = TOPWRAP
        cell.border = BORDER
    cat = row[1]
    if cat in CAT_FILL:
        ws.cell(row=r, column=2).fill = CAT_FILL[cat]
    src = row[7]
    if src in SRC_FONT:
        ws.cell(row=r, column=8).font = SRC_FONT[src]
    status = row[8]
    if status in STATUS_FILL:
        ws.cell(row=r, column=9).fill = STATUS_FILL[status]

# Column widths
for c, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + len(ROWS)}"

# Legend sheet
ls = wb.create_sheet("Legenda")
ls.column_dimensions["A"].width = 20
ls.column_dimensions["B"].width = 90
for r, (a, b) in enumerate(LEGEND, 1):
    ca = ls.cell(row=r, column=1, value=a)
    cb = ls.cell(row=r, column=2, value=b)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    if a in ("Modul", "URL", "URL post-success", "Total TC", "Kolom Sumber", "Kolom Kategori", "Kolom Status", "Catatan implementasi"):
        ca.font = Font(bold=True)

out = "docs/test-cases/onboarding/TC_Onboarding_Registrasi.xlsx"
wb.save(out)
print("saved:", out, "| rows:", len(ROWS))
