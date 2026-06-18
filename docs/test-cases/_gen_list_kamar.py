# -*- coding: utf-8 -*-
"""Generate Test Case sheet for List Kamar (per CLAUDE.md columns)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = ["ID", "Kategori", "Skenario", "Pre-condition", "Langkah",
           "Test Data", "Expected", "Sumber", "Status", "Catatan"]

# PRD List Kamar:
#   Kolom: Instansi | Nama Kamar | Lokasi | PIC | Status | Dibuat Pada | Aksi (Edit & Hapus)
#   Sort default: terbaru -> terlama (by Dibuat Pada)
#   Filter: Instansi ; Status (Semua/Aktif/Tidak Aktif)
#   Search: Instansi, Nama Kamar, PIC  (Lokasi BUKAN kriteria)
#   Empty: no-data / no-filter-result / no-search-result -> halaman kosong
ROWS = [
    # ---------------- HAPPY ----------------
    ["TC-KMR-LST-001", "Happy",
     "List menampilkan kolom lengkap sesuai PRD",
     "Login admin; ada >=1 data kamar",
     "1) Buka Pengaturan > Akademik > Kamar\n2) Amati header & isi tabel",
     "Seed 1 kamar (Instansi=SDIT)",
     "Tampil kolom: Instansi, Nama Kamar, Lokasi, PIC, Status, Dibuat Pada, Aksi (Edit & Hapus) = 8 sel/baris",
     "PRD", "Belum dijalankan", ""],

    ["TC-KMR-LST-002", "Happy",
     "Setiap baris punya aksi Edit & Hapus",
     "Ada >=1 data kamar",
     "1) Buka List Kamar\n2) Cek ikon aksi pada baris",
     "-",
     "Tiap baris menampilkan ikon Edit (pensil) & Hapus (trash)",
     "PRD", "Belum dijalankan", ""],

    ["TC-KMR-LST-003", "Positif",
     "Urutan default terbaru -> terlama (by Dibuat Pada)",
     "Login admin",
     "1) Tambah kamar baru (nama unik)\n2) Buka List Kamar tanpa filter/search\n3) Amati baris teratas",
     "Instansi=SDIT; Nama=QA<ts><seq>",
     "Kamar yang baru dibuat tampil di BARIS PALING ATAS (newest first)",
     "PRD", "Belum dijalankan", ""],

    # ---------------- FILTER ----------------
    ["TC-KMR-LST-004", "Happy",
     "Filter by Instansi menampilkan hanya kamar instansi tsb",
     "Ada kamar di Instansi SDIT",
     "1) Seed kamar SDIT\n2) Set filter Instansi=SDIT\n3) Amati semua baris",
     "Instansi=SDIT",
     "Semua baris ber-Instansi = SDIT; minimal seed muncul",
     "PRD", "Belum dijalankan", ""],

    ["TC-KMR-LST-005", "Positif",
     "Ganti filter Instansi -> list ter-update",
     "Ada kamar di SDIT & Sekolah Alam",
     "1) Seed kamar SDIT & Sekolah Alam\n2) Filter Instansi=SDIT (amati)\n3) Ganti Instansi=Sekolah Alam (amati)",
     "Instansi A=SDIT; Instansi B=Sekolah Alam",
     "Setelah ganti, semua baris ber-Instansi = instansi yang dipilih terakhir",
     "PRD", "Belum dijalankan", ""],

    ["TC-KMR-LST-006", "Happy",
     "Filter Status=Aktif menampilkan hanya kamar Aktif",
     "Ada kamar berstatus Aktif (default tambah = Aktif)",
     "1) Seed kamar (default Aktif)\n2) Filter Status=Aktif\n3) Amati badge status tiap baris",
     "Status=Aktif",
     "Semua baris berstatus Aktif; minimal seed muncul",
     "PRD", "FAIL",
     "BUG-016: filter Status balik empty state padahal data Aktif ada (seed pasti ada). Filter Status TUNGGAL rusak"],

    ["TC-KMR-LST-007", "Positif",
     "Filter Status=Semua mereset -> tampil semua status",
     "Ada data kamar",
     "1) Filter Status=Aktif\n2) Ganti Status=Semua\n3) Amati list",
     "Status=Semua",
     "List menampilkan semua kamar tanpa batasan status",
     "PRD", "Belum dijalankan", ""],

    ["TC-KMR-LST-008", "Positif",
     "Filter Status=Tidak Aktif menampilkan hanya kamar Tidak Aktif",
     "Ada kamar berstatus Tidak Aktif (dikonfirmasi manual; status diubah via Edit)",
     "1) Filter Status=Tidak Aktif\n2) Amati list & Network (XHR)",
     "Status=Tidak Aktif",
     "Semua baris berstatus Tidak Aktif (data Tidak Aktif ADA)",
     "PRD", "FAIL",
     "BUG-016: filter Status balik empty state padahal data Tidak Aktif ada. Butuh >=1 kamar Tidak Aktif di staging"],

    ["TC-KMR-LST-009", "Negatif",
     "Filter kombinasi Instansi + Status=Aktif",
     "Ada kamar SDIT berstatus Aktif",
     "1) Seed kamar SDIT (Aktif)\n2) Filter Instansi=SDIT\n3) Tambah filter Status=Aktif\n4) Amati list & Network (XHR)",
     "Instansi=SDIT; Status=Aktif",
     "Tampil kamar yang cocok KEDUA filter (minimal seed)",
     "PRD", "FAIL",
     "BUG-016: kombinasi Instansi+Status balik empty padahal data ada (pola BUG-009/014). Di Kamar filter Status tunggal pun rusak"],

    # ---------------- SEARCH ----------------
    ["TC-KMR-LST-010", "Happy",
     "Cari by Nama Kamar",
     "Ada kamar dgn nama unik",
     "1) Seed kamar nama unik\n2) Ketik nama di Cari\n3) Amati hasil",
     "Nama=QA<ts><seq>",
     "List ter-filter menampilkan kamar dgn nama tsb",
     "PRD", "Belum dijalankan", ""],

    ["TC-KMR-LST-011", "Positif",
     "Cari by PIC",
     "Ada kamar dgn PIC terisi (guru instansi)",
     "1) Seed kamar + pilih PIC (guru pertama)\n2) Ketik nama PIC di Cari\n3) Amati hasil",
     "Keyword=<nama PIC seed>",
     "Hasil memuat kamar dgn PIC tsb",
     "PRD", "Belum dijalankan",
     "Bergantung instansi punya guru; bila SDIT tak punya guru -> tahap PIC di-skip"],

    ["TC-KMR-LST-012", "Edge",
     "Cari case-insensitive (huruf kecil) menemukan data",
     "Ada kamar dgn nama unik",
     "1) Seed kamar nama unik (uppercase QA...)\n2) Ketik versi huruf kecil di Cari\n3) Amati hasil",
     "Nama=QA<ts><seq>; keyword=qa<ts><seq>",
     "Search tidak case-sensitive -> kamar tetap ditemukan",
     "Asumsi", "Belum dijalankan",
     "PRD tak eksplisit soal case-insensitive -> asumsi konvensi modul lain"],

    ["TC-KMR-LST-013", "Negatif",
     "Cari keyword tidak ada -> halaman kosong",
     "List Kamar terbuka",
     "1) Ketik keyword acak yang pasti tak match\n2) Amati hasil",
     "Keyword=ZZZQA000NOTEXIST",
     "Tidak ada baris; tampil halaman kosong (empty state)",
     "PRD", "Belum dijalankan", ""],

    ["TC-KMR-LST-014", "Positif",
     "Clear search -> list kembali tampil penuh",
     "Sudah melakukan search",
     "1) Seed kamar\n2) Search nama tsb\n3) Kosongkan kolom Cari\n4) Amati list",
     "Nama=QA<ts><seq>",
     "Setelah dikosongkan, list kembali menampilkan banyak baris (tidak ter-filter)",
     "PRD", "Belum dijalankan", ""],

    # ---------------- EDGE ----------------
    ["TC-KMR-LST-015", "Edge",
     "Cari pakai teks Lokasi -> TIDAK menemukan (Lokasi bukan kriteria search)",
     "Ada kamar dgn Lokasi unik",
     "1) Seed kamar + Lokasi unik\n2) Ketik teks Lokasi tsb di Cari\n3) Amati hasil",
     "Lokasi=LOK<ts><seq> (unik); Nama=QA<ts><seq>",
     "Kamar TIDAK muncul via pencarian Lokasi (PRD: search hanya Instansi/Nama/PIC) -> halaman kosong",
     "Asumsi", "Belum dijalankan",
     "PRD eksplisit sebut 3 kriteria search (tanpa Lokasi). Jika kamar MALAH muncul = behavior beda dari PRD"],
]

LEGEND = [
    ["Kolom Sumber", ""],
    ["PRD", "Skenario diturunkan langsung dari PRD List Kamar"],
    ["PRD-ambigu", "PRD menyinggung tapi tidak jelas/typo -> perlu konfirmasi"],
    ["Asumsi", "Tidak ada di PRD; asumsi berbasis konvensi/fixture/modul lain"],
    ["", ""],
    ["Kolom Kategori", ""],
    ["Happy", "Alur utama sukses sesuai desain"],
    ["Positif", "Variasi valid non-utama"],
    ["Negatif", "Input/kondisi yang harus ditolak / mengembalikan kosong"],
    ["Edge", "Batas/kondisi tepi"],
    ["", ""],
    ["Kolom Status", ""],
    ["Belum dijalankan", "Sudah di-generate, menunggu eksekusi Cypress"],
    ["FAIL", "Gagal saat run; lihat ref BUG-### di Catatan"],
    ["", ""],
    ["Bug ditemukan", ""],
    ["BUG-016", "Filter Status (Aktif/Tidak Aktif), tunggal & kombinasi, balik empty padahal data ada. Pola BUG-009/014 tapi lebih luas (filter status tunggal pun rusak). Lihat TC-006/008/009"],
    ["Catatan", "Tanpa kategori security/injection (sesuai kesepakatan)"],
]

# ---- styling ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CAT_FILL = {
    "Happy":   PatternFill("solid", fgColor="E2EFDA"),
    "Positif": PatternFill("solid", fgColor="DDEBF7"),
    "Negatif": PatternFill("solid", fgColor="FCE4D6"),
    "Edge":    PatternFill("solid", fgColor="FFF2CC"),
}
SRC_FONT = {
    "PRD":        Font(color="375623"),
    "PRD-ambigu": Font(color="BF8F00", bold=True),
    "Asumsi":     Font(color="C00000", bold=True),
}
STATUS_FILL = {
    "FAIL": PatternFill("solid", fgColor="FFC7CE"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPWRAP = Alignment(vertical="top", wrap_text=True)
WIDTHS = [16, 10, 38, 30, 44, 28, 50, 12, 16, 46]

wb = Workbook()
ws = wb.active
ws.title = "List Kamar"

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
t = ws.cell(row=1, column=1, value="Test Case - List Kamar (CARDS School v3)  |  Modul: Pengaturan Akademik > Kamar  |  Sumber PRD: List Kamar")
t.font = Font(bold=True, size=12)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22

for c, name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=2, column=c, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = BORDER

for r, row in enumerate(ROWS, start=3):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = TOPWRAP
        cell.border = BORDER
    cat = row[1]
    if cat in CAT_FILL:
        ws.cell(row=r, column=2).fill = CAT_FILL[cat]
    src = row[7]
    if src in SRC_FONT:
        ws.cell(row=r, column=8).font = SRC_FONT[src]
    status = row[8]
    if status in STATUS_FILL:
        ws.cell(row=r, column=9).fill = STATUS_FILL[status]

for c, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + len(ROWS)}"

ls = wb.create_sheet("Legenda")
ls.column_dimensions["A"].width = 22
ls.column_dimensions["B"].width = 72
for r, (a, b) in enumerate(LEGEND, 1):
    ca = ls.cell(row=r, column=1, value=a)
    cb = ls.cell(row=r, column=2, value=b)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    if a in ("Kolom Sumber", "Kolom Kategori", "Kolom Status", "Bug-target (pola lama)", "Catatan"):
        ca.font = Font(bold=True)

out = "docs/test-cases/TC_Kamar_List.xlsx"
wb.save(out)
print("saved:", out, "| rows:", len(ROWS))
