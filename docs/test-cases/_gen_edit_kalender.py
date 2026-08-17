# -*- coding: utf-8 -*-
"""Generate Test Case sheet for Edit Kalender Akademik (per CLAUDE.md columns).

Catatan PRD:
  PRD Kalender Pendidikan (satu-satunya artefak formal) SECARA EKSPLISIT membahas
  flow Edit dengan 3 field:
    1) Foto header kalender pendidikan (opsional, max 2MB)
    2) Awal minggu dimulai (Minggu / Senin)
    3) Tampilan nama minggu (Minggu / Ahad)

  Namun kenyataan di app (dikonfirmasi user 27 Juni 2026): modal Edit ternyata
  MENAMPILKAN field Instansi & bisa diubah (BUKAN read-only). Ini di luar
  spek PRD -> potensi bug design (kalender bisa "reassign" ke instansi lain
  atau trigger duplicate constraint). Perlu klarifikasi PM/Safki.

  Skenario positif ubah-instansi & negatif duplicate-instansi masuk sebagai
  Sumber=Asumsi (mencover behavior aktual). Skenario field kanonik PRD masuk
  Sumber=PRD.

Modal Edit (proyeksi, sebelum element analysis):
  - Title "Edit Kalender Akademik"
  - Field: Instansi (editable - anomali), Awal Pekan (select), Nama Pekan (select),
           Header (upload optional, max 2MB, dgn preview card + trash kalau existing)
  - Aksi: Simpan / Batal / Close X

Fixture existing di-reuse dari kalender.json (labels/messages/instansi/timeouts).
Aset upload: assets.headerValid (<2MB) & assets.headerOversize (>2MB) - sudah tersedia.

Test data strategi:
  - Pakai instansi `existing` (Sekolah Digital Indonesia (SDI)) sebagai row target edit
    (dari data seeded permanen; sudah punya header image).
  - Instansi `primary` (SMP+) sebagai row alt (Senin/Minggu, no header).
  - Untuk TC "ubah instansi ke yg BELUM punya" pakai instansi `secondary` / `tertiary`
    (perlu konfirmasi ketersediaan saat element analysis).
  - Untuk TC "ubah instansi ke yg SUDAH punya" pakai instansi primary saat editing existing.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = ["ID", "Kategori", "Skenario", "Pre-condition", "Langkah",
           "Test Data", "Expected", "Sumber", "Status", "Catatan"]

ROWS = [
    # ---------------- HAPPY ----------------
    ["TC-KLD-EDT-001", "Happy",
     "Klik icon Edit di row -> modal Edit terbuka, title benar, form pre-populate dengan data existing",
     "Login admin; row target existing di list (mis. Sekolah Digital Indonesia (SDI))",
     "1) Buka List Kalender Akademik\n2) Klik icon Edit (svg.lucide-square-pen) di row target\n3) Amati modal terbuka & isi field",
     "Row: instansi=Sekolah Digital Indonesia (SDI) (existing dgn header image)",
     "Modal terbuka (data-slot=\"dialog-content\") dgn title \"Edit Kalender Akademik\"; field Instansi pre-populate nama instansi row; field Awal Pekan pre-populate value existing; field Nama Pekan pre-populate value existing; preview card header muncul dgn image existing; tombol Simpan/Batal visible",
     "PRD", "Belum dijalankan",
     "Anomali PRD: Instansi editable di modal Edit (belum spek). Element analysis akan konfirmasi field render actual."],

    ["TC-KLD-EDT-002", "Happy",
     "Ubah Awal Pekan (mis. Senin -> Minggu) -> Simpan -> toast success + row col-1 badge update",
     "Login admin; modal Edit terbuka dgn row primary (SMP+, Awal Pekan=Senin)",
     "1) Buka modal Edit row primary\n2) Klik dropdown Awal Pekan -> pilih 'Minggu'\n3) Klik Simpan\n4) Amati toast & badge col-1",
     "Row: primary (SMP+); Awal Pekan lama=Senin; Awal Pekan baru=Minggu",
     "Toast sukses (Sonner data-type=success) muncul dgn text \"Kalender Akademik berhasil diperbarui\"; modal close; row primary col-1 badge update jadi \"Minggu\"; field lain (Nama Pekan, Header) tidak berubah",
     "PRD", "Belum dijalankan",
     "PRD explicit: awal minggu editable (Minggu/Senin)"],

    ["TC-KLD-EDT-003", "Happy",
     "Ubah Nama Pekan (mis. Minggu -> Ahad) -> Simpan -> toast success + row col-2 badge update",
     "Login admin; modal Edit terbuka dgn row primary (SMP+, Nama Pekan=Minggu)",
     "1) Buka modal Edit row primary\n2) Klik dropdown Nama Pekan -> pilih 'Ahad'\n3) Klik Simpan\n4) Amati toast & badge col-2",
     "Row: primary (SMP+); Nama Pekan lama=Minggu; Nama Pekan baru=Ahad",
     "Toast sukses \"Kalender Akademik berhasil diperbarui\"; modal close; row primary col-2 badge update jadi \"Ahad\"; field lain tidak berubah",
     "PRD", "Belum dijalankan",
     "PRD explicit: tampilan nama minggu editable (Minggu/Ahad)"],

    ["TC-KLD-EDT-004", "Happy",
     "Upload header baru pada row TANPA header existing -> preview muncul -> Simpan -> col-3 tampil <img>",
     "Login admin; row target = primary (SMP+, col-3 = '-'); file valid <2MB tersedia (assets.headerValid)",
     "1) Buka modal Edit row primary\n2) Klik area upload / tombol 'Pilih File' -> select file valid\n3) Preview card muncul dgn nama file\n4) Klik Simpan\n5) Amati toast & row col-3",
     "File: cypress/fixtures/kalender/header-valid.png (<2MB)",
     "Preview card muncul di modal setelah selectFile (nama file visible, trash icon visible); klik Simpan -> toast sukses; modal close; row primary col-3 update dari '-' jadi <img alt='header'> (backend generate URL image)",
     "PRD", "Belum dijalankan",
     "PRD explicit: foto header opsional max 2MB"],

    # ---------------- POSITIF ----------------
    ["TC-KLD-EDT-005", "Positif",
     "Ubah semua field sekaligus (Instansi + Awal Pekan + Nama Pekan + Header) -> Simpan -> semua update persist",
     "Login admin; row target = primary; instansi alt yg BELUM punya kalender tersedia (mis. secondary); file valid tersedia",
     "1) Buka modal Edit row primary\n2) Ubah dropdown Instansi ke <secondary>\n3) Ubah dropdown Awal Pekan\n4) Ubah dropdown Nama Pekan\n5) Upload file header baru\n6) Klik Simpan\n7) Amati toast & row",
     "Instansi baru=<secondary>; Awal Pekan baru=Minggu; Nama Pekan baru=Ahad; File header baru",
     "Toast sukses; modal close; row primary lama TIDAK ADA lagi (karena instansi dipindah); row baru dgn instansi=<secondary> muncul dgn Awal Pekan=Minggu, Nama Pekan=Ahad, kol-3 <img>",
     "Asumsi", "Belum dijalankan",
     "Combine semua perubahan; behavior instansi editable perlu klarifikasi PM"],

    ["TC-KLD-EDT-006", "Positif",
     "Remove header existing (klik trash preview) -> Simpan -> row col-3 balik ke '-'",
     "Login admin; row target = existing (SMA, sudah punya header image); TC-004 preview card DIKONFIRMASI ada trash icon (element analysis)",
     "1) Buka modal Edit row existing\n2) Preview card header muncul (image + trash icon)\n3) Klik trash icon di preview\n4) Preview card hilang (state jadi kosong)\n5) Klik Simpan\n6) Amati row col-3",
     "Row: existing (SMA); action: remove header",
     "Preview card hilang setelah klik trash; klik Simpan -> toast sukses; row existing col-3 update dari <img> jadi '-'",
     "PRD-ambigu", "DITUNDA",
     "Fitur remove header BELUM DIKONFIRMASI ada di app - defer sampai element analysis (kalau ga ada, TC diganti jadi 'Replace header')"],

    ["TC-KLD-EDT-007", "Positif",
     "Ubah Instansi ke instansi yg BELUM punya kalender -> row instansi update (behavior anomali)",
     "Login admin; row target = primary (SMP+); instansi alt yg BELUM punya kalender (mis. secondary) tersedia di master",
     "1) Buka modal Edit row primary\n2) Klik dropdown Instansi -> pilih <secondary> (blm punya kalender)\n3) Klik Simpan\n4) Amati toast & list",
     "Instansi lama=primary; instansi baru=secondary",
     "Toast sukses; modal close; row instansi=primary TIDAK ADA lagi; row instansi=secondary muncul dgn field lain (Awal Pekan/Nama Pekan/Header) sama seperti data primary sebelum edit",
     "Asumsi", "Belum dijalankan",
     "ANOMALI PRD - PM/Safki perlu klarifikasi apakah behavior ini by design"],

    ["TC-KLD-EDT-008", "Positif",
     "Klik tombol Batal di modal Edit -> modal close, tidak ada perubahan di list",
     "Login admin; row target = primary di list; modal Edit terbuka",
     "1) Buka modal Edit row primary\n2) Ubah beberapa field (mis. Awal Pekan)\n3) Klik tombol Batal\n4) Amati modal & list",
     "Row: primary; edit uncommitted",
     "Modal close; row primary di list TIDAK ter-update (field tetap sama sebelum edit); toast tidak muncul",
     "Asumsi", "Belum dijalankan",
     "Cancel confirm pattern - standar untuk modul lain"],

    ["TC-KLD-EDT-009", "Positif",
     "Klik icon X (close) di header modal Edit -> modal close, tidak ada perubahan",
     "Login admin; row target di list; modal Edit terbuka",
     "1) Buka modal Edit\n2) Ubah beberapa field\n3) Klik icon X (svg.lucide-x) di header modal\n4) Amati modal & list",
     "Row: primary; edit uncommitted",
     "Modal close; row TIDAK ter-update; toast tidak muncul; sama behavior dgn Batal",
     "Asumsi", "Belum dijalankan",
     "Close X pattern - konsisten"],

    ["TC-KLD-EDT-010", "Positif",
     "Persistence: Edit -> Simpan -> reload halaman -> data tetap update (backend persist)",
     "Login admin; row target di list; edit sukses",
     "1) Buka modal Edit row target\n2) Ubah field (mis. Awal Pekan)\n3) Klik Simpan\n4) Toast sukses + row update\n5) Reload halaman (F5)\n6) Amati row target",
     "Row: primary; field yg diubah: Awal Pekan",
     "Setelah reload, row primary tetap menampilkan value baru Awal Pekan (bukan optimistic UI state, tapi backend persisted)",
     "Asumsi", "Belum dijalankan",
     "assertPersisted standard pattern"],

    ["TC-KLD-EDT-011", "Positif",
     "Buka modal Edit -> tutup -> buka lagi: form reset ke data terkini dari server (bukan stale)",
     "Login admin; row target di list",
     "1) Buka modal Edit row target -> catat nilai field\n2) Ubah beberapa field\n3) Klik Batal (tanpa Simpan)\n4) Buka lagi modal Edit row yg sama\n5) Amati field value",
     "Row: primary; action: uncommitted edit",
     "Modal buka ulang -> field value SAMA dgn nilai server terkini (bukan draft/stale dari sesi sebelumnya); edit uncommitted hilang",
     "Asumsi", "FAIL",
     "BUG-030 - Form Edit persist draft value setelah Batal + reopen (tidak reset). Assertion tetap correct expected behavior (form reset). Test FAIL sampai bug fix."],

    # ---------------- NEGATIF ----------------
    ["TC-KLD-EDT-012", "Negatif",
     "Upload header >2MB di Edit -> alert inline muncul, Simpan tidak apply file",
     "Login admin; row target di list; file oversize tersedia (assets.headerOversize)",
     "1) Buka modal Edit row target\n2) Klik area upload -> select file >2MB\n3) Amati alert inline\n4) (Optional) Klik Simpan -> observasi apa file benar-benar tidak ter-upload",
     "File: cypress/fixtures/kalender/header-oversize.jpg (>2MB)",
     "Alert inline (data-slot=\"alert\") muncul dgn title 'Gagal mengunggah file' & desc 'File melebihi ukuran maksimal 2MB.'; preview card TIDAK muncul; kalaupun Simpan tetap diklik, header row TIDAK update (BE tolak / FE prevent)",
     "PRD", "Belum dijalankan",
     "PRD explicit: max 2MB"],

    ["TC-KLD-EDT-013", "Negatif",
     "Ubah Instansi ke instansi yg SUDAH punya kalender -> duplicate error (behavior anomali)",
     "Login admin; row target = primary (SMP+); row lain dgn instansi=existing (SMA) sudah ada di list",
     "1) Buka modal Edit row primary\n2) Klik dropdown Instansi -> pilih <existing> (SMA, sudah punya kalender)\n3) Klik Simpan\n4) Amati toast error / behavior",
     "Instansi lama=primary; instansi baru=existing (target sudah punya kalender)",
     "Toast error / dialog error dgn text 'Pengaturan kalender untuk office ini sudah ada' (reuse constraint dari Tambah); modal TIDAK close (biar user bisa koreksi); row TIDAK ter-update",
     "Asumsi", "Belum dijalankan",
     "ANOMALI PRD - duplicate constraint expected sesuai backend"],

    # ---------------- EDGE ----------------
    ["TC-KLD-EDT-014", "Edge",
     "Simpan tanpa perubahan (buka Edit -> langsung Simpan) -> discovery behavior",
     "Login admin; row target di list",
     "1) Buka modal Edit row target\n2) TIDAK ubah field apapun\n3) Klik Simpan\n4) Amati behavior (allow / disabled / no-op)",
     "Row: primary; edit: no-change",
     "Behavior discovery: (a) Simpan disabled sampai ada perubahan; ATAU (b) BE tetap accept PUT & toast sukses muncul; ATAU (c) FE cegah request (no-op). Test log actual behavior sebagai info.",
     "Asumsi", "Belum dijalankan",
     "Cypress cek actual, output jadi info; kalau (a) atau (c) direkomendasikan sebagai UX terbaik"],
]

LEGEND = [
    ["Kolom Sumber", ""],
    ["PRD", "Skenario diturunkan langsung dari PRD Kalender Pendidikan"],
    ["PRD-ambigu", "PRD menyinggung tapi tidak detail; turunan logis"],
    ["Asumsi", "PRD tidak menyebut sama sekali; behavior turunan / anomali actual app"],
    ["", ""],
    ["Kolom Kategori", ""],
    ["Happy", "Alur utama sukses sesuai desain PRD"],
    ["Positif", "Variasi valid non-utama (all-field, cancel, close, persist, remove-header)"],
    ["Negatif", "Input invalid / constraint violation (oversize, duplicate instansi)"],
    ["Edge", "Behavior discovery (no-change save)"],
    ["", ""],
    ["Kolom Status", ""],
    ["Belum dijalankan", "Sudah di-generate, menunggu eksekusi Cypress"],
    ["DITUNDA", "Defer sampai element analysis konfirmasi UI element terkait"],
    ["FAIL", "Diketahui FAIL karena bug app (lihat kolom Catatan untuk BUG-### ref)"],
    ["", ""],
    ["Anomali PRD - Instansi editable", "PRD sebut edit 3 field (Header/AwalMinggu/NamaMinggu). Actual: Instansi juga editable (konfirm user 27 Juni 2026). Perlu klarifikasi PM/Safki: by design atau leak?"],
    ["", "Kalau by design: TC-005, TC-007, TC-013 tetap valid. Kalau bug: TC-005 & TC-007 & TC-013 jadi negatif (Instansi hrsnya locked)"],
    ["Fitur Remove Header", "BELUM DIKONFIRMASI ada di app. TC-006 DITUNDA sampai element analysis konfirmasi trash icon di preview modal Edit. Kalau ga ada -> ganti jadi Replace Header (upload baru menimpa lama)."],
    ["", ""],
    ["Catatan", "Test data reuse fixture kalender.json (instansi/labels/messages/assets)"],
    ["", "Element analysis SETELAH ACC TC sheet (memory rule: STOP & tunggu ACC sebelum element analysis)"],
    ["", "Modul referensi pattern: Kamar/Kelas/Tingkat (semua punya Edit sheet serupa)"],
    ["", "Tanpa kategori security/injection (CLAUDE.md kesepakatan)"],
]

# ---- styling (copy pattern dari _gen_list_kalender.py) ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CAT_FILL = {
    "Happy":   PatternFill("solid", fgColor="E2EFDA"),
    "Positif": PatternFill("solid", fgColor="DDEBF7"),
    "Negatif": PatternFill("solid", fgColor="FCE4D6"),
    "Edge":    PatternFill("solid", fgColor="FFF2CC"),
}
SRC_FONT = {
    "PRD":        Font(color="375623"),
    "PRD-ambigu": Font(color="BF8F00", bold=True),
    "Asumsi":     Font(color="C00000", bold=True),
}
STATUS_FILL = {
    "DITUNDA": PatternFill("solid", fgColor="D9D9D9"),
    "FAIL":    PatternFill("solid", fgColor="F4B084"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPWRAP = Alignment(vertical="top", wrap_text=True)
WIDTHS = [16, 12, 42, 36, 48, 28, 56, 14, 16, 50]

wb = Workbook()
ws = wb.active
ws.title = "Edit Kalender"

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
t = ws.cell(row=1, column=1, value="Test Case - Edit Kalender Akademik (CARDS School v3)  |  Modul: Pengaturan Akademik > Kalender Akademik  |  Sumber campur PRD (3 field kanonik) & Asumsi (anomali Instansi editable)")
t.font = Font(bold=True, size=12)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22

for c, name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=2, column=c, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = BORDER

for r, row in enumerate(ROWS, start=3):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = TOPWRAP
        cell.border = BORDER
    cat = row[1]
    if cat in CAT_FILL:
        ws.cell(row=r, column=2).fill = CAT_FILL[cat]
    src = row[7]
    if src in SRC_FONT:
        ws.cell(row=r, column=8).font = SRC_FONT[src]
    status = row[8]
    if status in STATUS_FILL:
        ws.cell(row=r, column=9).fill = STATUS_FILL[status]

for c, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + len(ROWS)}"

ls = wb.create_sheet("Legenda")
ls.column_dimensions["A"].width = 30
ls.column_dimensions["B"].width = 78
for r, (a, b) in enumerate(LEGEND, 1):
    ca = ls.cell(row=r, column=1, value=a)
    cb = ls.cell(row=r, column=2, value=b)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    if a in ("Kolom Sumber", "Kolom Kategori", "Kolom Status", "Catatan",
             "Anomali PRD - Instansi editable", "Fitur Remove Header"):
        ca.font = Font(bold=True)

out = "docs/test-cases/TC_Kalender_Edit.xlsx"
wb.save(out)
print("saved:", out, "| rows:", len(ROWS))
