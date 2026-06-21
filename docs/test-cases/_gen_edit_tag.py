# -*- coding: utf-8 -*-
"""Generate Test Case sheet for Edit Tag (per CLAUDE.md columns)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = ["ID", "Kategori", "Skenario", "Pre-condition", "Langkah",
           "Test Data", "Expected", "Sumber", "Status", "Catatan"]

# Form Edit Tag (PRD):
#   a. Instansi*    (required, select)
#   b. Nama Tag*    (required, text)
#   c. Kode Tag*    (required, text)
#   d. Status*      (required: Aktif / Tidak Aktif)
#   e. Tipe Member* (required, dropdown: Semua / Siswa / Guru & Staff)
#
# Beda dgn Tambah Tag:
#   - Edit punya field Status (Tambah tidak; Tambah default=Aktif)
#   - Field di-pre-fill dari data existing
#
# Bug-pola yg sengaja dicek kemunculannya di Edit Tag (hasil run + element analysis):
#   - BUG-024 (Edit kode duplikat -> FE silent) -- ter-konfirmasi via element analysis -> TC-017
#   - BUG-022 (no maxlength >255 -> raw SQL + FE silent) -- ter-reproduce di Edit -> TC-018/019
#   - BUG-021 (toast description salah copy "data Kelas") -- ter-reproduce di Edit toast.
#     Mitigasi: assertSuccessToast() di POM hanya assert [data-title], bukan description
#
# DEFERRED (per instruksi user, butuh fitur konsumen yg belum ada):
#   - "Tag berhasil dibuat muncul di data diri siswa, presensi, tagihan"
#   - "Tipe member ngatur visibility tag di feature konsumen"
ROWS = [
    # ---------------- DISPLAY & PRE-FILL ----------------
    ["TC-TAG-EDT-001", "Happy",
     "Form Edit pre-filled dgn data existing (Instansi/Nama/Kode/Status/Tipe)",
     "Login admin; ada tag di list (seed)",
     "1) Seed tag QA (Instansi=SDIT, Tipe=Semua, default Status=Aktif)\n2) Klik ikon Edit pada baris tsb\n3) Amati nilai semua field di modal",
     "Tag QA<ts><seq>",
     "Modal terbuka berjudul \"Edit Tag\"; SEMUA field terisi nilai existing: Instansi=SDIT, Nama=QA..., Kode=QA..., Status=Aktif, Tipe Member=Semua",
     "PRD", "Belum dijalankan",
     "Asumsi: field Status & Tipe Member ditampilkan sbg select dgn value preselected"],

    ["TC-TAG-EDT-002", "Happy",
     "Dropdown Tipe Member di Edit menampilkan tepat 3 opsi PRD (urutan UI bebas)",
     "Modal Edit terbuka",
     "1) Buka modal Edit tag\n2) Klik dropdown Tipe Member\n3) Hitung & catat opsi yg muncul",
     "-",
     "Dropdown menampilkan TEPAT 3 opsi: Semua, Siswa, Guru & Staff (set, urutan UI configurable)",
     "PRD", "Belum dijalankan",
     "Pola sama TC-TAG-ADD-007 (Tambah). Pakai assertion SET sorted, bukan array order"],

    # ---------------- HAPPY (single-field edit) ----------------
    ["TC-TAG-EDT-003", "Happy",
     "Edit Nama Tag berhasil -> perubahan tersimpan & tampil di list",
     "Ada tag QA di list",
     "1) Seed tag QA\n2) Buka Edit\n3) Ubah Nama Tag ke nilai baru unik\n4) Klik \"Simpan\"",
     "Nama baru=QA<ts><seq>-EDIT",
     "Toast sukses (\"Tag berhasil diperbarui\" atau setara); modal tertutup; kolom Nama Tag di list ter-update; reload -> nama baru tetap ada",
     "PRD", "Belum dijalankan",
     "Teks toast = Asumsi (PRD hanya sebut \"pesan sukses\"). Assertion title-only (mitigasi BUG-021)"],

    ["TC-TAG-EDT-004", "Happy",
     "Edit Kode Tag berhasil -> perubahan tersimpan & tampil di list",
     "Ada tag QA di list",
     "1) Seed tag QA\n2) Buka Edit\n3) Ubah Kode Tag ke nilai baru unik\n4) Klik \"Simpan\"",
     "Kode baru=QA<ts><seq>-EDIT",
     "Toast sukses; modal tutup; kolom Kode Tag di baris tag ter-update ke kode baru",
     "PRD", "Belum dijalankan",
     "Verifikasi kolom Kode (td ke-3)"],

    ["TC-TAG-EDT-005", "Happy",
     "Edit Tipe Member berhasil -> badge Tipe di list ter-update",
     "Ada tag QA Tipe=Semua",
     "1) Seed tag QA Tipe=Semua\n2) Buka Edit\n3) Ganti Tipe Member ke \"Siswa\"\n4) Klik \"Simpan\"",
     "Tipe lama=Semua; Tipe baru=Siswa",
     "Toast sukses; badge Tipe di baris tag berubah jadi \"Siswa\"",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-EDT-006", "Happy",
     "Edit Status Aktif -> Tidak Aktif -> badge di list berubah",
     "Ada tag QA berstatus Aktif",
     "1) Seed tag (default Aktif)\n2) Buka Edit\n3) Ubah Status ke \"Tidak Aktif\"\n4) Klik \"Simpan\"",
     "Status awal=Aktif; Status baru=Tidak Aktif",
     "Toast sukses; badge Status di list jadi \"Tidak Aktif\"",
     "PRD", "Belum dijalankan",
     "Verifikasi badge Status (badge terakhir di row)"],

    ["TC-TAG-EDT-007", "Happy",
     "Edit Status Tidak Aktif -> Aktif -> badge di list berubah (reverse)",
     "Ada tag QA Tidak Aktif (hasil TC-006)",
     "1) Pakai tag dari TC-006 (Tidak Aktif)\n2) Buka Edit\n3) Ubah Status ke \"Aktif\"\n4) Klik \"Simpan\"",
     "Status awal=Tidak Aktif; Status baru=Aktif",
     "Toast sukses; badge Status jadi \"Aktif\"",
     "PRD", "Belum dijalankan",
     "Dependent: TC-006 jalan dulu, atau seed tag Tidak Aktif via Edit di before()"],

    # ---------------- POSITIF (multi-field, no-regression, cancel paths) ----------------
    ["TC-TAG-EDT-008", "Positif",
     "Edit semua field sekaligus (Nama+Kode+Status+Tipe) -> semua tersimpan",
     "Ada tag QA",
     "1) Seed tag QA\n2) Buka Edit\n3) Ubah Nama+Kode+Status+Tipe\n4) Simpan",
     "Semua field baru unik",
     "Toast sukses; di list semua kolom ter-update sesuai input",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-EDT-009", "Positif",
     "Edit hanya 1 field (Nama) -> field lain tetap (no regression)",
     "Ada tag QA dgn semua field terisi",
     "1) Seed tag QA Tipe=Siswa Status=Aktif\n2) Buka Edit\n3) Ubah HANYA Nama Tag\n4) Simpan",
     "Nama baru saja",
     "Nama ter-update; Kode/Status/Tipe/Instansi TIDAK berubah di list",
     "PRD", "Belum dijalankan",
     "PRD validasi b/c: implicit no-regression -> jangan ngubah field yg tidak diutak-atik"],

    ["TC-TAG-EDT-010", "Positif",
     "Klik Batal -> tidak menyimpan perubahan, kembali ke list (PRD validasi b)",
     "Ada tag QA di list",
     "1) Buka Edit\n2) Ubah Nama Tag ke nilai baru\n3) Klik \"Batal\"",
     "Nama edit=QA<ts><seq>-CANCEL",
     "Modal tertutup; baris tag di list TIDAK berubah (nama lama tetap)",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-EDT-011", "Positif",
     "Klik X close -> tidak menyimpan perubahan",
     "Ada tag QA di list",
     "1) Buka Edit\n2) Ubah Nama Tag\n3) Klik tombol X (close icon)",
     "Nama edit=QA<ts><seq>-X",
     "Modal tertutup; data tidak berubah",
     "Asumsi", "Belum dijalankan",
     "PRD tak eksplisit soal X, asumsi konvensi modul lain"],

    ["TC-TAG-EDT-012", "Positif",
     "Ganti Instansi -> tag pindah ke instansi baru (field editable)",
     "Ada tag QA di Instansi SDIT",
     "1) Seed tag SDIT\n2) Buka Edit\n3) Ganti Instansi ke \"Sekolah Alam\"\n4) Simpan",
     "Instansi lama=SDIT; baru=Sekolah Alam",
     "Toast sukses; kolom Instansi di baris tag jadi \"Sekolah Alam\"",
     "PRD-ambigu", "Belum dijalankan",
     "Konfirmasi via element analysis: apakah Instansi editable di Edit modal? Beberapa modul read-only. Adjust TC kalau ternyata read-only"],

    # ---------------- NEGATIF (required) ----------------
    ["TC-TAG-EDT-013", "Negatif",
     "Clear Nama Tag -> Simpan -> error required, tidak tersimpan",
     "Ada tag QA",
     "1) Buka Edit\n2) Clear field Nama Tag\n3) Klik \"Simpan\"",
     "Nama=\"\"",
     "Pesan error \"Nama Tag wajib diisi\" (atau setara); modal tetap terbuka; data lama tidak berubah",
     "PRD", "Belum dijalankan",
     "PRD validasi a: required wajib diisi"],

    ["TC-TAG-EDT-014", "Negatif",
     "Clear Kode Tag -> Simpan -> error required, tidak tersimpan",
     "Ada tag QA",
     "1) Buka Edit\n2) Clear field Kode Tag\n3) Klik \"Simpan\"",
     "Kode=\"\"",
     "Pesan error \"Kode Tag wajib diisi\"; modal tetap terbuka; data lama tidak berubah",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-EDT-015", "Negatif",
     "Isi Nama Tag dgn whitespace saja -> error required",
     "Ada tag QA",
     "1) Buka Edit\n2) Clear Nama, ketik \"     \" (5 spasi)\n3) Klik \"Simpan\"",
     "Nama=\"     \"",
     "Pesan error required; data tidak tersimpan",
     "Asumsi", "Belum dijalankan",
     "Asumsi konsisten BUG-trim modul lain"],

    ["TC-TAG-EDT-016", "Negatif",
     "Isi Kode Tag dgn whitespace saja -> error required",
     "Ada tag QA",
     "1) Buka Edit\n2) Clear Kode, ketik \"     \"\n3) Klik \"Simpan\"",
     "Kode=\"     \"",
     "Pesan error required; data tidak tersimpan",
     "Asumsi", "Belum dijalankan", ""],

    # ---------------- NEGATIF (duplikat code, instansi sama) ----------------
    ["TC-TAG-EDT-017", "Negatif",
     "Edit Kode Tag jadi duplikat (instansi sama) -> pesan error, perubahan tidak tersimpan",
     "Ada 2 tag QA di Instansi SDIT (Tag A & Tag B)",
     "1) Seed 2 tag SDIT (Tag A kode K1, Tag B kode K2)\n2) Buka Edit Tag B\n3) Ubah Kode B jadi K1 (duplikat A)\n4) Klik \"Simpan\"\n5) Monitor Network (XHR)",
     "Kode B baru = Kode A",
     "Toast/inline error \"Kode tag sudah digunakan di instansi ini\" (PRD validasi d); modal tetap terbuka; data Tag B di list TIDAK berubah",
     "PRD", "Belum dijalankan",
     "KANDIDAT BUG-019 (Kamar Edit) / BUG-020 (Tag Tambah): FE silent saat duplikat - BE balikin error tapi user tidak dapat feedback. Assertion PRD-correct supaya fail sampai bug fixed"],

    # ---------------- EDGE (boundaries) ----------------
    ["TC-TAG-EDT-018", "Edge",
     "Edit Nama Tag >255 char -> dibatasi/validasi FE",
     "Ada tag QA",
     "1) Buka Edit\n2) Clear Nama, ketik string ~300 char\n3) Klik \"Simpan\"\n4) Monitor UI & Network",
     "Nama panjang=300 char",
     "FE membatasi (maxlength) ATAU menampilkan pesan validasi ramah \"Nama Tag maksimal 255 karakter\"; data tidak tersimpan kalau melebihi batas",
     "Asumsi", "FAIL",
     "BUG-022 ter-reproduce di Edit: FE silent saat input >255 char (raw SQL error dari BE, FE diam). Pola identik Tambah. Assertion PRD-correct -> tetap FAIL sampai FE catch & batasi maxlength"],

    ["TC-TAG-EDT-019", "Edge",
     "Edit Kode Tag >255 char -> dibatasi/validasi FE",
     "Ada tag QA",
     "1) Buka Edit\n2) Clear Kode, ketik string ~300 char\n3) Klik \"Simpan\"",
     "Kode panjang=300 char",
     "FE membatasi/validasi; data tidak tersimpan kalau melebihi batas",
     "Asumsi", "FAIL",
     "BUG-022 ter-reproduce di Edit sisi Kode (pola sama)"],

    ["TC-TAG-EDT-020", "Edge",
     "Edit Nama Tag jadi 1 karakter -> tersimpan (min boundary)",
     "Ada tag QA",
     "1) Buka Edit\n2) Clear Nama, ketik \"A\"\n3) Klik \"Simpan\"",
     "Nama=\"A\"",
     "Toast sukses; baris di list menampilkan Nama=\"A\"",
     "Asumsi", "Belum dijalankan",
     "Konsisten TC Tambah Tag min-boundary"],

    ["TC-TAG-EDT-021", "Edge",
     "Edit Kode Tag jadi 1 karakter -> tersimpan (min boundary)",
     "Ada tag QA",
     "1) Buka Edit\n2) Clear Kode, ketik \"A\"\n3) Klik \"Simpan\"",
     "Kode=\"A\"",
     "Toast sukses; kolom Kode jadi \"A\"",
     "Asumsi", "Belum dijalankan", ""],

    # ---------------- PERSISTENCE ----------------
    ["TC-TAG-EDT-022", "Positif",
     "Persistence: reload setelah Edit -> perubahan tetap ada",
     "Sudah berhasil Edit (gabung dgn TC-003 atau seed segar)",
     "1) Seed + Edit Nama tag\n2) Reload halaman List Tag\n3) Cari nama baru",
     "Nama baru hasil Edit",
     "Setelah reload, nama baru tetap muncul di list (data persist di BE)",
     "PRD", "Belum dijalankan",
     "Verifikasi via assertPersisted-style"],

    # ---------------- DEFERRED (asosiasi Member -- per instruksi user) ----------------
    ["TC-TAG-EDT-023", "Positif",
     "Tag hasil Edit (Tipe=Siswa) muncul di picker Tag pada Data Diri Siswa",
     "Tag QA sudah di-Edit Tipe=Siswa",
     "1) Edit tag QA Tipe=Siswa\n2) Buka form Data Diri Siswa\n3) Buka picker Tag\n4) Cek tag QA muncul",
     "Tag QA Tipe=Siswa",
     "Tag muncul & dapat dipilih di picker Data Siswa, TIDAK muncul di picker Data Guru/Staff",
     "PRD", "DITUNDA",
     "DEFERRED per instruksi user: asosiasi cross-fitur (Data Siswa/Presensi/Tagihan) belum bisa diakses"],

    ["TC-TAG-EDT-024", "Positif",
     "Ubah Tipe Member tag -> visibility di picker feature konsumen ikut berubah",
     "Tag QA awal Tipe=Semua",
     "1) Edit tag dari Tipe=Semua ke Tipe=Guru & Staff\n2) Buka picker Tag di Data Siswa -> harusnya HILANG\n3) Buka picker Tag di Data Guru/Staff -> harusnya MUNCUL",
     "Tipe lama=Semua; baru=Guru & Staff",
     "Tag berpindah konteks: hilang dari picker Data Siswa, muncul di picker Data Guru/Staff",
     "PRD", "DITUNDA",
     "DEFERRED per instruksi user: butuh modul Member/Data Siswa/Data Guru-Staff yg belum tersedia"],
]

LEGEND = [
    ["Kolom Sumber", ""],
    ["PRD", "Skenario diturunkan langsung dari PRD Edit Tag"],
    ["PRD-ambigu", "PRD menyinggung tapi tidak jelas -> perlu konfirmasi via element analysis"],
    ["Asumsi", "Tidak ada di PRD; asumsi berbasis konvensi/fixture/modul lain"],
    ["", ""],
    ["Kolom Kategori", ""],
    ["Happy", "Alur utama sukses sesuai desain"],
    ["Positif", "Variasi valid non-utama / cancel paths / no-regression"],
    ["Negatif", "Input/kondisi yang harus mengembalikan error"],
    ["Edge", "Batas/kondisi tepi (maxlength, min-length, dll)"],
    ["", ""],
    ["Kolom Status", ""],
    ["Belum dijalankan", "Sudah di-generate, menunggu eksekusi Cypress"],
    ["DITUNDA", "Sengaja belum di-run; menunggu fitur konsumen (Data Siswa / Data Guru-Staff / Member) tersedia"],
    ["FAIL", "Gagal saat run; lihat ref BUG-### di Catatan"],
    ["", ""],
    ["Bug ter-konfirmasi (Edit Tag)", ""],
    ["BUG-024 (Edit kode duplikat)", "FE silent saat duplikat -- BE balikin error, FE diam. TC-017 sengaja FAIL"],
    ["BUG-022 (no maxlength)", "Ter-reproduce di Edit: input >255 char -> raw SQL + FE silent. TC-018/019 FAIL"],
    ["BUG-021 (toast description)", "Description toast Edit juga salah copy \"data Kelas\". Mitigasi: assertion [data-title] only"],
    ["", ""],
    ["Catatan", "Tanpa kategori security/injection (sesuai kesepakatan)"],
    ["", "Form Edit punya 5 field required: Instansi, Nama Tag, Kode Tag, Status, Tipe Member"],
    ["", "Beda dgn Tambah: Edit punya field Status (Tambah tidak); semua field di-pre-fill dari data existing"],
    ["", "DEFERRED: kemunculan tag di picker Data Siswa/Guru/Staff & filter tipe member - butuh fitur konsumen yg belum ada"],
]

# ---- styling (mirror _gen_list_tag.py) ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CAT_FILL = {
    "Happy":   PatternFill("solid", fgColor="E2EFDA"),
    "Positif": PatternFill("solid", fgColor="DDEBF7"),
    "Negatif": PatternFill("solid", fgColor="FCE4D6"),
    "Edge":    PatternFill("solid", fgColor="FFF2CC"),
}
SRC_FONT = {
    "PRD":        Font(color="375623"),
    "PRD-ambigu": Font(color="BF8F00", bold=True),
    "Asumsi":     Font(color="C00000", bold=True),
}
STATUS_FILL = {
    "DITUNDA": PatternFill("solid", fgColor="D9D9D9"),
    "FAIL":    PatternFill("solid", fgColor="FFC7CE"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPWRAP = Alignment(vertical="top", wrap_text=True)
WIDTHS = [16, 10, 38, 30, 44, 28, 50, 12, 16, 46]

wb = Workbook()
ws = wb.active
ws.title = "Edit Tag"

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
t = ws.cell(row=1, column=1, value="Test Case - Edit Tag (CARDS School v3)  |  Modul: Pengaturan Akademik > Tag  |  Sumber PRD: Edit Tag")
t.font = Font(bold=True, size=12)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22

for c, name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=2, column=c, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = BORDER

for r, row in enumerate(ROWS, start=3):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = TOPWRAP
        cell.border = BORDER
    cat = row[1]
    if cat in CAT_FILL:
        ws.cell(row=r, column=2).fill = CAT_FILL[cat]
    src = row[7]
    if src in SRC_FONT:
        ws.cell(row=r, column=8).font = SRC_FONT[src]
    status = row[8]
    if status in STATUS_FILL:
        ws.cell(row=r, column=9).fill = STATUS_FILL[status]

for c, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + len(ROWS)}"

ls = wb.create_sheet("Legenda")
ls.column_dimensions["A"].width = 32
ls.column_dimensions["B"].width = 72
for r, (a, b) in enumerate(LEGEND, 1):
    ca = ls.cell(row=r, column=1, value=a)
    cb = ls.cell(row=r, column=2, value=b)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    if a in ("Kolom Sumber", "Kolom Kategori", "Kolom Status",
             "Kandidat bug-pola (cek saat run)", "Catatan"):
        ca.font = Font(bold=True)

out = "docs/test-cases/TC_Tag_Edit.xlsx"
wb.save(out)
print("saved:", out, "| rows:", len(ROWS))
