# -*- coding: utf-8 -*-
"""Generate Test Case sheet for Tambah Tag (per CLAUDE.md columns)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = ["ID", "Kategori", "Skenario", "Pre-condition", "Langkah",
           "Test Data", "Expected", "Sumber", "Status", "Catatan"]

# Form Tambah Tag (UI aktual, mengikuti label yang dirender):
#   a. Instansi*      (required, select)
#   b. Nama Tag*      (required, text, input[name=name])
#   c. Kode Tag*      (required, text, input[name=code])
#   d. Tipe Anggota*  (required, select)   <-- PRD nyebut "Tipe Member", UI render "Tipe Anggota" (label configurable)
#   Opsi Tipe Anggota: Semua | Siswa | Guru & Staff (value=ALL | STUDENT | TEACHER) — label configurable, ikut PRD
#
# Validation PRD vs realita UI:
#   - Field required wajib diisi -> inline error "<nama field> wajib diisi" (OK)
#   - Batal -> kembali ke list tanpa menambah data (OK)
#   - Simpan -> toast success "Tag berhasil ditambahkan" (OK)
#       NOTE: toast description salah copy "data Kelas" -> BUG-021 (jangan di-lock di assert)
#   - Kode Tag duplikat pada instansi sama -> PRD harus tampil error, REALITA FE silent -> BUG-020
#   - Tag dipakai utk data siswa/presensi/tagihan -> DEFERRED (fitur konsumen belum ada)
ROWS = [
    # ---------------- HAPPY ----------------
    ["TC-TAG-ADD-001", "Happy",
     "Tambah tag valid lengkap (semua field required terisi)",
     "Login admin; halaman List Tag terbuka",
     "1) Klik \"Tambah Tag\"\n2) Pilih Instansi\n3) Isi Nama Tag\n4) Isi Kode Tag\n5) Pilih Tipe Anggota = Semua\n6) Klik \"Simpan\"",
     "Instansi=SDIT; Nama=QA<6digit-ts><seq>; Kode=QA<ts><seq>; Tipe=Semua",
     "Toast sukses \"Tag berhasil ditambahkan\"; modal tertutup; baris tag baru muncul di list",
     "PRD", "Belum dijalankan",
     "Teks toast = Asumsi (PRD hanya \"pesan succes\"). Semua field PRD required -> wajib semua diisi"],

    ["TC-TAG-ADD-002", "Happy",
     "Tambah tag dengan Tipe Anggota = Siswa",
     "Login admin; List Tag terbuka",
     "1) Klik \"Tambah Tag\"\n2) Pilih Instansi\n3) Isi Nama & Kode Tag\n4) Pilih Tipe Anggota = Siswa\n5) Klik \"Simpan\"",
     "Instansi=SDIT; Nama=QA<ts><seq>; Kode=QA<ts><seq>; Tipe=Siswa",
     "Toast sukses; baris baru muncul dengan Tipe Anggota = Siswa (badge sesuai)",
     "PRD", "Belum dijalankan",
     "Verifikasi Tipe Anggota tersimpan -> reload utk cek persist"],

    ["TC-TAG-ADD-003", "Happy",
     "Tambah tag dengan Tipe Anggota = Guru & Staff",
     "Login admin; List Tag terbuka",
     "1) Klik \"Tambah Tag\"\n2) Pilih Instansi\n3) Isi Nama & Kode Tag\n4) Pilih Tipe Anggota = Guru & Staff\n5) Klik \"Simpan\"",
     "Instansi=SDIT; Nama=QA<ts><seq>; Kode=QA<ts><seq>; Tipe=Guru & Staff",
     "Toast sukses; baris baru muncul dengan Tipe Anggota = Guru & Staff",
     "PRD", "Belum dijalankan", ""],

    # ---------------- POSITIF ----------------
    ["TC-TAG-ADD-004", "Positif",
     "Form Tambah Tag terbuka dengan field kosong",
     "Login admin; List Tag terbuka",
     "1) Klik \"Tambah Tag\"\n2) Amati kondisi awal form",
     "-",
     "Modal \"Tambah Tag\" tampil; Instansi placeholder, Nama Tag kosong, Kode Tag kosong, Tipe Anggota placeholder; tombol Simpan & Batal ada",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-ADD-005", "Positif",
     "Batal menutup form tanpa menambah data",
     "Modal Tambah Tag terbuka",
     "1) Klik \"Tambah Tag\"\n2) Klik \"Batal\"",
     "-",
     "Modal tertutup; kembali ke List Tag; tidak ada data baru",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-ADD-006", "Positif",
     "Isi form valid lalu Batal -> data tidak tersimpan",
     "Modal Tambah Tag terbuka",
     "1) Isi seluruh field valid (Instansi+Nama+Kode+Tipe)\n2) Klik \"Batal\"\n3) Cari nama tsb di list",
     "Instansi=SDIT; Nama=QA<ts><seq>; Kode=QA<ts><seq>; Tipe=Semua",
     "Modal tertutup; nama tag TIDAK muncul di list (tidak tersimpan)",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-ADD-007", "Positif",
     "Dropdown Tipe Anggota menampilkan tepat 3 opsi sesuai PRD (urutan UI bebas)",
     "Modal Tambah Tag terbuka",
     "1) Buka dropdown Tipe Anggota\n2) Amati daftar opsi (cek SET tepat 3 opsi, urutan diabaikan)",
     "-",
     "Dropdown berisi tepat 3 opsi PRD: Semua, Siswa, Guru & Staff (value=ALL/STUDENT/TEACHER). Urutan UI configurable di Pengaturan, jadi tidak di-lock.",
     "PRD", "Belum dijalankan",
     "Label & urutan opsi configurable di app. Assertion cek SET (sorted), bukan array order"],

    ["TC-TAG-ADD-008", "Positif",
     "Tambah beberapa tag berbeda dalam 1 instansi",
     "Login admin",
     "1) Tambah tag A (Instansi=SDIT)\n2) Tambah tag B (Instansi=SDIT)",
     "Nama & Kode A != B (Instansi sama)",
     "Kedua tag sukses ditambahkan; keduanya muncul di list",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-ADD-009", "Positif",
     "Kode tag sama pada instansi berbeda -> sukses",
     "Login admin",
     "1) Tambah Kode X @ Instansi SDIT\n2) Tambah Kode X @ Instansi Sekolah Alam",
     "Kode X sama; Instansi berbeda",
     "Keduanya sukses (PRD batasi duplikat hanya pada \"instansi yang sama\")",
     "PRD", "Belum dijalankan",
     "PRD: \"code tag dengan instansi yang sama\" -> beda instansi diasumsikan boleh"],

    ["TC-TAG-ADD-010", "Positif",
     "Nama tag sama pada instansi berbeda -> sukses",
     "Login admin",
     "1) Tambah Nama X @ Instansi SDIT\n2) Tambah Nama X @ Instansi Sekolah Alam",
     "Nama X sama; Kode berbeda; Instansi berbeda",
     "Keduanya sukses (PRD hanya larang duplikat kode, bukan nama)",
     "Asumsi", "Belum dijalankan",
     "PRD tak larang nama duplikat -> Asumsi boleh"],

    # ---------------- NEGATIF ----------------
    ["TC-TAG-ADD-011", "Negatif",
     "Simpan dengan Instansi kosong",
     "Form terbuka",
     "1) Isi Nama, Kode, Tipe Anggota\n2) Biarkan Instansi kosong\n3) Klik \"Simpan\"",
     "Instansi=(kosong); Nama=QA<ts><seq>; Kode=QA<ts><seq>; Tipe=Semua",
     "Muncul pesan error required di Instansi; Simpan tidak bekerja; modal tetap terbuka; tanpa toast sukses",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-ADD-012", "Negatif",
     "Simpan dengan Nama Tag kosong",
     "Form terbuka",
     "1) Pilih Instansi, isi Kode, pilih Tipe\n2) Kosongkan Nama Tag\n3) Klik \"Simpan\"",
     "Instansi=SDIT; Nama=(kosong); Kode=QA<ts><seq>; Tipe=Semua",
     "Muncul pesan error required di Nama Tag; Simpan tidak bekerja; modal tetap terbuka; tanpa toast sukses",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-ADD-013", "Negatif",
     "Simpan dengan Kode Tag kosong",
     "Form terbuka",
     "1) Pilih Instansi, isi Nama, pilih Tipe\n2) Kosongkan Kode Tag\n3) Klik \"Simpan\"",
     "Instansi=SDIT; Nama=QA<ts><seq>; Kode=(kosong); Tipe=Semua",
     "Muncul pesan error required di Kode Tag; Simpan tidak bekerja; modal tetap terbuka; tanpa toast sukses",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-ADD-014", "Negatif",
     "Simpan dengan Tipe Anggota kosong",
     "Form terbuka",
     "1) Pilih Instansi, isi Nama & Kode\n2) Biarkan Tipe Anggota kosong\n3) Klik \"Simpan\"",
     "Instansi=SDIT; Nama=QA<ts><seq>; Kode=QA<ts><seq>; Tipe=(kosong)",
     "Muncul pesan error required di Tipe Anggota; Simpan tidak bekerja; modal tetap terbuka; tanpa toast sukses",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-ADD-015", "Negatif",
     "Simpan dengan semua field required kosong",
     "Form terbuka",
     "1) Klik \"Simpan\" tanpa isi apa pun",
     "Semua field=(kosong)",
     "Muncul pesan error required pada keempat field; Simpan tidak bekerja; tanpa toast sukses",
     "PRD", "Belum dijalankan", ""],

    ["TC-TAG-ADD-016", "Negatif",
     "Tambah duplikat: Kode Tag sama pada instansi yang sama",
     "Sudah ada tag dengan kode QA pada instansi target",
     "1) Tambah tag (seed)\n2) Tambah lagi kode & instansi yang sama (nama boleh beda)",
     "Instansi=SDIT; Kode=QA<ts><seq> (sama persis); Nama berbeda",
     "Ditolak: muncul pesan error duplikat kode (toast/inline) berisi 'Kode tag sudah digunakan di instansi ini'; tanpa toast sukses; modal tetap terbuka",
     "PRD", "FAIL",
     "BUG-020: BE benar tolak ({status:false, message:'Kode tag sudah digunakan di instansi ini'}) tapi FE SILENT - tidak tampilkan toast/inline. Pola sama BUG-007/008/010/019. Test pakai assertNotSilent() -> akan fail sampai bug fix"],

    # ---------------- EDGE ----------------
    ["TC-TAG-ADD-017", "Edge",
     "Input Nama / Kode whitespace-only (spasi saja)",
     "Form terbuka",
     "1) Pilih Instansi & Tipe\n2) Isi Nama & Kode = spasi saja\n3) Klik \"Simpan\"",
     "Instansi=SDIT; Nama=\"     \"; Kode=\"     \"; Tipe=Semua",
     "Spasi ter-trim -> dianggap kosong -> error required di Nama & Kode; tanpa toast sukses",
     "Asumsi", "Belum dijalankan",
     "PRD tak bahas trim -> Asumsi konsistensi modul lain (Kamar/Mapel)"],

    ["TC-TAG-ADD-018", "Edge",
     "Leading/trailing space pada Nama & Kode -> ter-trim saat simpan",
     "Form terbuka",
     "1) Isi Nama=\"  <core>  \"; Kode=\"  <core>  \"\n2) Simpan\n3) Cari <core> di list",
     "Instansi=SDIT; Nama=\"  QA<ts><seq>  \"; Kode=\"  QA<ts><seq>  \"; Tipe=Semua",
     "Sukses; data tersimpan tanpa spasi tepi (ditemukan saat cari <core>)",
     "Asumsi", "Belum dijalankan", ""],

    ["TC-TAG-ADD-019", "Edge",
     "Nama Tag >255 char (batas DB varchar 255)",
     "Form terbuka",
     "1) Pilih Instansi & Tipe\n2) Isi Kode valid\n3) Isi Nama 300 karakter\n4) Simpan\n5) Cek Network (XHR) response body",
     "Instansi=SDIT; Nama=QA<ts> + 'X' s/d 300 char; Kode=QA<ts><seq>; Tipe=Semua",
     "FE batasi maxlength / tampilkan pesan validasi ramah (cth 'maksimal 255 karakter'); data tidak tersimpan; tidak ada raw SQL error ke client",
     "Asumsi", "FAIL",
     "BUG-022: tidak ada maxlength, BE balas RAW SQL 'value too long ... varchar(255)' (SQLSTATE 22001), FE SILENT. Pola identik BUG-013 (Jurusan) & BUG-015 (Kamar). Test pakai assertNotSilent() -> akan fail sampai bug fix"],

    ["TC-TAG-ADD-020", "Edge",
     "Kode Tag >255 char (batas DB varchar 255)",
     "Form terbuka",
     "1) Pilih Instansi & Tipe\n2) Isi Nama valid\n3) Isi Kode 300 karakter\n4) Simpan\n5) Cek Network (XHR) response body",
     "Instansi=SDIT; Nama=QA<ts><seq>; Kode=QA<ts> + 'X' s/d 300 char; Tipe=Semua",
     "FE batasi maxlength / tampilkan pesan validasi ramah; data tidak tersimpan; tidak ada raw SQL error ke client",
     "Asumsi", "FAIL",
     "BUG-022 (sama akar BUG-013/015): field Kode juga tanpa maxlength, BE RAW SQL + FE silent. assertNotSilent() -> fail sampai bug fix"],

    ["TC-TAG-ADD-021", "Edge",
     "Nama & Kode 1 karakter (min boundary) -- tidak ditolak validasi panjang",
     "Form terbuka",
     "1) Pilih Instansi & Tipe\n2) Isi Nama=\"A\", Kode=\"A\"\n3) Simpan\n4) Cek field (data-invalid)",
     "Instansi=SDIT; Nama=\"A\"; Kode=\"A\"; Tipe=Semua",
     "1 char DITERIMA sbg panjang sah: field Nama & Kode tetap valid (data-invalid=false), tidak ada error required/min-length. PRD tak ada batas minimum.",
     "Asumsi", "Belum dijalankan",
     "Nama/Kode \"A\" statis (boundary 1 char tak bisa di-uniq-kan). TIDAK assert toast/persist krn app tolak duplikat -> false-fail run ke-2; jalur create sudah dicover TC-001"],

    # ---------------- DEFERRED (belum bisa di-run -- per instruksi user) ----------------
    ["TC-TAG-ADD-022", "Positif",
     "Tag yang dibuat muncul & dapat dipakai untuk Data Diri Siswa",
     "Tag QA sudah dibuat sukses (Tipe=Semua / Siswa)",
     "1) Buka form Data Diri Siswa\n2) Buka picker Tag\n3) Cek tag QA muncul sbg opsi",
     "Tag QA<ts><seq>",
     "Tag QA muncul & dapat dipilih pada form Data Diri Siswa",
     "PRD", "DITUNDA",
     "DEFERRED per instruksi user: fitur penggunaan tag di Data Siswa belum ada/belum bisa diakses. Aktifkan setelah modul tsb tersedia"],

    ["TC-TAG-ADD-023", "Positif",
     "Tag yang dibuat muncul & dapat dipakai untuk Pengaturan Presensi Kegiatan",
     "Tag QA sudah dibuat sukses",
     "1) Buka Pengaturan Presensi Kegiatan\n2) Buka picker Tag\n3) Cek tag QA muncul",
     "Tag QA<ts><seq>",
     "Tag QA muncul & dapat dipilih pada Pengaturan Presensi Kegiatan",
     "PRD", "DITUNDA",
     "DEFERRED per instruksi user: fitur Presensi Kegiatan belum ada"],

    ["TC-TAG-ADD-024", "Positif",
     "Tag yang dibuat muncul & dapat dipakai untuk membuat Tagihan",
     "Tag QA sudah dibuat sukses",
     "1) Buka form Buat Tagihan\n2) Buka picker Tag\n3) Cek tag QA muncul",
     "Tag QA<ts><seq>",
     "Tag QA muncul & dapat dipilih pada form Tagihan",
     "PRD", "DITUNDA",
     "DEFERRED per instruksi user: fitur Tagihan belum ada"],

    ["TC-TAG-ADD-025", "Positif",
     "Tag Tipe=Semua dapat dipakai untuk semua tipe anggota (siswa, guru, staff)",
     "Tag QA Tipe=Semua sudah dibuat",
     "1) Cek picker Tag di form Data Siswa\n2) Cek picker Tag di form Data Guru\n3) Cek picker Tag di form Data Staff",
     "Tag QA Tipe=Semua",
     "Tag muncul di ketiga picker (siswa, guru, staff)",
     "PRD", "DITUNDA",
     "DEFERRED: bergantung pada fitur Data Siswa/Guru/Staff yang belum tersedia"],

    ["TC-TAG-ADD-026", "Positif",
     "Tag Tipe=Siswa (STUDENT) hanya dapat dipakai untuk tipe Siswa",
     "Tag QA Tipe=Siswa sudah dibuat",
     "1) Cek picker Tag di form Data Siswa -> harus muncul\n2) Cek picker Tag di form Data Guru -> TIDAK muncul\n3) Cek picker Tag di form Data Staff -> TIDAK muncul",
     "Tag QA Tipe=Siswa (value=STUDENT)",
     "Tag hanya muncul di picker Siswa, tidak muncul di picker Guru/Staff",
     "PRD", "DITUNDA",
     "DEFERRED: bergantung pada fitur Data Siswa/Guru/Staff yang belum tersedia"],

    ["TC-TAG-ADD-027", "Positif",
     "Tag Tipe=Guru & Staff hanya dapat dipakai untuk tipe Guru & Staff",
     "Tag QA Tipe=Guru & Staff sudah dibuat",
     "1) Cek picker Tag di form Data Guru -> harus muncul\n2) Cek picker Tag di form Data Staff -> harus muncul\n3) Cek picker Tag di form Data Siswa -> TIDAK muncul",
     "Tag QA Tipe=Guru & Staff",
     "Tag muncul di picker Guru & Staff, tidak muncul di picker Siswa",
     "PRD", "DITUNDA",
     "DEFERRED: bergantung pada fitur Data Siswa/Guru/Staff yang belum tersedia"],

    # ---------------- BUG OBSERVATION (toast copy) ----------------
    ["TC-TAG-ADD-028", "Positif",
     "Toast sukses Tambah Tag: title benar, description SALAH modul (sebut 'Kelas')",
     "Modal Tambah Tag terbuka",
     "1) Isi seluruh field valid (Instansi+Nama+Kode+Tipe)\n2) Klik Simpan\n3) Amati toast (title & description)",
     "Instansi=SDIT; Nama=QA<ts><seq>; Kode=QA<ts><seq>; Tipe=Semua",
     "Toast Sonner muncul; title='Tag berhasil ditambahkan'; description menjelaskan konteks tag (BUKAN modul lain seperti 'Kelas')",
     "Asumsi", "FAIL",
     "BUG-021: title OK tapi description berbunyi 'Silahkan lanjutkan dengan menambahkan data Kelas pada menu'. Test assert title only & tandai description bermasalah -> akan lock issue copy"],
]

# Legend reference data
LEGEND = [
    ["Kolom Sumber", ""],
    ["PRD", "Skenario diturunkan langsung dari PRD Tambah Tag"],
    ["PRD-ambigu", "PRD menyinggung tapi tidak jelas/typo -> perlu konfirmasi"],
    ["Asumsi", "Tidak ada di PRD; asumsi berbasis konvensi/fixture/modul lain"],
    ["", ""],
    ["Kolom Kategori", ""],
    ["Happy", "Alur utama sukses sesuai desain"],
    ["Positif", "Variasi valid non-utama"],
    ["Negatif", "Input/kondisi yang harus ditolak sistem"],
    ["Edge", "Batas/kondisi tepi & verifikasi persistence"],
    ["", ""],
    ["Kolom Status", ""],
    ["Belum dijalankan", "Sudah di-generate, menunggu eksekusi Cypress"],
    ["DITUNDA", "Sengaja belum di-run; menunggu fitur lain (Data Siswa/Guru/Staff/Presensi/Tagihan) tersedia"],
    ["", ""],
    ["Catatan", "Tanpa kategori security/injection (sesuai kesepakatan)"],
    ["", "Bug ditandai inline: Status=FAIL + ref BUG-### di kolom Catatan"],
    ["", "Field PRD: Instansi* + Nama Tag* + Kode Tag* + Tipe Anggota* (semua required)"],
    ["", "Tipe Anggota = dropdown 3 opsi: Semua | Siswa | Guru & Staff"],
    ["", "PRD larang duplikat kode tag pada instansi yang sama"],
]

# ---- styling ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CAT_FILL = {
    "Happy":   PatternFill("solid", fgColor="E2EFDA"),
    "Positif": PatternFill("solid", fgColor="DDEBF7"),
    "Negatif": PatternFill("solid", fgColor="FCE4D6"),
    "Edge":    PatternFill("solid", fgColor="FFF2CC"),
}
SRC_FONT = {
    "PRD":        Font(color="375623"),
    "PRD-ambigu": Font(color="BF8F00", bold=True),
    "Asumsi":     Font(color="C00000", bold=True),
}
STATUS_FILL = {
    "DITUNDA": PatternFill("solid", fgColor="D9D9D9"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPWRAP = Alignment(vertical="top", wrap_text=True)
WIDTHS = [16, 10, 36, 30, 42, 28, 48, 12, 16, 44]

wb = Workbook()
ws = wb.active
ws.title = "Tambah Tag"

# Title row
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
t = ws.cell(row=1, column=1, value="Test Case - Tambah Tag (CARDS School v3)  |  Modul: Pengaturan Akademik > Tag  |  Sumber PRD: Tambah Tag")
t.font = Font(bold=True, size=12)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22

# Header row (row 2)
for c, name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=2, column=c, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = BORDER

# Data rows
for r, row in enumerate(ROWS, start=3):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = TOPWRAP
        cell.border = BORDER
    cat = row[1]
    if cat in CAT_FILL:
        ws.cell(row=r, column=2).fill = CAT_FILL[cat]
    src = row[7]
    if src in SRC_FONT:
        ws.cell(row=r, column=8).font = SRC_FONT[src]
    status = row[8]
    if status in STATUS_FILL:
        ws.cell(row=r, column=9).fill = STATUS_FILL[status]

# Column widths
for c, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + len(ROWS)}"

# Legend sheet
ls = wb.create_sheet("Legenda")
ls.column_dimensions["A"].width = 18
ls.column_dimensions["B"].width = 72
for r, (a, b) in enumerate(LEGEND, 1):
    ca = ls.cell(row=r, column=1, value=a)
    cb = ls.cell(row=r, column=2, value=b)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    if a in ("Kolom Sumber", "Kolom Kategori", "Kolom Status", "Catatan"):
        ca.font = Font(bold=True)

out = "docs/test-cases/TC_Tag_Tambah.xlsx"
wb.save(out)
print("saved:", out, "| rows:", len(ROWS))
