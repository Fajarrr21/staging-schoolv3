# -*- coding: utf-8 -*-
"""Generate Test Case sheet for Hapus Kamar (per CLAUDE.md columns)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = ["ID", "Kategori", "Skenario", "Pre-condition", "Langkah",
           "Test Data", "Expected", "Sumber", "Status", "Catatan"]

# PRD Hapus Kamar:
#   1) Pada list kamar, klik Aksi (ikon Hapus) -> popup delete confirmation muncul
#   2) Klik "Hapus" di popup -> kamar terhapus, sistem menampilkan pesan success
#   3) Klik "Batal" di popup -> popup tutup, kembali ke list kamar (tidak terhapus)
# SKIP (per instruksi user): "Kamar yang berhasil dihapus tidak dapat digunakan kembali
#   untuk data diri siswa serta pengaturan presensi kegiatan" -> fitur downstream
#   (data siswa & presensi) belum bisa diuji dari modul Kamar.

ROWS = [
    # ---------------- HAPPY ----------------
    ["TC-KMR-DEL-001", "Happy",
     "Klik ikon Hapus pada baris kamar → popup konfirmasi Hapus muncul",
     "Login admin; ada minimal 1 kamar di list",
     "1) Seed 1 kamar (Tambah)\n2) Cari kamar tsb\n3) Klik ikon Hapus (trash) pada baris\n4) Amati popup",
     "Kamar=QA<ts><seq>",
     "Popup konfirmasi terbuka dengan judul \"Hapus Kamar\"; ada tombol \"Hapus\" dan \"Batal\"; nama kamar yang akan dihapus terlihat di popup",
     "PRD", "Belum dijalankan",
     "Teks tombol & isi pesan konfirmasi sebagian = Asumsi (PRD hanya sebut popup delete confirmation)"],

    ["TC-KMR-DEL-002", "Happy",
     "Konfirmasi Hapus → toast sukses & kamar hilang dari list",
     "Login admin; popup konfirmasi Hapus terbuka untuk kamar tsb",
     "1) Seed 1 kamar\n2) Buka popup Hapus untuk kamar tsb\n3) Klik tombol \"Hapus\"",
     "Kamar=QA<ts><seq>",
     "Toast sukses \"Kamar berhasil dihapus\"; popup tertutup; kamar TIDAK lagi muncul di list saat di-search",
     "PRD", "Belum dijalankan",
     "Teks toast = Asumsi (PRD hanya sebut \"pesan success\"). Verifikasi via search → assertRowNotExists"],

    ["TC-KMR-DEL-003", "Happy",
     "Hapus persist setelah reload halaman",
     "Login admin; kamar baru saja dihapus",
     "1) Seed 1 kamar\n2) Hapus kamar (sukses)\n3) Reload halaman (kamar.visit())\n4) Cari nama kamar yang dihapus",
     "Kamar=QA<ts><seq>",
     "Setelah reload, kamar tetap tidak ada di list (search → empty / row tidak ditemukan). Backend benar-benar menghapus, bukan hanya optimistic UI",
     "PRD", "Belum dijalankan",
     "Verifikasi backend persist via reload+search → assertRowNotExists"],

    # ---------------- POSITIF ----------------
    ["TC-KMR-DEL-004", "Positif",
     "Klik Batal di popup → popup tutup, data tetap ada di list",
     "Login admin; popup konfirmasi Hapus terbuka",
     "1) Seed 1 kamar\n2) Buka popup Hapus\n3) Klik tombol \"Batal\"\n4) Cari kamar tsb",
     "Kamar=QA<ts><seq>",
     "Popup tertutup; tidak ada toast sukses/error; kamar TETAP ada di list (data tidak dihapus)",
     "PRD", "Belum dijalankan",
     "PRD eksplisit: 'Klik Batal → kembali pada halaman list kamar (tidak hapus)'"],

    ["TC-KMR-DEL-005", "Positif",
     "Hapus kamar yang punya PIC terisi → tetap berhasil",
     "Login admin; ada kamar dengan PIC terisi",
     "1) Seed kamar + PIC (pilih guru/staff pertama)\n2) Hapus kamar via popup\n3) Verifikasi list",
     "Kamar dengan PIC=<guru/staff>",
     "Toast sukses; kamar hilang dari list. PIC (guru/staff) tetap ada di sistem (tidak ikut terhapus)",
     "PRD", "Belum dijalankan",
     "Memastikan delete tidak gagal karena ada relasi PIC. PRD tidak melarang"],

    ["TC-KMR-DEL-006", "Positif",
     "Hapus kamar yang punya Lokasi terisi → tetap berhasil",
     "Login admin; ada kamar dengan Lokasi terisi",
     "1) Seed kamar + Lokasi\n2) Hapus kamar via popup\n3) Verifikasi list",
     "Kamar dengan Lokasi=LOK<ts><seq>",
     "Toast sukses; kamar hilang dari list",
     "PRD", "Belum dijalankan",
     "Field optional terisi tidak boleh menghalangi delete"],

    ["TC-KMR-DEL-007", "Positif",
     "Hapus kamar berstatus Tidak Aktif → tetap berhasil",
     "Login admin; ada kamar berstatus Tidak Aktif",
     "1) Seed kamar (default Aktif)\n2) Edit → set Status=Tidak Aktif\n3) Hapus kamar via popup\n4) Verifikasi list",
     "Kamar Status=Tidak Aktif",
     "Toast sukses; kamar hilang dari list. Status Tidak Aktif TIDAK menghalangi penghapusan",
     "PRD", "Belum dijalankan",
     "PRD tidak melarang hapus kamar Tidak Aktif. Reuse modul Edit untuk setup"],

    ["TC-KMR-DEL-008", "Positif",
     "Hapus beberapa kamar berturut-turut → masing-masing popup independen",
     "Login admin; ada minimal 2 kamar di list",
     "1) Seed 2 kamar (A & B) di instansi yang sama\n2) Hapus A via popup (sukses)\n3) Tanpa reload, hapus B via popup\n4) Verifikasi list",
     "Kamar A & Kamar B",
     "Kedua kamar terhapus berurutan; masing-masing memunculkan popup konfirmasi tersendiri & toast sukses; tidak ada state popup yang nyangkut",
     "PRD", "Belum dijalankan",
     "Mengantisipasi bug state popup (modal-stack atau focus-trap nyangkut antar delete)"],

    # ---------------- NEGATIF ----------------
    ["TC-KMR-DEL-009", "Negatif",
     "Hapus kamar yang masih digunakan di data diri siswa / pengaturan presensi → sistem harus mencegah / beri peringatan",
     "Login admin; ada kamar yang sudah dipakai oleh data siswa atau dijadikan kamar presensi kegiatan",
     "1) Setup: pastikan ada kamar X yang sudah dipakai di data siswa / presensi\n2) Buka popup Hapus untuk kamar X\n3) Konfirmasi Hapus",
     "Kamar X (sudah ter-link ke siswa / presensi)",
     "Sistem mencegah / memberi peringatan bahwa kamar masih digunakan; ATAU hapus berhasil tapi referensi di data siswa/presensi otomatis dilepas (perilaku resmi belum didefinisikan)",
     "PRD", "DITUNDA",
     "PRD: \"Kamar yang berhasil dihapus tidak dapat digunakan kembali untuk data diri siswa serta pengaturan presensi kegiatan\". Fitur downstream (modul data siswa & presensi) belum tersedia untuk diuji dari sini -> di-pending sampai modul terkait bisa diakses"],

    # ---------------- EDGE ----------------
    ["TC-KMR-DEL-010", "Edge",
     "Tekan ESC saat popup konfirmasi terbuka → popup tutup, tidak terhapus",
     "Login admin; popup konfirmasi Hapus terbuka",
     "1) Seed 1 kamar\n2) Buka popup Hapus\n3) Tekan tombol ESC pada keyboard\n4) Cari kamar tsb",
     "Kamar=QA<ts><seq>",
     "Popup tertutup tanpa konfirmasi; tidak ada toast; kamar TETAP ada di list",
     "PRD", "Belum dijalankan",
     "Dikonfirmasi oleh user: ESC menutup popup (konvensi Radix UI). Setara fungsional dengan tombol Batal"],

    ["TC-KMR-DEL-011", "Edge",
     "Setelah hapus, nama kamar bisa dipakai ulang untuk Tambah Kamar baru (bukan soft-delete)",
     "Login admin; kamar baru saja dihapus",
     "1) Seed kamar nama N di Instansi A\n2) Hapus kamar tsb\n3) Tambah kamar baru dengan nama N di Instansi A yang sama\n4) Verifikasi success",
     "Nama N reused di instansi yang sama",
     "Tambah sukses (toast \"Kamar berhasil ditambahkan\"); kamar baru muncul di list. Nama yang sudah dihapus TIDAK menabrak unique constraint -> konfirmasi delete benar-benar menghilangkan record",
     "Asumsi", "Belum dijalankan",
     "Memastikan delete = hard delete (bukan soft-delete yang masih meng-occupy unique constraint). Kalau tabrakan -> kandidat bug baru"],
]

LEGEND = [
    ["Kolom Sumber", ""],
    ["PRD", "Skenario diturunkan langsung dari PRD Hapus Kamar"],
    ["PRD-ambigu", "PRD menyinggung tapi tidak jelas/typo -> perlu konfirmasi"],
    ["Asumsi", "Tidak ada di PRD; asumsi berbasis konvensi/fixture/modul lain"],
    ["", ""],
    ["Kolom Kategori", ""],
    ["Happy", "Alur utama sukses sesuai desain"],
    ["Positif", "Variasi valid non-utama"],
    ["Negatif", "Input/kondisi yang harus ditolak / mengembalikan error"],
    ["Edge", "Batas/kondisi tepi"],
    ["", ""],
    ["Kolom Status", ""],
    ["Belum dijalankan", "Sudah di-generate, menunggu eksekusi Cypress"],
    ["DITUNDA", "Skenario valid namun fitur/dependensi belum bisa diuji dari modul ini"],
    ["FAIL", "Gagal saat run; lihat ref BUG-### di kolom Catatan"],
    ["", ""],
    ["Skenario SKIP (tidak ada di sheet ini)", ""],
    ["SKIP", "Validasi bahwa kamar terhapus tidak bisa digunakan di data siswa / presensi -> butuh modul downstream yang belum tersedia"],
    ["", ""],
    ["Catatan", "Tanpa kategori security/injection (sesuai kesepakatan)"],
    ["", "Bug ditandai inline: Status=FAIL + ref BUG-### di kolom Catatan"],
    ["", "Sub-order modul: Tambah -> List -> Edit -> Hapus -> cleanup utility (zzz_cleanup_kamar)"],
    ["", "Cleanup utility akan reuse delete flow dari TC-002 untuk hapus residu QA<ts><seq>"],
]

# ---- styling ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CAT_FILL = {
    "Happy":   PatternFill("solid", fgColor="E2EFDA"),
    "Positif": PatternFill("solid", fgColor="DDEBF7"),
    "Negatif": PatternFill("solid", fgColor="FCE4D6"),
    "Edge":    PatternFill("solid", fgColor="FFF2CC"),
}
SRC_FONT = {
    "PRD":        Font(color="375623"),
    "PRD-ambigu": Font(color="BF8F00", bold=True),
    "Asumsi":     Font(color="C00000", bold=True),
}
STATUS_FILL = {
    "FAIL":    PatternFill("solid", fgColor="FF0000"),
    "DITUNDA": PatternFill("solid", fgColor="D9D9D9"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPWRAP = Alignment(vertical="top", wrap_text=True)
WIDTHS = [16, 10, 40, 34, 46, 30, 52, 12, 16, 48]

wb = Workbook()
ws = wb.active
ws.title = "Hapus Kamar"

# Title row
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
t = ws.cell(row=1, column=1,
            value="Test Case - Hapus Kamar (CARDS School v3)  |  Modul: Pengaturan Akademik > Kamar  |  Sumber PRD: Hapus Kamar")
t.font = Font(bold=True, size=12)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22

# Header row
for c, name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=2, column=c, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = BORDER

# Data rows
for r, row in enumerate(ROWS, start=3):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = TOPWRAP
        cell.border = BORDER
    cat = row[1]
    if cat in CAT_FILL:
        ws.cell(row=r, column=2).fill = CAT_FILL[cat]
    src = row[7]
    if src in SRC_FONT:
        ws.cell(row=r, column=8).font = SRC_FONT[src]
    status = row[8]
    if status in STATUS_FILL:
        ws.cell(row=r, column=9).fill = STATUS_FILL[status]

# Column widths
for c, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + len(ROWS)}"

# Legend sheet
ls = wb.create_sheet("Legenda")
ls.column_dimensions["A"].width = 26
ls.column_dimensions["B"].width = 72
for r, (a, b) in enumerate(LEGEND, 1):
    ca = ls.cell(row=r, column=1, value=a)
    cb = ls.cell(row=r, column=2, value=b)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    if a in ("Kolom Sumber", "Kolom Kategori", "Kolom Status",
             "Skenario SKIP (tidak ada di sheet ini)", "Catatan"):
        ca.font = Font(bold=True)

out = "docs/test-cases/TC_Kamar_Hapus.xlsx"
wb.save(out)
print("saved:", out, "| rows:", len(ROWS))
