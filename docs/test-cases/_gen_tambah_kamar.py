# -*- coding: utf-8 -*-
"""Generate Test Case sheet for Tambah Kamar (per CLAUDE.md columns)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = ["ID", "Kategori", "Skenario", "Pre-condition", "Langkah",
           "Test Data", "Expected", "Sumber", "Status", "Catatan"]

# Form Tambah Kamar (PRD):
#   a. Instansi*      (required, select)
#   b. Nama Kamar*    (required, text)
#   c. PIC            (optional, select -> dropdown guru & staff pada instansi terpilih)
#   d. Lokasi         (optional, text)
ROWS = [
    # ---------------- HAPPY ----------------
    ["TC-KMR-ADD-001", "Happy",
     "Tambah kamar valid (hanya field required: Instansi + Nama Kamar)",
     "Login admin; halaman List Kamar terbuka",
     "1) Klik \"Tambah Kamar\"\n2) Pilih Instansi\n3) Isi Nama Kamar\n4) Biarkan PIC & Lokasi kosong (optional)\n5) Klik \"Simpan\"",
     "Instansi=SDIT; Nama=QA<6digit-ts><seq>",
     "Toast sukses \"Kamar berhasil ditambahkan\"; modal tertutup; baris kamar baru muncul di list",
     "PRD", "Belum dijalankan",
     "PIC & Lokasi optional -> form valid tanpa keduanya. Teks toast = Asumsi (PRD hanya \"pesan succes\")"],

    ["TC-KMR-ADD-002", "Happy",
     "Tambah kamar lengkap (Instansi + Nama + PIC + Lokasi)",
     "Login admin; ada guru/staff pada instansi terpilih untuk opsi PIC",
     "1) Klik \"Tambah Kamar\"\n2) Pilih Instansi\n3) Isi Nama Kamar\n4) Pilih PIC dari dropdown\n5) Isi Lokasi\n6) Klik \"Simpan\"",
     "Instansi=SDIT; Nama=QA<ts><seq>; PIC=<guru/staff pertama>; Lokasi=\"Lantai 2\"",
     "Toast sukses; modal tertutup; baris baru muncul dengan PIC & Lokasi sesuai input",
     "PRD", "Belum dijalankan",
     "Verifikasi PIC & Lokasi tersimpan -> cek kolom/persist via reload"],

    # ---------------- POSITIF ----------------
    ["TC-KMR-ADD-003", "Positif",
     "Form Tambah Kamar terbuka dengan field kosong",
     "Login admin; List Kamar terbuka",
     "1) Klik \"Tambah Kamar\"\n2) Amati kondisi awal form",
     "-",
     "Modal \"Tambah Kamar\" tampil; Instansi placeholder, Nama kosong, PIC placeholder, Lokasi kosong; tombol Simpan & Batal ada",
     "PRD", "Belum dijalankan", ""],

    ["TC-KMR-ADD-004", "Positif",
     "Batal menutup form tanpa menambah data",
     "Modal Tambah Kamar terbuka",
     "1) Klik \"Tambah Kamar\"\n2) Klik \"Batal\"",
     "-",
     "Modal tertutup; kembali ke List Kamar; tidak ada data baru",
     "PRD", "Belum dijalankan", ""],

    ["TC-KMR-ADD-005", "Positif",
     "Isi form valid lalu Batal -> data tidak tersimpan",
     "Modal Tambah Kamar terbuka",
     "1) Pilih Instansi + isi Nama Kamar\n2) Klik \"Batal\"\n3) Cari nama tsb di list",
     "Instansi=SDIT; Nama=QA<ts><seq>",
     "Modal tertutup; nama kamar TIDAK muncul di list (tidak tersimpan)",
     "PRD", "Belum dijalankan", ""],

    ["TC-KMR-ADD-006", "Positif",
     "Tambah beberapa kamar berbeda dalam 1 instansi",
     "Login admin",
     "1) Tambah kamar A (Instansi=SDIT)\n2) Tambah kamar B (Instansi=SDIT)",
     "Nama A & B = QA<ts><seq> (berbeda)",
     "Kedua kamar sukses ditambahkan; keduanya muncul di list",
     "PRD", "Belum dijalankan", ""],

    ["TC-KMR-ADD-007", "Positif",
     "PIC menampilkan dropdown guru & staff pada instansi yang dipilih",
     "Ada minimal 1 guru/staff terdaftar pada instansi terpilih",
     "1) Buka form\n2) Pilih Instansi\n3) Buka dropdown PIC\n4) Amati daftar opsi",
     "Instansi=SDIT",
     "Dropdown PIC berisi guru & staff milik instansi terpilih (bukan instansi lain)",
     "PRD", "Belum dijalankan",
     "PRD: \"memilih PIC -> dropdown guru & staff pada instansi yang dipilih\". Perlu cek HTML opsi PIC"],

    ["TC-KMR-ADD-008", "Positif",
     "Ganti Instansi -> opsi PIC ikut berubah sesuai instansi baru",
     "Form terbuka; ada guru/staff berbeda antar instansi",
     "1) Pilih Instansi A -> lihat opsi PIC\n2) Ganti ke Instansi B -> lihat opsi PIC",
     "Instansi A=SDIT; Instansi B=Sekolah Alam",
     "Daftar opsi PIC ter-refresh mengikuti instansi yang dipilih; PIC yang sudah terpilih dari instansi lama di-reset",
     "Asumsi", "Belum dijalankan",
     "PRD tak eksplisit soal reset PIC saat ganti instansi -> Asumsi, validasi ke behavior/HTML"],

    ["TC-KMR-ADD-009", "Positif",
     "Nama kamar sama pada instansi berbeda -> sukses",
     "Login admin",
     "1) Tambah Nama X @ Instansi SDIT\n2) Tambah Nama X @ Instansi Sekolah Alam",
     "Nama X = QA<ts><seq> (sama)",
     "Keduanya sukses (uniqueness diasumsikan per-instansi, bukan global)",
     "Asumsi", "Belum dijalankan",
     "PRD hanya larang duplikat \"instansi yang sama\" -> beda instansi diasumsikan boleh"],

    # ---------------- NEGATIF ----------------
    ["TC-KMR-ADD-010", "Negatif",
     "Simpan dengan Nama Kamar kosong",
     "Form terbuka",
     "1) Pilih Instansi\n2) Kosongkan Nama Kamar\n3) Klik \"Simpan\"",
     "Instansi=SDIT; Nama=(kosong)",
     "Muncul pesan error required di Nama Kamar; Simpan tidak bekerja; modal tetap terbuka; tanpa toast sukses",
     "PRD", "Belum dijalankan", ""],

    ["TC-KMR-ADD-011", "Negatif",
     "Simpan dengan Instansi kosong",
     "Form terbuka",
     "1) Isi Nama Kamar\n2) Biarkan Instansi kosong\n3) Klik \"Simpan\"",
     "Instansi=(kosong); Nama=QA<ts><seq>",
     "Muncul pesan error required di Instansi; Simpan tidak bekerja; modal tetap terbuka; tanpa toast sukses",
     "PRD", "Belum dijalankan", ""],

    ["TC-KMR-ADD-012", "Negatif",
     "Simpan dengan kedua field required kosong",
     "Form terbuka",
     "1) Klik \"Simpan\" tanpa isi apa pun",
     "Instansi=(kosong); Nama=(kosong)",
     "Muncul pesan error required di Instansi & Nama Kamar; Simpan tidak bekerja; tanpa toast sukses",
     "PRD", "Belum dijalankan", ""],

    ["TC-KMR-ADD-013", "Negatif",
     "Tambah duplikat: Nama kamar sama pada instansi yang sama",
     "Sudah ada kamar QA pada instansi target",
     "1) Tambah kamar (seed)\n2) Tambah lagi nama & instansi yang sama",
     "Instansi=SDIT; Nama=QA<ts><seq> (sama persis)",
     "Ditolak: muncul pesan error duplikat; tanpa toast sukses; modal tetap terbuka",
     "PRD", "Belum dijalankan",
     "PRD eksplisit larang nama sama pada instansi sama. Jika malah sukses & nutup modal = FE-silent -> log BUG"],

    # ---------------- EDGE ----------------
    ["TC-KMR-ADD-014", "Edge",
     "Input Nama Kamar whitespace-only (spasi saja)",
     "Form terbuka",
     "1) Pilih Instansi\n2) Isi Nama = beberapa spasi\n3) Klik \"Simpan\"",
     "Instansi=SDIT; Nama=\"     \"",
     "Spasi ter-trim -> dianggap kosong -> error required; tanpa toast sukses",
     "Asumsi", "Belum dijalankan",
     "PRD tak bahas trim -> Asumsi konsistensi modul lain"],

    ["TC-KMR-ADD-015", "Edge",
     "Leading/trailing space pada Nama -> ter-trim saat simpan",
     "Form terbuka",
     "1) Isi Nama = \"  <core>  \"\n2) Simpan\n3) Cari <core> di list",
     "Instansi=SDIT; Nama=\"  QA<ts><seq>  \"",
     "Sukses; nama tersimpan tanpa spasi tepi (ditemukan saat cari <core>)",
     "Asumsi", "Belum dijalankan", ""],

    ["TC-KMR-ADD-016", "Edge",
     "Nama kamar >255 char ditolak rapi (batas DB varchar 255)",
     "Form terbuka",
     "1) Pilih Instansi\n2) Isi Nama 300 karakter\n3) Simpan\n4) Cek Network (XHR) response body",
     "Instansi=SDIT; Nama=QA<ts> + 'X' s/d 300 char",
     "FE batasi maxlength / tampilkan pesan validasi ramah (cth 'maksimal 255 karakter'); data tidak tersimpan; tidak ada raw SQL error ke client",
     "Asumsi", "FAIL",
     "BUG-015: tanpa maxlength, API balas RAW SQL 'value too long ... varchar(255) (SQLSTATE 22001)' & FE SILENT (no toast sukses/error). Pola sama BUG-013 (Jurusan)"],

    ["TC-KMR-ADD-017", "Edge",
     "Nama kamar 1 karakter (min boundary) — tidak ditolak validasi panjang",
     "Form terbuka",
     "1) Pilih Instansi\n2) Isi Nama 1 huruf\n3) Simpan\n4) Cek field Kamar (data-invalid)",
     "Instansi=SDIT; Nama=\"A\"",
     "1 char DITERIMA sbg panjang sah: field Kamar tetap valid (data-invalid=false), tidak ada error required/min-length. Tidak ada batas minimum di PRD.",
     "Asumsi", "Belum dijalankan",
     "Nama \"A\" statis (boundary 1 char tak bisa di-uniq-kan). TIDAK assert toast/persist krn app tolak duplikat -> false-fail run ke-2; jalur create sudah dicover TC-001 (nama unik)"],

    ["TC-KMR-ADD-018", "Edge",
     "Lokasi diisi panjang / karakter beragam (optional field)",
     "Form terbuka",
     "1) Isi Instansi + Nama valid\n2) Isi Lokasi teks panjang\n3) Simpan",
     "Lokasi=\"Gedung A Lantai 3 Sayap Kanan dekat tangga darurat\"",
     "Sukses; Lokasi tersimpan apa adanya",
     "Asumsi", "Belum dijalankan",
     "Lokasi optional & bebas teks -> validasi ke HTML apakah ada batasan"],

    # ---------------- DEFERRED (belum bisa di-run) ----------------
    ["TC-KMR-ADD-019", "Positif",
     "Kamar yang dibuat dapat dipakai di Data Diri Siswa & Pengaturan Presensi Kegiatan",
     "Kamar QA sudah dibuat sukses",
     "1) Buka form Data Diri Siswa -> cek kamar muncul sebagai opsi\n2) Buka Pengaturan Presensi Kegiatan -> cek kamar muncul",
     "Kamar QA<ts><seq>",
     "Kamar muncul & dapat dipilih pada Data Diri Siswa dan Pengaturan Presensi Kegiatan",
     "PRD", "DITUNDA",
     "DEFERRED: dibuat sesuai PRD tapi BELUM dijalankan -- fitur Data Siswa & Presensi belum dikerjakan (masih fase CRUD Kamar). Aktifkan setelah modul tsb tersedia"],
]

# Legend reference data
LEGEND = [
    ["Kolom Sumber", ""],
    ["PRD", "Skenario diturunkan langsung dari PRD Tambah Kamar"],
    ["PRD-ambigu", "PRD menyinggung tapi tidak jelas/typo -> perlu konfirmasi"],
    ["Asumsi", "Tidak ada di PRD; asumsi berbasis konvensi/fixture/modul lain"],
    ["", ""],
    ["Kolom Kategori", ""],
    ["Happy", "Alur utama sukses sesuai desain"],
    ["Positif", "Variasi valid non-utama"],
    ["Negatif", "Input/kondisi yang harus ditolak sistem"],
    ["Edge", "Batas/kondisi tepi & verifikasi persistence"],
    ["", ""],
    ["Kolom Status", ""],
    ["Belum dijalankan", "Sudah di-generate, menunggu eksekusi Cypress"],
    ["DITUNDA", "Sengaja belum di-run; menunggu fitur lain (Data Siswa/Presensi) tersedia"],
    ["", ""],
    ["Catatan", "Tanpa kategori security/injection (sesuai kesepakatan)"],
    ["", "Bug ditandai inline: Status=FAIL + ref BUG-### di kolom Catatan"],
    ["", "Field PRD: Instansi* + Nama Kamar* (required), PIC + Lokasi (optional)"],
    ["", "PIC = dropdown guru & staff pada instansi yang dipilih"],
]

# ---- styling ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CAT_FILL = {
    "Happy":   PatternFill("solid", fgColor="E2EFDA"),
    "Positif": PatternFill("solid", fgColor="DDEBF7"),
    "Negatif": PatternFill("solid", fgColor="FCE4D6"),
    "Edge":    PatternFill("solid", fgColor="FFF2CC"),
}
SRC_FONT = {
    "PRD":        Font(color="375623"),
    "PRD-ambigu": Font(color="BF8F00", bold=True),
    "Asumsi":     Font(color="C00000", bold=True),
}
STATUS_FILL = {
    "DITUNDA": PatternFill("solid", fgColor="D9D9D9"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPWRAP = Alignment(vertical="top", wrap_text=True)
WIDTHS = [16, 10, 36, 30, 42, 28, 48, 12, 16, 44]

wb = Workbook()
ws = wb.active
ws.title = "Tambah Kamar"

# Title row
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
t = ws.cell(row=1, column=1, value="Test Case - Tambah Kamar (CARDS School v3)  |  Modul: Pengaturan Akademik > Kamar  |  Sumber PRD: Tambah Kamar")
t.font = Font(bold=True, size=12)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22

# Header row (row 2)
for c, name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=2, column=c, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = BORDER

# Data rows
for r, row in enumerate(ROWS, start=3):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = TOPWRAP
        cell.border = BORDER
    cat = row[1]
    if cat in CAT_FILL:
        ws.cell(row=r, column=2).fill = CAT_FILL[cat]
    src = row[7]
    if src in SRC_FONT:
        ws.cell(row=r, column=8).font = SRC_FONT[src]
    status = row[8]
    if status in STATUS_FILL:
        ws.cell(row=r, column=9).fill = STATUS_FILL[status]

# Column widths
for c, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + len(ROWS)}"

# Legend sheet
ls = wb.create_sheet("Legenda")
ls.column_dimensions["A"].width = 18
ls.column_dimensions["B"].width = 72
for r, (a, b) in enumerate(LEGEND, 1):
    ca = ls.cell(row=r, column=1, value=a)
    cb = ls.cell(row=r, column=2, value=b)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    if a in ("Kolom Sumber", "Kolom Kategori", "Kolom Status", "Catatan"):
        ca.font = Font(bold=True)

out = "docs/test-cases/TC_Kamar_Tambah.xlsx"
wb.save(out)
print("saved:", out, "| rows:", len(ROWS))
