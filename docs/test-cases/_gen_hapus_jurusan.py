# -*- coding: utf-8 -*-
"""Generate Test Case sheet for Hapus Jurusan (per CLAUDE.md columns)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = ["ID", "Kategori", "Skenario", "Pre-condition", "Langkah",
           "Test Data", "Expected", "Sumber", "Status", "Catatan"]

ROWS = [
    # ID, Kategori, Skenario, Pre-condition, Langkah, Test Data, Expected, Sumber, Status, Catatan
    # CATATAN STATUS: Hapus Jurusan saat ini PLAIN DELETE - belum terkoneksi ke menu apa pun
    # (termasuk PPDB), tanpa kriteria/validasi khusus. Case PPDB DITUNDA sampai fitur tersambung.
    ["TC-JRS-HPS-001", "Happy",
     "Buka popup konfirmasi hapus via Aksi -> Hapus",
     "Login admin; minimal 1 jurusan tampil di list",
     "1) Buka list Jurusan\n2) Pada baris target klik ikon Hapus (svg.lucide-trash) -> .closest('button')",
     "Jurusan existing apa saja",
     "Popup konfirmasi tampil dengan judul \"Hapus Jurusan\"; ada tombol \"Hapus\" dan \"Batal\"",
     "PRD", "Belum dijalankan", ""],

    ["TC-JRS-HPS-002", "Happy",
     "Hapus jurusan -> sukses (hapus biasa, tanpa kriteria khusus)",
     "Ada jurusan QA baru di list",
     "1) Aksi -> Hapus pada baris target\n2) Klik \"Hapus\" di popup",
     "QA<6digit-ts><seq> (mis. QA48211203)",
     "Toast sukses \"Jurusan berhasil dihapus\"; baris hilang dari list",
     "PRD", "Belum dijalankan",
     "Saat ini langsung terhapus (belum terkoneksi PPDB/menu lain). Teks toast dari fixture deleteSuccess; PRD hanya menyebut \"pesan sukses\" -> teks = Asumsi"],

    ["TC-JRS-HPS-003", "Happy",
     "Klik Batal di popup -> tidak terhapus",
     "Popup konfirmasi terbuka untuk baris target",
     "1) Aksi -> Hapus\n2) Klik \"Batal\"",
     "Jurusan existing",
     "Popup tertutup; kembali ke list Jurusan; baris target TETAP ada",
     "PRD-ambigu", "Belum dijalankan",
     "PRD menulis \"kembali pada halaman list kelas\" - diasumsikan typo, maksudnya list Jurusan"],

    ["TC-JRS-HPS-004", "Positif",
     "Hapus jurusan pada instansi sekunder (Sekolah Alam)",
     "Ada jurusan QA pada instansi sekunder",
     "1) Filter/identifikasi baris instansi sekunder\n2) Aksi -> Hapus -> Hapus",
     "QA<ts><seq> @ Sekolah Alam",
     "Sukses terhapus; hanya baris instansi tsb hilang; instansi lain tidak terpengaruh",
     "Asumsi", "Belum dijalankan",
     "PRD tidak membahas multi-instansi pada hapus -> Asumsi konsistensi dengan modul lain"],

    ["TC-JRS-HPS-005", "Positif",
     "Hapus jurusan setelah Cari by nama",
     "Ada jurusan QA yang bisa dicari",
     "1) Ketik nama di Cari -> cy.wait('@alias') (intercept list)\n2) Aksi -> Hapus -> Hapus",
     "QA<ts><seq>",
     "Baris hasil pencarian terhapus; toast sukses",
     "Asumsi", "Belum dijalankan",
     "Gunakan cy.intercept + wait('@alias'), bukan cy.wait(angka)"],

    ["TC-JRS-HPS-006", "Edge",
     "Tutup popup tanpa konfirmasi (Escape / klik overlay)",
     "Popup konfirmasi terbuka",
     "1) Aksi -> Hapus\n2) Tekan Escape ATAU klik area di luar dialog",
     "Jurusan existing",
     "Popup tertutup; tidak ada perubahan data; baris TETAP ada",
     "Asumsi", "Belum dijalankan",
     "PRD hanya menyebut tombol Batal; perilaku Escape/overlay = Asumsi, perlu validasi ke HTML/behavior"],

    ["TC-JRS-HPS-007", "Edge",
     "Hapus baris terakhir pada hasil filter/list -> empty state",
     "Hasil filter/list menyisakan tepat 1 baris",
     "1) Hapus baris terakhir -> Hapus",
     "Jurusan QA tunggal pada hasil filter",
     "Setelah terhapus muncul empty state \"Data Jurusan tidak ditemukan\"",
     "Asumsi", "Belum dijalankan",
     "emptyStateTitle diambil dari fixture"],

    ["TC-JRS-HPS-008", "Edge",
     "Hapus sukses lalu reload -> data tetap hilang (persistence)",
     "TC-JRS-HPS-002 sukses",
     "1) Setelah hapus sukses, reload (visit)\n2) Cari nama jurusan yang dihapus",
     "QA<ts><seq>",
     "Jurusan tidak muncul setelah reload (persist di backend, bukan sekadar update optimistik)",
     "Asumsi", "Belum dijalankan",
     "Sesuai aturan persistence CLAUDE.md (verifikasi via reload)"],

    ["TC-JRS-HPS-009", "Edge",
     "Popup konfirmasi menampilkan nama jurusan yang akan dihapus",
     "Popup konfirmasi terbuka untuk baris target",
     "1) Aksi -> Hapus\n2) Baca isi teks popup",
     "Jurusan existing",
     "Teks popup memuat nama jurusan target",
     "PRD-ambigu", "Belum dijalankan",
     "Perlu cek HTML asli apakah dialog konfirmasi menampilkan nama jurusan"],
]

# Legend reference data
LEGEND = [
    ["Kolom Sumber", ""],
    ["PRD", "Skenario diturunkan langsung dari PRD Hapus Jurusan"],
    ["PRD-ambigu", "PRD menyinggung tapi tidak jelas/typo -> perlu konfirmasi"],
    ["Asumsi", "Tidak ada di PRD; asumsi berbasis konvensi/fixture/modul lain"],
    ["", ""],
    ["Kolom Kategori", ""],
    ["Happy", "Alur utama sukses sesuai desain"],
    ["Positif", "Variasi valid non-utama"],
    ["Negatif", "Input/kondisi yang harus ditolak sistem"],
    ["Edge", "Batas/kondisi tepi & verifikasi persistence"],
    ["", ""],
    ["Catatan", "Tanpa kategori security/injection (sesuai kesepakatan)"],
    ["", "Bug ditandai inline: Status=FAIL + ref BUG-### di kolom Catatan"],
    ["", "Hapus Jurusan saat ini PLAIN DELETE - belum terkoneksi PPDB/menu lain."],
    ["", "Case validasi PPDB (ditolak + tombol Jadwal PPDB) DITUNDA sampai fitur tersambung."],
]

# ---- styling ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CAT_FILL = {
    "Happy":   PatternFill("solid", fgColor="E2EFDA"),
    "Positif": PatternFill("solid", fgColor="DDEBF7"),
    "Negatif": PatternFill("solid", fgColor="FCE4D6"),
    "Edge":    PatternFill("solid", fgColor="FFF2CC"),
}
SRC_FONT = {
    "PRD":        Font(color="375623"),
    "PRD-ambigu": Font(color="BF8F00", bold=True),
    "Asumsi":     Font(color="C00000", bold=True),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPWRAP = Alignment(vertical="top", wrap_text=True)
WIDTHS = [16, 10, 34, 30, 40, 22, 46, 12, 16, 40]

wb = Workbook()
ws = wb.active
ws.title = "Hapus Jurusan"

# Title row
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
t = ws.cell(row=1, column=1, value="Test Case - Hapus Jurusan (CARDS School v3)  |  Modul: Pengaturan Jurusan  |  Sumber PRD: Hapus Jurusan")
t.font = Font(bold=True, size=12)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22

# Header row (row 2)
for c, name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=2, column=c, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = BORDER

# Data rows
for r, row in enumerate(ROWS, start=3):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = TOPWRAP
        cell.border = BORDER
    cat = row[1]
    if cat in CAT_FILL:
        ws.cell(row=r, column=2).fill = CAT_FILL[cat]
    src = row[7]
    if src in SRC_FONT:
        ws.cell(row=r, column=8).font = SRC_FONT[src]

# Column widths
for c, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + len(ROWS)}"

# Legend sheet
ls = wb.create_sheet("Legenda")
ls.column_dimensions["A"].width = 18
ls.column_dimensions["B"].width = 70
for r, (a, b) in enumerate(LEGEND, 1):
    ca = ls.cell(row=r, column=1, value=a)
    cb = ls.cell(row=r, column=2, value=b)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    if a in ("Kolom Sumber", "Kolom Kategori", "Catatan"):
        ca.font = Font(bold=True)

out = "docs/test-cases/TC_Jurusan_Hapus.xlsx"
wb.save(out)
print("saved:", out, "| rows:", len(ROWS))
