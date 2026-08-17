# -*- coding: utf-8 -*-
"""Generate Test Case sheet for Hapus Kalender Akademik (per CLAUDE.md columns).

Catatan PRD:
  PRD Kalender Pendidikan TIDAK secara eksplisit menyebut flow Hapus. Skenario TC
  di sheet ini disusun berdasarkan pola modul Hapus standar di project (Hapus Kelas,
  Hapus Tag, Hapus Kamar): dialog konfirmasi (Radix AlertDialog) dgn tombol
  Hapus/Batal, toast sukses setelah confirm, row hilang dari list.

  Semua Sumber = Asumsi (pola shared UI). Tanpa TC kategori Negatif karena
  tidak ada validasi khusus yg diketahui (mis. "kalender aktif tidak bisa dihapus"
  — tidak ada dependency flag di modul ini).

Data strategi (rerun-safe):
  - Target delete pakai instansi yg BELUM punya kalender di staging (unusedForEdit,
    unusedAlt) + di-Tambah-dulu di beforeEach / dalam TC -> delete -> verify.
  - Setelah delete sukses, state kembali bersih otomatis (idempotent).
  - Instansi seeded permanen (primary/existing/tertiary) TIDAK boleh dihapus di
    TC ini (dipakai di modul lain).

Modal Hapus (proyeksi, sebelum element analysis):
  - Radix AlertDialog: title "Hapus Kalender Akademik", description menyebut
    nama instansi target, tombol Hapus (destructive) + Batal + close X.
  - Confirm: toast sukses "Kalender Akademik berhasil dihapus", row hilang dari list.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = ["ID", "Kategori", "Skenario", "Pre-condition", "Langkah",
           "Test Data", "Expected", "Sumber", "Status", "Catatan"]

ROWS = [
    # ---------------- HAPPY ----------------
    ["TC-KLD-HPS-001", "Happy",
     "Klik icon trash di row -> dialog konfirmasi -> klik Hapus -> toast sukses + row hilang dari list",
     "Login admin; row target (unusedForEdit) sudah di-Tambah lebih dulu di beforeEach; row muncul di list",
     "1) Buka List Kalender Akademik\n2) Pastikan row target ada di list\n3) Klik icon trash (svg.lucide-trash) di row target\n4) Dialog konfirmasi muncul\n5) Klik tombol 'Hapus' (destructive action)\n6) Amati toast & list",
     "Row: instansi=unusedForEdit (Academy Cazh); Awal Pekan=Minggu; Nama Pekan=Ahad",
     "Toast sukses (Sonner data-type=success) muncul dgn text 'Kalender Akademik berhasil dihapus'; dialog konfirmasi close; row target HILANG dari list; total row berkurang 1",
     "Asumsi", "Belum dijalankan",
     "Pola standar Hapus. beforeEach seed via kalender.addKalender ke unusedForEdit"],

    ["TC-KLD-HPS-002", "Happy",
     "Persistence: setelah delete sukses + reload halaman -> row tetap tidak muncul (backend persist)",
     "Login admin; row target di-Tambah di beforeEach; delete sukses di step Langkah",
     "1) Setup: Tambah row target\n2) Delete row target (via icon trash + confirm)\n3) Toast sukses + row hilang di UI\n4) Reload halaman (F5)\n5) Amati list setelah reload",
     "Row: instansi=unusedForEdit",
     "Setelah reload, row target TETAP tidak muncul di list (BE benar-benar persist delete, bukan optimistic UI); tidak ada state ghost",
     "Asumsi", "Belum dijalankan",
     "assertPersisted-negative pattern"],

    # ---------------- POSITIF ----------------
    ["TC-KLD-HPS-003", "Positif",
     "Dialog konfirmasi tampil: title 'Hapus Kalender Akademik', deskripsi menyebut nama instansi target",
     "Login admin; row target di-Tambah di beforeEach; row visible di list",
     "1) Klik icon trash di row target\n2) Dialog konfirmasi muncul\n3) Amati title, deskripsi, tombol",
     "Row: instansi=unusedForEdit",
     "Dialog terbuka (Radix AlertDialog): title text 'Hapus Kalender Akademik'; deskripsi/konten mengandung nama instansi target (mis. 'Academy Cazh') utk clarity ke user; tombol Hapus (destructive) & Batal visible",
     "Asumsi", "Belum dijalankan",
     "Element analysis WAJIB konfirmasi: role=alertdialog, description mention instansi, tombol Hapus label + styling destructive"],

    ["TC-KLD-HPS-004", "Positif",
     "Klik Batal di dialog konfirmasi -> dialog close, row TIDAK terhapus, tidak ada toast",
     "Login admin; row target di list; dialog konfirmasi terbuka",
     "1) Klik icon trash di row target\n2) Dialog konfirmasi muncul\n3) Klik tombol 'Batal'\n4) Amati dialog & list & toast",
     "Row: instansi=unusedForEdit; action: cancel confirm",
     "Dialog close; row target TETAP ada di list; toast sukses TIDAK muncul; total row tidak berubah",
     "Asumsi", "Belum dijalankan",
     "Cancel confirm pattern - safety net"],

    ["TC-KLD-HPS-005", "Positif",
     "Klik icon X di header dialog konfirmasi -> dialog close, row TIDAK terhapus",
     "Login admin; row target di list; dialog konfirmasi terbuka",
     "1) Klik icon trash di row target\n2) Dialog konfirmasi muncul\n3) Klik icon X (svg.lucide-x) di pojok kanan atas dialog\n4) Amati dialog & list",
     "Row: instansi=unusedForEdit; action: cancel via X",
     "Dialog close; row target TETAP ada di list; toast sukses TIDAK muncul; sama behavior dgn tombol Batal",
     "Asumsi", "Belum dijalankan",
     "Close X pattern - konsisten dgn modul lain. Element analysis konfirmasi apakah AlertDialog ada X (Radix AlertDialog default TIDAK punya X, biasanya cuma dua tombol)"],

    ["TC-KLD-HPS-006", "Positif",
     "Tekan ESC key saat dialog konfirmasi terbuka -> dialog close, row TIDAK terhapus",
     "Login admin; row target di list; dialog konfirmasi terbuka",
     "1) Klik icon trash di row target\n2) Dialog konfirmasi muncul\n3) Tekan tombol ESC di keyboard\n4) Amati dialog & list",
     "Row: instansi=unusedForEdit; action: cancel via ESC",
     "Dialog close; row target TETAP ada di list; toast TIDAK muncul (keyboard shortcut standar Radix Dialog)",
     "Asumsi", "Belum dijalankan",
     "Radix Dialog default handle ESC. Kalau AlertDialog dev-nya override (disable ESC), TC ini FAIL -> flag di element analysis"],

    ["TC-KLD-HPS-007", "Positif",
     "Klik overlay (area gelap luar dialog) -> dialog close, row TIDAK terhapus",
     "Login admin; row target di list; dialog konfirmasi terbuka",
     "1) Klik icon trash di row target\n2) Dialog konfirmasi muncul\n3) Klik area overlay (di luar dialog card)\n4) Amati dialog & list",
     "Row: instansi=unusedForEdit; action: cancel via overlay",
     "Dialog close; row target TETAP ada di list; toast TIDAK muncul",
     "Asumsi", "Belum dijalankan",
     "Radix AlertDialog by default TIDAK close via overlay click (biar user tidak accidental cancel destructive). Kalau app override -> TC OK; kalau default -> TC bisa DITUNDA. Element analysis konfirmasi."],

    # ---------------- EDGE ----------------
    ["TC-KLD-HPS-008", "Edge",
     "Delete row lalu Tambah ulang instansi yg sama -> Tambah sukses (uniqueness constraint lepas setelah delete)",
     "Login admin; row target di list (sudah di-Tambah); belum ada TC lain manipulate row target",
     "1) Delete row target (icon trash -> confirm)\n2) Toast sukses + row hilang\n3) Klik 'Tambah Kalender Akademik'\n4) Isi instansi = SAMA dgn yg baru di-delete\n5) Isi Awal Pekan + Nama Pekan\n6) Klik Simpan\n7) Amati toast & list",
     "Row lama: instansi=unusedForEdit (deleted); Tambah baru: instansi=unusedForEdit (sama)",
     "Tambah baru sukses (toast 'Kalender Akademik berhasil ditambahkan'); row unusedForEdit muncul kembali di list; TIDAK muncul toast duplicate 'Pengaturan kalender untuk office ini sudah ada'",
     "Asumsi", "Belum dijalankan",
     "Verify uniqueness constraint di-lepas properly setelah delete (bukan soft-delete leaked ke unique check)"],
]

LEGEND = [
    ["Kolom Sumber", ""],
    ["PRD", "Skenario diturunkan langsung dari PRD Kalender Pendidikan"],
    ["PRD-ambigu", "PRD menyinggung tapi tidak detail; turunan logis"],
    ["Asumsi", "PRD tidak menyebut sama sekali; behavior turunan dari pola modul Hapus lain"],
    ["", ""],
    ["Kolom Kategori", ""],
    ["Happy", "Alur utama sukses (delete + persist)"],
    ["Positif", "Variasi cancel valid (Batal, X, ESC, overlay) + verify dialog content"],
    ["Negatif", "TIDAK ADA - tidak ada validasi khusus utk delete di modul ini"],
    ["Edge", "Delete lalu re-add (uniqueness constraint verify)"],
    ["", ""],
    ["Kolom Status", ""],
    ["Belum dijalankan", "Sudah di-generate, menunggu eksekusi Cypress"],
    ["DITUNDA", "Defer sampai element analysis konfirmasi UI element"],
    ["FAIL", "Diketahui FAIL karena bug app (lihat kolom Catatan untuk BUG-### ref)"],
    ["", ""],
    ["Catatan strategi", ""],
    ["Data rerun-safe", "Setiap TC seed via kalender.addKalender di beforeEach ke instansi unusedForEdit (Academy Cazh) yg belum punya kalender permanen. Delete lalu clean state otomatis."],
    ["Instansi seeded permanen", "Primary/existing/tertiary DIPAKAI di modul lain (Tambah/List/Edit) - TIDAK boleh dihapus di sheet ini."],
    ["Element analysis pending", "Radix AlertDialog vs Dialog: title/description selector, tombol Hapus destructive class, close X presence, ESC/overlay behavior - semua di-konfirmasi setelah ACC TC sheet."],
    ["Pola referensi", "Hapus Kelas / Hapus Kamar / Hapus Tag - semua modul Pengaturan Akademik lain punya pattern serupa"],
    ["Tanpa kategori security", "Injection & security cases dilarang per CLAUDE.md kesepakatan"],
]

# ---- styling (copy pattern dari _gen_edit_kalender.py) ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CAT_FILL = {
    "Happy":   PatternFill("solid", fgColor="E2EFDA"),
    "Positif": PatternFill("solid", fgColor="DDEBF7"),
    "Negatif": PatternFill("solid", fgColor="FCE4D6"),
    "Edge":    PatternFill("solid", fgColor="FFF2CC"),
}
SRC_FONT = {
    "PRD":        Font(color="375623"),
    "PRD-ambigu": Font(color="BF8F00", bold=True),
    "Asumsi":     Font(color="C00000", bold=True),
}
STATUS_FILL = {
    "DITUNDA": PatternFill("solid", fgColor="D9D9D9"),
    "FAIL":    PatternFill("solid", fgColor="F4B084"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPWRAP = Alignment(vertical="top", wrap_text=True)
WIDTHS = [16, 12, 42, 36, 48, 28, 56, 14, 16, 50]

wb = Workbook()
ws = wb.active
ws.title = "Hapus Kalender"

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
t = ws.cell(row=1, column=1, value="Test Case - Hapus Kalender Akademik (CARDS School v3)  |  Modul: Pengaturan Akademik > Kalender Akademik  |  Sumber: Asumsi (pola standar modul Hapus, tidak ada PRD spesifik)")
t.font = Font(bold=True, size=12)
t.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22

for c, name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=2, column=c, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = BORDER

for r, row in enumerate(ROWS, start=3):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = TOPWRAP
        cell.border = BORDER
    cat = row[1]
    if cat in CAT_FILL:
        ws.cell(row=r, column=2).fill = CAT_FILL[cat]
    src = row[7]
    if src in SRC_FONT:
        ws.cell(row=r, column=8).font = SRC_FONT[src]
    status = row[8]
    if status in STATUS_FILL:
        ws.cell(row=r, column=9).fill = STATUS_FILL[status]

for c, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + len(ROWS)}"

ls = wb.create_sheet("Legenda")
ls.column_dimensions["A"].width = 30
ls.column_dimensions["B"].width = 78
for r, (a, b) in enumerate(LEGEND, 1):
    ca = ls.cell(row=r, column=1, value=a)
    cb = ls.cell(row=r, column=2, value=b)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    if a in ("Kolom Sumber", "Kolom Kategori", "Kolom Status", "Catatan strategi"):
        ca.font = Font(bold=True)

out = "docs/test-cases/TC_Kalender_Hapus.xlsx"
wb.save(out)
print("saved:", out, "| rows:", len(ROWS))
